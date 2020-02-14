from variable.models import Variable, Unidad
# librerias para manejar los archivos EXCEL
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class Titulos:
    # Estilo de Fuentes
    fonts = {
        'font_bold_11': Font(bold=True),
        'font_bold_10': Font(bold=True, size=10),
        'font_10': Font(size=10),
        'font_11': Font(size=11),
        'font_8': Font(size=8)
    }
    # Alineacion de las celdas
    alignments = {
        'center': Alignment(horizontal="center", vertical="center"),
        'right': Alignment(horizontal="right", vertical="center"),
        'left': Alignment(horizontal="left", vertical="center"),
        'wrap': Alignment(horizontal="center", vertical="center", wrap_text=True)
    }
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    # bordes
    borders = {
        'border_thin': Border(top=thin, left=thin, bottom=thin, right=thin),
        'border_medium': Border(top=medium, left=medium, right=medium, bottom=medium)
    }
    # relleno
    colors = {
        'orange': PatternFill("solid", fgColor="FDE9D9"),
        'light_salmon': PatternFill("solid", fgColor="FFA07A")
    }

    @staticmethod
    def titulo_grafico(variable):
        # returns var_nombre given var_id
        consulta = list(Variable.objects.filter(var_id=variable))

        return consulta[0]

    @staticmethod
    def titulo_unidad(variable):
        var = Variable.objects.get(var_id=variable)
        return var.uni_id.uni_sigla

    def set_encabezado_excel(self, ws, estacion, periodo):
        fila = 1
        col = 1
        col_fin = 11

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col+2)
        title = ws.cell(row=fila, column=col)
        title.value = estacion.est_codigo
        self.set_style(cell=title, font='font_bold_11', alignment='center',
                       border='border_thin', fill='light_salmon')

        ws.merge_cells(start_row=fila, start_column=col+3, end_row=fila, end_column=col+7)
        title = ws.cell(row=fila, column=col+3)
        title.value = estacion.est_nombre
        self.set_style(cell=title, font='font_bold_11', alignment='center',
                       border='border_thin', fill='light_salmon')

        ws.merge_cells(start_row=fila, start_column=col+8, end_row=fila, end_column=col_fin)
        title = ws.cell(row=fila, column=col+8)
        title.value = periodo
        self.set_style(cell=title, font='font_bold_11', alignment='center',
                       border='border_thin', fill='light_salmon')

        fila += 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col_fin)
        title = ws.cell(row=fila, column=col)
        title.value = 'Coordenadas Geográfica UTM (DATUM WGS 84)'
        self.set_style(cell=title, font='font_bold_11', alignment='center',
                       border='border_thin')

        fila += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Latitud"
        self.set_style(cell=cell, font='font_bold_10', alignment='center',
                       border='border_thin')

        ws.merge_cells(start_row=fila, start_column=col + 1, end_row=fila, end_column=col + 3)
        title = ws.cell(row=fila, column=col + 1)
        title.value = estacion.est_latitud
        self.set_style(cell=title, font='font_10', alignment='center',
                       border='border_thin')

        cell = ws.cell(row=fila, column=col+4)
        cell.value = "Longitud"
        self.set_style(cell=cell, font='font_bold_10', alignment='center',
                       border='border_thin')

        ws.merge_cells(start_row=fila, start_column=col + 5, end_row=fila, end_column=col + 7)
        title = ws.cell(row=fila, column=col + 5)
        title.value = estacion.est_longitud
        self.set_style(cell=title, font='font_10', alignment='center',
                       border='border_thin')

        cell = ws.cell(row=fila, column=col + 8)
        cell.value = "Altura"
        self.set_style(cell=cell, font='font_bold_10', alignment='center',
                       border='border_thin')

        ws.merge_cells(start_row=fila, start_column=col + 9, end_row=fila, end_column=col_fin)
        title = ws.cell(row=fila, column=col + 9)
        title.value = estacion.est_altura
        self.set_style(cell=title, font='font_10', alignment='center',
                       border='border_thin')

    @staticmethod
    def get_mes_anio(int_mes):
        # meses = ['Enero', 'Febrero','Marzo', 'Abril', 'Mayo', 'Junio', 'Julio','Agosto',
        # 'Septiembre','Octubre','Noviembre','Diciembre']
        meses = ['ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEP', 'OCT',
                 'NOV', 'DIC']

        return meses[int_mes-1]

    def set_style(self, cell, font=None, alignment=None, border=None, fill = None):
        if font is not None:
            cell.font = self.fonts[font]
        if alignment is not None:
            cell.alignment = self.alignments[alignment]
        if border is not None:
            cell.border = self.borders[border]
        if fill is not None:
            cell.fill = self.colors[fill]
