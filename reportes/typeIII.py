# -*- coding: utf-8 -*-
import plotly.offline as opy
import plotly.graph_objs as go
from reportes.titulos import Titulos
from anuarios.models import TemperaturaAire
from django.db.models import Avg
import calendar

from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)


# clase para anuario de la variable TAI
class TypeIII(Titulos):

    '''@staticmethod
    def consulta(estacion, periodo):
        # annotate agrupa los valores en base a un campo y a una operacion
        informacion = list(TemperaturaAire.objects.filter(est_id=estacion).filter(tai_periodo=periodo))
        return informacion'''

    def matriz(self, estacion, variable, periodo):
        print("matriz", variable)
        datos = self.consulta(estacion, variable, periodo)
        return datos

    def grafico(self, estacion, variable, periodo):
        print("grafico", variable)
        datos = self.consulta(estacion, variable, periodo)
        if datos:
            meses = []
            max_simple = []
            min_simple = []
            avg_simple = []
            for item in datos:
                meses.append(str(calendar.month_abbr[item.tai_mes]))
                max_simple.append(item.tai_maximo_abs)
                min_simple.append(item.tai_minimo_abs)
                avg_simple.append(item.tai_promedio)

            trace0 = go.Scatter(
                x=meses,
                y=max_simple,
                name='Máx. Abs.',
                line=dict(
                    color=('rgb(22, 96, 167)'),
                    width=4)
            )
            trace1 = go.Scatter(
                x=meses,
                y=min_simple,
                name='Mín. Abs',
                line=dict(
                    color=('rgb(205, 12, 24)'),
                    width=4, )
            )
            trace2 = go.Scatter(
                x=meses,
                y=avg_simple,
                name='Media',
                line=dict(
                    color=('rgb(50, 205, 50)'),
                    width=4, )
            )
            data = [trace0, trace1, trace2]
            layout = go.Layout(
                title=str(variable.var_nombre) + str(" (") + str(variable.uni_id.uni_sigla) + str(") ")
            )
            figure = go.Figure(data=data, layout=layout)
            div = opy.plot(figure, auto_open=False, output_type='div')
            return div
        return False

    def tabla_excel(self, ws, estacion, variable, periodo):
        fila = 5
        col_fin = 11
        col = 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col_fin)
        subtitle = ws.cell(row=fila, column=col)
        subtitle.value = "Temperatura del Aire - Valores medios mensuales," \
                         " absolutos maximos y mimimos, medios maximos y minimos "
        self.set_style(cell=subtitle, font='font_bold_10', alignment='center',
                       border='border_thin', fill='light_salmon')

        fila += 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila + 2, end_column=col)
        cell = ws.cell(row=fila, column=col)
        cell.value = "MES"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col += 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col + 7)
        cell = ws.cell(row=fila, column=col)
        cell_final = ws.cell(row=fila, column=col + 7)
        cell.value = "Temperatura del Aire (ºC)"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')
        self.set_style(cell=cell_final, font='font_10', alignment='center',
                       border='border_thin')

        fila += 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col + 3)
        cell = ws.cell(row=fila, column=col)
        cell.value = "Absoluta"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col = 6

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col + 3)
        cell = ws.cell(row=fila, column=col)
        cell_final = ws.cell(row=fila, column=col+3)
        cell.value = "Media"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')
        self.set_style(cell=cell_final, font='font_10', alignment='center',
                       border='border_thin')

        fila += 1
        col = 2

        cell = ws.cell(row=fila, column=col)
        cell.value = "Max"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Día"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Min"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Día"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Max"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Min"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Mensual"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Histórica"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        matriz = self.matriz(estacion, variable, periodo)
        fila += 1
        col = 1

        media_historica = self.datos_historicos(estacion, variable, periodo)

        for item in matriz:
            cell = ws.cell(row=fila, column=col)
            cell.value = self.get_mes_anio(item.tai_mes)
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')
            cell = ws.cell(row=fila, column=col+1)
            cell.value = item.tai_maximo_abs
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')

            cell = ws.cell(row=fila, column=col+2)
            cell.value = item.tai_maximo_dia
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')

            cell = ws.cell(row=fila, column=col+3)
            cell.value = item.tai_minimo_abs
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')

            cell = ws.cell(row=fila, column=col+4)
            cell.value = item.tai_minimo_dia
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')

            cell = ws.cell(row=fila, column=col + 5)
            cell.value = item.tai_maximo
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')

            cell = ws.cell(row=fila, column=col + 6)
            cell.value = item.tai_minimo
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')

            cell = ws.cell(row=fila, column=col + 7)
            cell.value = item.tai_promedio
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')

            cell = ws.cell(row=fila, column=col + 8)
            cell.value = round(media_historica[item.tai_mes-1], 2)
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')

            fila += 1

    @staticmethod
    def grafico_excel(ws, variable, periodo):
        c1 = ScatterChart()
        c1.title = str(variable.var_nombre) + str(" (") + \
            str(variable.uni_id.uni_sigla) + str(") ") + str(periodo)
        # c1.style = 10
        # c1.y_axis.title = str(variable.uni_id.uni_sigla)
        c1.x_axis.title = 'Meses'

        xvalues = Reference(ws, min_col=1, min_row=9, max_row=20)

        for i in range(6, 10):
            values = Reference(ws, min_col=i, min_row=8, max_row=20)
            series = Series(values, xvalues, title_from_data=True)
            c1.series.append(series)

        serie_max = c1.series[0]
        serie_max.marker.symbol = "diamond"
        serie_max.marker.graphicalProperties.solidFill = "1660a7"
        serie_max.marker.graphicalProperties.line.solidFill = "1660a7"
        serie_max.graphicalProperties.line.solidFill = "1660a7"

        serie_min = c1.series[1]
        serie_min.marker.symbol = "triangle"
        serie_min.marker.graphicalProperties.solidFill = "cd0c18"
        serie_min.marker.graphicalProperties.line.solidFill = "cd0c18"
        serie_min.graphicalProperties.line.solidFill = "cd0c18"

        serie_pro = c1.series[2]
        serie_pro.marker.symbol = "square"
        serie_pro.marker.graphicalProperties.solidFill = "32cd32"
        serie_pro.marker.graphicalProperties.line.solidFill = "32cd32"
        serie_pro.graphicalProperties.line.solidFill = "32cd32"

        serie_pro = c1.series[3]
        serie_pro.marker.symbol = "x"
        serie_pro.marker.graphicalProperties.solidFill = "7d60a0"
        serie_pro.marker.graphicalProperties.line.solidFill = "7d60a0"
        serie_pro.graphicalProperties.line.solidFill = "7d60a0"

        cats = Reference(ws, min_col=1, min_row=8, max_row=19)
        c1.set_categories(cats)
        c1.legend.position = "b"
        ws.add_chart(c1, "A21")
