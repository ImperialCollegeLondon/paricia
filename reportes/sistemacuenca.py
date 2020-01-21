# -*- coding: utf-8 -*-

# Modelos
from reportes.models import (
    ConsultaGenericaFechaHoraGrafico,
    ConsultaGenericaFechaHora,
    ConsultaGenericaFecha)

# Funcionalidades Comunes
from reportes.functions import (
    get_layout_grafico,
    get_elemento_data_json,
    get_data_graph
)

# Funciones de django
from datetime import datetime
from django.http import HttpResponse
import csv


# librerias para manejar los archivos EXCEL
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.drawing.image import Image

from openpyxl.chart import (
    LineChart,
    Reference,
)

from sedc.settings import BASE_DIR

__frecuencia__dict = {
    "subhorario-crudo": "Dato crudo",
    "subhorario-validado": "Dato validado",
    "horario": "Horario",
    "diario": "Diario",
    "mensual": "Mensual",
}


# consultar datos crudos, validados, horarios, diarios y mensuales
def get_datos_graficar(estacion, variable, inicio, fin, frecuencia, profundidad):
    titulo = estacion.est_codigo + " " + estacion.est_nombre + " - " + variable.var_nombre
    titulo_grafico = titulo + "<br>(" + __frecuencia__dict[frecuencia] + ")"

    if frecuencia == "subhorario-crudo":
        datos = consulta_crudos_graficar(estacion.est_id, variable, inicio, fin, profundidad)
    elif frecuencia == "subhorario-validado":
        datos = consulta_validados_graficar(estacion.est_id, variable, inicio, fin, profundidad)
    elif frecuencia == "horario":
        datos = consulta_horario_graficar(estacion.est_id, variable, inicio, fin, profundidad)
    elif frecuencia == "diario":
        datos = consulta_diario_graficar(estacion.est_id, variable, inicio, fin, profundidad)
    else:
        datos = consulta_mensual_graficar(estacion.est_id, variable, inicio, fin, profundidad)

    titulo_yaxis = variable.var_nombre + " (" + variable.uni_id.uni_sigla + ")"
    layout = get_layout_grafico(titulo_grafico, titulo_yaxis, inicio, fin)

    if variable.var_id != 1:

        data_valor = get_elemento_data_json('scatter', datos['fecha'], datos['valor'], 'Valor', '#1660A7')
    else:
        data_valor = get_elemento_data_json('bar', datos['fecha'], datos['valor'], 'Promedio', '#1660A7')

    data = get_data_graph(data_valor)


    grafico = dict(
        data=data,
        layout=layout,
    )

    return grafico


def get_datos_exportar(estacion, variable, inicio, fin, frecuencia, profundidad):
    nombre_archivo = estacion.est_codigo + "-" + estacion.est_nombre + " " + variable.var_nombre
    if profundidad:
        nombre_archivo = nombre_archivo + " a " + str(profundidad/100.0) + "[m]"
    nombre_archivo = nombre_archivo + " " + __frecuencia__dict[frecuencia]
    nombre_archivo = nombre_archivo.replace(" ", "_")

    if frecuencia == "subhorario-crudo":
        datos = consulta_crudos(estacion.est_id, variable, inicio, fin, profundidad)
        formato_fecha = 'yyyy/mm/dd hh:mm:ss'
    elif frecuencia == "subhorario-validado":
        datos = consulta_validados(estacion.est_id, variable, inicio, fin, profundidad)
        formato_fecha = 'yyyy/mm/dd hh:mm:ss'
    elif frecuencia == "horario":
        datos = consulta_horario(estacion.est_id, variable, inicio, fin, profundidad)
        formato_fecha = 'yyyy/mm/dd hh:mm:ss'
    elif frecuencia == "diario":
        datos = consulta_diario(estacion.est_id, variable, inicio, fin, profundidad)
        formato_fecha = 'yyyy/mm/dd'
    else:
        datos = consulta_mensual(estacion.est_id, variable, inicio, fin, profundidad)
        formato_fecha = 'yyyy/mm'

    return datos, nombre_archivo, formato_fecha


def export_csv(estacion, variable, inicio, fin, frecuencia, profundidad):
    datos, nombre_archivo, formato_fecha = get_datos_exportar(estacion, variable, inicio, fin, frecuencia, profundidad)
    for fila in datos:
        hay_datos = True
        break
    if not hay_datos:
        return None

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="' + nombre_archivo + '.csv"'
    writer = csv.writer(response)
    writer.writerow(['Fecha', 'Valor (' + variable.uni_id.uni_sigla + ')'])
    formato_fecha = formato_fecha.replace("yyyy", "%Y").replace("/mm", "/%m").replace("dd", "%d")
    formato_fecha = formato_fecha.replace("hh", "%H").replace(":mm", ":%M").replace("ss", "%S")
    for fila in datos:
        writer.writerow([fila.fecha.strftime(formato_fecha), fila.valor])
    return response


def export_excel(estacion, variable, inicio, fin, frecuencia, profundidad):
    datos, nombre_archivo, formato_fecha = get_datos_exportar(estacion, variable, inicio, fin, frecuencia, profundidad)
    for fila in datos:
        hay_datos = True
        break
    if not hay_datos:
        return None

    # ruta de la imagen
    ruta_fonag = str(BASE_DIR) + '/media/logo_fonag.jpg'
    ruta_epmaps = str(BASE_DIR) + '/media/logo_EPMAPS.jpg'
    img_fonag = Image(ruta_fonag)
    img_epmaps = Image(ruta_epmaps)
    # estilo de negrita
    font_bold = Font(bold=True)
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    ws.add_image(img_fonag, 'A1')
    ws.add_image(img_epmaps, 'G1')

    ws['B4'] = 'Reporte de Datos Hidrometerológicos'
    ws['B4'].font = font_bold

    ws.merge_cells('B4:F4')
    ws['A7'] = 'Estación'
    ws['A7'].font = font_bold
    ws['B7'] = estacion.est_codigo
    ws['C7'] = estacion.est_nombre
    ws.merge_cells('C4:E4')
    ws['F7'] = 'Variable'
    ws['F7'].font = font_bold
    ws['G7'] = variable.var_nombre
    ws['B9'] = 'Coordenadas Geográfica UTM (DATUM WGS 84)'
    ws['B9'].font = font_bold
    ws.merge_cells('B6:G6')
    ws['A10'] = 'Latitud'
    ws['A10'].font = font_bold
    ws['B10'] = estacion.est_latitud
    ws['F10'] = 'Longitud'
    ws['F10'].font = font_bold
    ws['G10'] = estacion.est_longitud

    # Creamos los encabezados desde la celda B9 hasta la E9
    ws['A12'] = 'Fecha'
    ws['B12'] = 'Valor'
    cont = 13
    for fila in datos:
        ws.cell(row=cont, column=1).value = fila.fecha
        ws.cell(row=cont, column=2).value = fila.valor
        cont = cont + 1

    # grafico
    chart = LineChart()
    chart.title = variable.var_nombre
    chart.style = 12
    chart.x_axis.title = 'Tiempo'
    chart.x_axis.number_format = 'dd-mm-yyyy'
    chart.x_axis.majorTimeUnit = 'days'
    chart.y_axis.title = variable.var_nombre

    final = 13 + len(datos)
    data = Reference(ws, min_col=2, min_row=12, max_col=3, max_row=final - 1)
    chart.add_data(data, titles_from_data=True)
    s0 = chart.series[0]
    s0.graphicalProperties.line.solidFill = "32CD32"
    s0.graphicalProperties.line.width = 10

    dates = Reference(ws, min_col=1, min_row=13, max_row=final)
    chart.set_categories(dates)

    ws.add_chart(chart, "H9")

    nombre_archivo = nombre_archivo + str('.xlsx')

    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


def consulta_crudos(estacion_id, variable, inicio, fin, profundidad):
    modelo = str(variable.var_modelo)
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    sql = """
    SELECT m.fecha, m.valor
    FROM medicion_var101 m
    WHERE m.estacion = %s
    """
    if profundidad:
        sql = sql + '        AND m.profundidad = ' + str(profundidad)
    if inicio:
        sql = sql + '        AND m.fecha>=\'' + str(inicio) + '\''
    if fin:
        sql = sql + '        AND m.fecha<=\'' + str(fin) + '\''
    sql = sql + """
    ORDER BY m.fecha ASC, m.id ASC;
    """
    sql = sql.replace("var101", modelo)
    consulta = ConsultaGenericaFechaHora.objects.raw(sql, [estacion_id, ])
    return consulta


# consultar datos crudos para graficar
def consulta_crudos_graficar(estacion_id, variable, inicio, fin, profundidad):
    modelo = str(variable.var_modelo)
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    sql = """
    with
    estacion AS (SELECT * FROM estacion_estacion est WHERE est.est_id = %s),
    variable AS (SELECT * FROM variable_variable var WHERE var.var_id = %s),
    medicion AS (
        SELECT m.id, m.fecha, m.valor
        FROM medicion_var101 m
        WHERE m.estacion = (SELECT est_id FROM estacion)
    """
    if profundidad:
        sql = sql + '        AND profundidad = ' + str(profundidad)
    if inicio:
        sql = sql + '        AND fecha>=\'' + str(inicio) + '\''
    if fin:
        sql = sql + '        AND fecha<=\'' + str(fin) + '\''
    sql = sql + """
        ORDER BY fecha ASC
    ),
    fechas AS (
        SELECT 
            row_number() OVER (ORDER BY mm.fecha ASC, mm.id ASC) as fila,
            mm.id, mm.fecha, mm.valor, 
            EXTRACT(EPOCH FROM mm.fecha - lag(mm.fecha) OVER (ORDER BY mm.fecha ASC))/60  as lapso_tiempo,
            (SELECT fre.fre_valor FROM frecuencia_frecuencia fre
                    WHERE fre.var_id_id = (SELECT var_id FROM variable) AND fre.est_id_id = (SELECT est_id FROM estacion) AND fre.fre_fecha_ini < mm.fecha
                    ORDER BY fre.fre_fecha_ini DESC LIMIT 1) AS periodo_esperado
        FROM medicion mm ORDER BY mm.fecha ASC
    ),
    tabla AS (
        SELECT 
            ff.fila, 
            ff.id, 
            ff.fecha, ff.valor, 
            --ff.lapso_tiempo, ff.periodo_esperado, 
            CASE WHEN ff.fila = 1 THEN FALSE ELSE 
                CASE WHEN ff.lapso_tiempo > ff.periodo_esperado * 1.005 THEN TRUE ELSE FALSE END
            END AS salto
        FROM fechas ff ORDER BY fecha
    )
    select * from tabla;    
    """
    sql = sql.replace("var101", modelo)
    consulta = ConsultaGenericaFechaHoraGrafico.objects.raw(sql, [estacion_id, variable.var_id])

    fecha = []
    valor = []
    for fila in consulta:
        if fila.salto is True:
            fecha_intermedia = fecha_anterior + (fila.fecha - fecha_anterior)/2
            fecha.append(fecha_intermedia)
            valor.append(None)
        fecha.append(fila.fecha)
        valor.append(fila.valor)
        fecha_anterior = fila.fecha
    datos = {"fecha": fecha, "valor": valor}
    return datos


# datos validaos
def consulta_validados(estacion_id, variable, inicio, fin, profundidad):
    modelo = str(variable.var_modelo)
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    sql = """
    WITH 
    seleccion AS (
        SELECT fecha, valor, validacion 
        FROM validacion_%var_id% v 
        WHERE estacion_id = %s
    """
    if inicio:
        sql = sql + '        AND v.fecha>=\'' + str(inicio) + '\''
    if fin:
        sql = sql + '        AND v.fecha<=\'' + str(fin) + '\''
    sql = sql + """
    )
    SELECT ss.fecha, ss.valor FROM (
        SELECT fecha, MAX(validacion) AS validacion FROM seleccion GROUP BY fecha
    ) AS tbl_max
    INNER JOIN seleccion ss
    ON ss.fecha = tbl_max.fecha AND ss.validacion = tbl_max.validacion
    ORDER BY ss.fecha;
    """

    sql = sql.replace("%var_id%", modelo)
    consulta = ConsultaGenericaFechaHora.objects.raw(sql, [estacion_id, ])
    return consulta


# consulta de datos validados para graficar
def consulta_validados_graficar(estacion_id, variable, inicio, fin, profundidad):
    modelo = str(variable.var_modelo)
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    sql = """
    WITH 
    estacion AS (SELECT * FROM estacion_estacion est WHERE est.est_id = %s),
    variable AS (SELECT * FROM variable_variable var WHERE var.var_id = %s),    
    seleccion AS (
        SELECT id, fecha, valor, validacion 
        FROM validacion_%var_id% v 
        WHERE estacion_id = (SELECT est_id FROM estacion)
    """
    if inicio:
        sql = sql + '        AND v.fecha>=\'' + str(inicio) + '\''
    if fin:
        sql = sql + '        AND v.fecha<=\'' + str(fin) + '\''
    sql = sql + """
    ),
    validados AS (
        SELECT ss.id, ss.fecha, ss.valor FROM (
            SELECT fecha, MAX(validacion) AS validacion FROM seleccion GROUP BY fecha
        ) AS tbl_max
        INNER JOIN seleccion ss
        ON ss.fecha = tbl_max.fecha AND ss.validacion = tbl_max.validacion
        ORDER BY ss.fecha
    ),
    fechas AS (
        SELECT 
            row_number() OVER (ORDER BY vv.fecha ASC, vv.id ASC) as fila,
            vv.id, vv.fecha, vv.valor, 
            EXTRACT(EPOCH FROM vv.fecha - lag(vv.fecha) OVER (ORDER BY vv.fecha ASC))/60  as lapso_tiempo,
            (SELECT fre.fre_valor FROM frecuencia_frecuencia fre
                    WHERE fre.var_id_id = (SELECT var_id FROM variable) AND fre.est_id_id = (SELECT est_id FROM estacion) AND fre.fre_fecha_ini < vv.fecha
                    ORDER BY fre.fre_fecha_ini DESC LIMIT 1) AS periodo_esperado
        FROM validados vv ORDER BY vv.fecha ASC
    ),
    tabla AS (
        SELECT 
            ff.fila, 
            ff.id, 
            ff.fecha, ff.valor, 
            --ff.lapso_tiempo, ff.periodo_esperado, 
            CASE WHEN ff.fila = 1 THEN FALSE ELSE 
                CASE WHEN ff.lapso_tiempo > ff.periodo_esperado * 1.005 THEN TRUE ELSE FALSE END
            END AS salto
        FROM fechas ff ORDER BY fecha
    )
    select * from tabla;      
    """

    sql = sql.replace("%var_id%", modelo)
    consulta = ConsultaGenericaFechaHoraGrafico.objects.raw(sql, [estacion_id, variable.var_id])
    fecha = []
    valor = []
    for fila in consulta:
        # if fila.valor is not None:
        #     resultado.append([fila.fecha, fila.valor])
        if fila.salto is True:
            fecha_intermedia = fecha_anterior + (fila.fecha - fecha_anterior)/2
            fecha.append(fecha_intermedia)
            valor.append(None)
        fecha.append(fila.fecha)
        valor.append(fila.valor)
        fecha_anterior = fila.fecha
    datos = {"fecha": fecha, "valor": valor}
    return datos


# consulta del modelo de datos horarios
def consulta_horario(estacion_id, variable, inicio, fin, profundidad):
    modelo = str(variable.var_modelo)
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)
    sql = """
    WITH 
    variable AS (
        SELECT * FROM variable_variable vv WHERE vv.var_id = %%var_id%%
    ),
    consulta AS (
        select * from horario_%%var_modelo%% c 
        WHERE c.estacion_id = %%est_id%% 
        %%filtro%%
        AND c.completo_mediciones >= (SELECT vv.umbral_completo FROM variable vv)
    ),
    secuencia AS (
        SELECT * FROM generate_series( 
            (SELECT c.fecha FROM consulta c ORDER BY c.fecha ASC LIMIT 1 ),
            (SELECT c.fecha FROM consulta c ORDER BY c.fecha DESC LIMIT 1 ),
            '1 hour'::interval) fecha
    )
    select fecha, valor from secuencia s LEFT JOIN consulta c USING (fecha) ORDER BY fecha ASC;    
    """
    filtro = ""
    if profundidad:
        filtro += ' AND c.profundidad = ' + str(profundidad)
    if inicio:
        filtro += ' AND c.fecha >=\'' + str(inicio) + '\''
    if fin:
        filtro += ' AND c.fecha <=\'' + str(fin) + '\''

    sql = sql.replace('%%est_id%%', str(estacion_id))
    sql = sql.replace('%%var_modelo%%', modelo)
    sql = sql.replace('%%var_id%%', str(variable.var_id))
    sql = sql.replace('%%filtro%%', filtro)

    consulta = ConsultaGenericaFechaHora.objects.raw(sql)
    return consulta


# consulta de datos horarios par graficar
def consulta_horario_graficar(estacion_id, variable_id, inicio, fin, profundidad):
    consulta = consulta_horario(estacion_id, variable_id, inicio, fin, profundidad)
    fecha = []
    valor = []
    for fila in consulta:
        # resultado.append([fila.fecha, None if fila.valor is None else round(fila.valor, 1)])
        fecha.append(fila.fecha)
        valor.append(fila.valor)
    datos = {"fecha": fecha, "valor": valor}
    return datos


# consulta de datos diarios
def consulta_diario(estacion_id, variable, inicio, fin, profundidad):
    modelo = str(variable.var_modelo)
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)
    sql = """
    WITH 
    variable AS (
        SELECT * FROM variable_variable vv WHERE vv.var_id = %%var_id%%
    ),
    consulta AS (
        select * from diario_%%var_modelo%% c 
        WHERE c.estacion_id = %%est_id%% 
        %%filtro%%
        AND c.completo_umbral >= (SELECT vv.umbral_completo FROM variable vv)
    ),
    secuencia AS (
        SELECT * FROM generate_series( 
            (SELECT c.fecha FROM consulta c ORDER BY c.fecha ASC LIMIT 1),
            (SELECT c.fecha FROM consulta c ORDER BY c.fecha DESC LIMIT 1),
            '1 day'::interval) fecha
    )
    select fecha, valor from secuencia s LEFT JOIN consulta c USING (fecha) ORDER BY fecha ASC;    
    """
    filtro = ""
    if profundidad:
        filtro += '        AND profundidad = ' + str(profundidad)
    if inicio:
        filtro += ' AND c.fecha >=\'' + str(inicio) + '\''
    if fin:
        filtro += ' AND c.fecha <=\'' + str(fin) + '\''

    sql = sql.replace('%%est_id%%', str(estacion_id))
    sql = sql.replace('%%var_modelo%%', modelo)
    sql = sql.replace('%%var_id%%', str(variable.var_id))
    sql = sql.replace('%%filtro%%', filtro)

    consulta = ConsultaGenericaFecha.objects.raw(sql)
    return consulta


#consulta de datos diarios para graficar
def consulta_diario_graficar(estacion_id, variable_id, inicio, fin, profundidad):
    consulta = consulta_diario(estacion_id, variable_id, inicio, fin, profundidad)
    fecha = []
    valor = []
    for fila in consulta:
        # resultado.append([fila.fecha, None if fila.valor is None else round(fila.valor, 1)])
        fecha.append(fila.fecha)
        valor.append(fila.valor)
    datos = {"fecha": fecha, "valor": valor}
    return datos


def consulta_mensual(estacion_id, variable, inicio, fin, profundidad):
    modelo = str(variable.var_modelo)
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)
    sql = """
    WITH 
    variable AS (
        SELECT * FROM variable_variable vv WHERE vv.var_id = %%var_id%%
    ),
    consulta AS (
        select * from mensual_%%var_modelo%%mensual c 
        WHERE c.estacion_id = %%est_id%% 
        %%filtro%%
        AND c.completo_umbral >= (SELECT vv.umbral_completo FROM variable vv)
    ),
    secuencia AS (
        SELECT * FROM generate_series( 
            (SELECT c.fecha FROM consulta c ORDER BY c.fecha ASC LIMIT 1),
            (SELECT c.fecha FROM consulta c ORDER BY c.fecha DESC LIMIT 1),
            '1 month'::interval) fecha
    )
    select fecha, valor from secuencia s LEFT JOIN consulta c USING (fecha) ORDER BY fecha ASC;    
    """
    filtro = ""
    if profundidad:
        filtro += '        AND profundidad = ' + str(profundidad)
    if inicio:
        filtro += ' AND c.fecha >=\'' + str(inicio) + '\''
    if fin:
        filtro += ' AND c.fecha <=\'' + str(fin) + '\''

    sql = sql.replace('%%est_id%%', str(estacion_id))
    sql = sql.replace('%%var_modelo%%', modelo)
    sql = sql.replace('%%var_id%%', str(variable.var_id))
    sql = sql.replace('%%filtro%%', filtro)

    consulta = ConsultaGenericaFecha.objects.raw(sql)
    return consulta


def consulta_mensual_graficar(estacion_id, variable_id, inicio, fin, profundidad):
    consulta = consulta_mensual(estacion_id, variable_id, inicio, fin, profundidad)
    fecha = []
    valor = []
    for fila in consulta:
        # resultado.append([fila.fecha, None if fila.valor is None else round(fila.valor, 1)])
        fecha.append(fila.fecha)
        valor.append(fila.valor)
    datos = {"fecha": fecha, "valor": valor}
    return datos



