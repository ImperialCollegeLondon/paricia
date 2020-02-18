# -*- coding: utf-8 -*-
from anuarios.models import HumedadSuelo
from anuarios.models import PresionAtmosferica
from anuarios.models import TemperaturaAgua
from anuarios.models import Caudal
from anuarios.models import NivelAgua
import plotly.offline as opy
import plotly.graph_objs as go
from reportes.titulos import Titulos
import calendar

from openpyxl.chart import (
    ScatterChart,
    LineChart,
    Reference,
    Series,
)
from openpyxl.chart.axis import DateAxis

# clase para anuario de las las variables HSU, PAT, TAG, CAU, NAG
class TypeI(Titulos):

    @staticmethod
    def consulta(estacion, variable, periodo):
        if variable == 6:
            informacion = list(HumedadSuelo.objects.filter(est_id=estacion).filter(hsu_periodo=periodo))
        elif variable == 8:
            informacion = list(PresionAtmosferica.objects.filter(est_id=estacion).filter(pat_periodo=periodo))
        elif variable == 9:
            informacion = list(TemperaturaAgua.objects.filter(est_id=estacion).filter(tag_periodo=periodo))
        elif variable == 10:
            informacion = list(Caudal.objects.filter(est_id=estacion).filter(cau_periodo=periodo))
        elif variable == 11:
            informacion = list(NivelAgua.objects.filter(est_id=estacion).filter(nag_periodo=periodo))
        return informacion

    def matriz(self, estacion, variable, periodo):
        datos = self.consulta(estacion, variable, periodo)
        return datos

    def grafico(self, estacion, variable, periodo):
        datos = self.consulta(estacion, variable, periodo)
        if len(datos)>0:
            meses = []
            max_simple = []
            min_simple = []
            avg_simple = []
            for item in datos:
                if variable == 6:
                    meses.append(str(calendar.month_abbr[item.hsu_mes]))
                    max_simple.append(item.hsu_maximo)
                    min_simple.append(item.hsu_minimo)
                    avg_simple.append(item.hsu_promedio)
                elif variable == 8:
                    meses.append(str(calendar.month_abbr[item.pat_mes]))
                    max_simple.append(item.pat_maximo)
                    min_simple.append(item.pat_minimo)
                    avg_simple.append(item.pat_promedio)
                elif variable == 9:
                    meses.append(str(calendar.month_abbr[item.tag_mes]))
                    max_simple.append(item.tag_maximo)
                    min_simple.append(item.tag_minimo)
                    avg_simple.append(item.tag_promedio)
                elif variable == 10:
                    meses.append(str(calendar.month_abbr[item.cau_mes]))
                    max_simple.append(item.cau_maximo)
                    min_simple.append(item.cau_minimo)
                    avg_simple.append(item.cau_promedio)
                elif variable == 11:
                    meses.append(str(calendar.month_abbr[item.nag_mes]))
                    max_simple.append(item.nag_maximo)
                    min_simple.append(item.nag_minimo)
                    avg_simple.append(item.nag_promedio)

            trace0 = go.Scatter(
                x=meses,
                y=max_simple,
                name='Max',
                line=dict(
                    color=('rgb(22, 96, 167)'),
                    width=4)
            )
            trace1 = go.Scatter(
                x=meses,
                y=min_simple,
                name='Min',
                line=dict(
                    color=('rgb(205, 12, 24)'),
                    width=4, )
            )
            trace2 = go.Scatter(
                x=meses,
                y=avg_simple,
                name='Media',
                line=dict(
                    color=('rgb(50, 205, 50)'),
                    width=4, )
            )
            data = [trace0, trace1, trace2]
            layout = go.Layout(
                title=str(self.titulo_grafico(variable)) + str(" (") + str(self.titulo_unidad(variable)) + str(")"))
            figure = go.Figure(data=data, layout=layout)
            div = opy.plot(figure, auto_open=False, output_type='div')
            return div
        return False

    def tabla_excel(self, ws, estacion, variable, periodo):
        fila = 5
        col_fin = 11
        col = 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col_fin)
        subtitle = ws.cell(row=fila, column=col)
        subtitle.value = self.get_titulo(variable.var_id)
        self.set_style(cell=subtitle, font='font_bold_10', alignment='center',
                       border='border_thin', fill='light_salmon')

        fila += 1
        col = 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila + 1, end_column=col)
        cell = ws.cell(row=fila, column=col)
        cell.value = "MES"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')
        col += 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col+2)
        cell = ws.cell(row=fila, column=col)
        cell.value = str(variable.var_nombre) + str(" (") + str(variable.uni_id.uni_sigla) + str(")")
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')
        fila += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Máximo"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Mínimo"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Media"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        matriz = self.matriz(estacion, variable.var_id, periodo)
        fila += 1
        col = 1

        for item in matriz:
            cell = ws.cell(row=fila, column=col)
            cell.value = self.get_mes_anio(self.get_item_mes(variable.var_id,item))
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')
            cell = ws.cell(row=fila, column=col+1)
            cell.value = self.get_item_maximo(variable.var_id,item)
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')

            cell = ws.cell(row=fila, column=col+2)
            cell.value = self.get_item_minimo(variable.var_id,item)
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')

            cell = ws.cell(row=fila, column=col+3)
            cell.value = self.get_item_promedio(variable.var_id,item)
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')

            fila += 1

    @staticmethod
    def grafico_excel(ws, variable):
        c1 = ScatterChart()
        c1.title = str(variable.var_nombre) + str(" (") + str(variable.uni_id.uni_sigla) + str(")")
        c1.style = 13
        c1.y_axis.title = str(variable.uni_id.uni_sigla)
        c1.x_axis.title = 'Meses'

        '''data = Reference(ws, min_col=2, min_row=7, max_col=4, max_row=19)
        cats = Reference(ws, min_col=1, min_row=8, max_row=19)

        c1.add_data(data, titles_from_data=True)

        s1 = c1.series[0]
        s1.marker.symbol = "triangle"
        s1.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
        s1.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outlin
        s1.graphicalProperties.line.noFill = True

        s2 = c1.series[1]
        s2.graphicalProperties.line.solidFill = "00AAAA"
        s2.graphicalProperties.line.dashStyle = "sysDot"
        s2.graphicalProperties.line.width = 100050  # width in EMUs

        c1.set_categories(cats)
        c1.grouping = "stacked"'''

        xvalues = Reference(ws, min_col=1, min_row=8, max_row=19)
        for i in range(2, 5):
            values = Reference(ws, min_col=i, min_row=7, max_row=19)
            series = Series(values, xvalues, title_from_data=True)
            c1.series.append(series)

        serie_max = c1.series[0]
        serie_max.marker.symbol = "diamond"
        serie_max.graphicalProperties.line.solidFill = "1660a7"

        serie_min = c1.series[1]
        serie_min.marker.symbol = "triangle"
        serie_min.graphicalProperties.line.solidFill = "cd0c18"

        serie_pro = c1.series[2]
        serie_pro.marker.symbol = "square"
        serie_pro.graphicalProperties.line.solidFill = "32cd32"

        ws.add_chart(c1, "E6")


    @staticmethod
    def get_titulo(variable):
        if variable == 6:
            titulo = 'Humedad del Suelo - Valores medios diarios, absolutos maximos y mimimos'
        elif variable == 8:
            titulo = 'Presion Atomsferica - Valores medios diarios, absolutos maximos y mimimos'
        elif variable == 9:
            titulo = 'Temperatura de Agua - Valores medios diarios, medios maximos y mimimos'
        elif variable == 10:
            titulo = 'Caudal - Valores medios diarios, medios maximos y mimimos'
        elif variable == 11:
            titulo = 'Nivel del agua - Valores medios diarios, medios maximos y mimimos'
        return titulo
    
    @staticmethod
    def get_item_promedio(variable, item):
        if variable == 6:
            promedio = item.hsu_promedio
        elif variable == 8:
            promedio = item.pat_promedio
        elif variable == 9:
            promedio = item.tag_promedio
        elif variable == 10:
            promedio = item.cau_promedio
        elif variable == 11:
            promedio = item.nag_promedio
        return promedio

    @staticmethod
    def get_item_maximo(variable, item):
        if variable == 6:
            maximo = item.hsu_maximo
        elif variable == 8:
            maximo = item.pat_maximo
        elif variable == 9:
            maximo = item.tag_maximo
        elif variable == 10:
            maximo = item.cau_maximo
        elif variable == 11:
            maximo = item.nag_maximo
        return maximo

    @staticmethod
    def get_item_minimo(variable, item):
        if variable == 6:
            minimo = item.hsu_minimo
        elif variable == 8:
            minimo = item.pat_minimo
        elif variable == 9:
            minimo = item.tag_minimo
        elif variable == 10:
            minimo = item.cau_minimo
        elif variable == 11:
            minimo = item.nag_minimo
        return minimo

    @staticmethod
    def get_item_mes(variable, item):
        if variable == 6:
            mes = item.hsu_mes
        elif variable == 8:
            mes = item.pat_mes
        elif variable == 9:
            mes = item.tag_mes
        elif variable == 10:
            mes = item.cau_mes
        elif variable == 11:
            mes = item.nag_mes
        return mes





