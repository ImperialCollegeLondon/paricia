from reportes.typeI import TypeI
from reportes.typeII import TypeII
from reportes.typeIII import TypeIII
from reportes.typeIV import TypeIV
from reportes.typeV import TypeV
from reportes.typeVI import TypeVI


from cruce.models import Cruce

from django.http import HttpResponse
from sedc.settings import BASE_DIR
# librerias para manejar los archivos EXCEL
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import Workbook
from openpyxl.drawing.image import Image

from openpyxl.chart import (
    LineChart,
    Reference,
)




def reporte_excel_anuario(form):
    estacion = form.cleaned_data['estacion']
    periodo = form.cleaned_data['anio']
    variables = list(Cruce.objects
                     .filter(est_id=form.cleaned_data['estacion'])

                     )

    # ruta de la imagen
    ruta = str(BASE_DIR) + '/media/logo_fonag.jpg'
    img = Image(ruta)

    # Creamos el libro de trabajo
    wb = Workbook()

    # humedadsuelo,presionatmosferica,temperaturaagua,caudal,nivelagua
    typei = [6, 8, 9, 10, 11]
    # precipitacion
    typeii = [1]
    # temperaturaaire
    typeiii = [2]
    # humedadaire
    typeiv = [3]
    # direccion y velocidad
    typev = [4, 5]
    # radiacion
    typevi = [7]

    obj_typei = TypeI()
    obj_typeii = TypeII()
    obj_typeiii = TypeIII()
    obj_typeiv = TypeIV()
    obj_typev = TypeV()
    obj_typevi = TypeVI()
    for item in variables:
        print(item.var_id.var_nombre)
        variable = item.var_id
        if variable.var_id == 1:
            if len(wb.sheetnames) == 1:
                # ws_pre = wb.active
                # ws_pre.title = variable.var_nombre
                ws_pre = wb.create_sheet(str(variable.var_nombre), 0)
            else:
                ws_pre = wb.create_sheet(str(variable.var_nombre), -1)

            obj_typeii.set_encabezado_excel(ws_pre, estacion, periodo)
            obj_typeii.tabla_excel(ws_pre, estacion, variable, periodo)
            obj_typeii.grafico_excel(ws_pre, variable, periodo)

        elif variable.var_id == 2:
            if len(wb.sheetnames) == 1:
                ws_tai = wb.create_sheet(str(variable.var_nombre), 0)
            else:
                ws_tai = wb.create_sheet(str(variable.var_nombre), -1)

            obj_typeiii.set_encabezado_excel(ws_tai, estacion, periodo)
            obj_typeiii.tabla_excel(ws_tai, estacion, variable, periodo)
            obj_typeiii.grafico_excel(ws_tai, variable, periodo)

        elif variable.var_id == 3:
            if len(wb.sheetnames) == 1:
                ws_hai = wb.create_sheet(str(variable.var_nombre), 0)
            else:
                ws_hai = wb.create_sheet(str(variable.var_nombre), -1)

            obj_typeiv.set_encabezado_excel(ws_hai, estacion, periodo)
            obj_typeiv.tabla_excel(ws_hai, estacion, variable, periodo)
            obj_typeiv.grafico_excel(ws_hai, variable, periodo)

        elif variable.var_id == 4:
            if len(wb.sheetnames) == 1:
                ws_vvi = wb.create_sheet(str(variable.var_nombre), 0)
            else:
                print(variable.var_nombre)
                ws_vvi = wb.create_sheet(str(variable.var_nombre), -1)

            obj_typev.set_encabezado_excel(ws_vvi, estacion, periodo)
            obj_typev.tabla_excel(ws_vvi, estacion, variable, periodo)
            # obj_typeiv.grafico_excel(ws_vvi, variable, periodo)

        elif variable.var_id == 6:
            if len(wb.sheetnames) == 1:
                ws_hsu = wb.create_sheet(str(variable.var_nombre), 0)
            else:
                ws_hsu = wb.create_sheet(str(variable.var_nombre), -1)

            obj_typei.set_encabezado_excel(ws_hsu, estacion, periodo)
            obj_typei.tabla_excel(ws_hsu, estacion, variable, periodo)
            obj_typei.grafico_excel(ws_hsu, variable, periodo)

        elif variable.var_id == 7:
            if len(wb.sheetnames) == 1:
                ws_rad = wb.create_sheet(str(variable.var_nombre), 0)
            else:
                ws_rad = wb.create_sheet(str(variable.var_nombre), -1)

            obj_typevi.set_encabezado_excel(ws_rad, estacion, periodo)
            obj_typevi.tabla_excel(ws_rad, estacion, variable, periodo, 'maxima')
            obj_typevi.tabla_excel(ws_rad, estacion, variable, periodo, 'minima')
            # obj_typei.grafico_excel(ws_hsu, variable, periodo)

        elif variable.var_id == 8:
            if len(wb.sheetnames) == 1:
                ws_pat = wb.create_sheet(str(variable.var_nombre), 0)
            else:
                ws_pat = wb.create_sheet(str(variable.var_nombre), -1)

            obj_typei.set_encabezado_excel(ws_pat, estacion, periodo)
            obj_typei.tabla_excel(ws_pat, estacion, variable, periodo)
            obj_typei.grafico_excel(ws_pat, variable, periodo)

        elif variable.var_id == 9:
            if len(wb.sheetnames) == 1:
                ws_tag = wb.create_sheet(str(variable.var_nombre), 0)
            else:
                ws_tag = wb.create_sheet(str(variable.var_nombre), -1)

            obj_typei.set_encabezado_excel(ws_tag, estacion, periodo)
            obj_typei.tabla_excel(ws_tag, estacion, variable, periodo)
            obj_typei.grafico_excel(ws_tag, variable, periodo)

        elif variable.var_id == 10:
            if len(wb.sheetnames) == 1:
                ws_cau = wb.create_sheet(str(variable.var_nombre), 0)
            else:
                ws_cau = wb.create_sheet(str(variable.var_nombre), -1)

            obj_typei.set_encabezado_excel(ws_cau, estacion, periodo)
            obj_typei.tabla_excel(ws_cau, estacion, variable, periodo)
            obj_typei.grafico_excel(ws_cau, variable, periodo)

        elif variable.var_id == 11:
            if len(wb.sheetnames) == 1:
                ws_nag = wb.create_sheet(str(variable.var_nombre), 0)
            else:
                ws_nag = wb.create_sheet(str(variable.var_nombre), -1)

            obj_typei.set_encabezado_excel(ws_nag, estacion, periodo)
            obj_typei.tabla_excel(ws_nag, estacion, variable, periodo)
            obj_typei.grafico_excel(ws_nag, variable, periodo)




    # Establecemos el nombre del archivo
    nombre_archivo = str('"') + str(estacion.est_codigo) + str("_") + str(periodo) + str('.xlsx"')
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response
