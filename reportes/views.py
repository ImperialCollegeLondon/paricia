from django.views.generic.base import TemplateView
from django.views.generic import FormView
from reportes.forms import AnuarioForm
from reportes.consultas.forms import MedicionSearchForm, ComparacionForm, VariableForm, EstacionVariableSearchForm
import csv
from django.http import HttpResponse
# from django.template import loader, Context
from reportes.consultas.functions import (datos_horarios_json, datos_diarios, datos_5minutos, datos_horarios,
                                          datos_instantaneos, datos_mensuales, datos_estacion)
from reportes.functions import filtrar, comparar, comparar_variables, consultar_datos
from datetime import date
from django.shortcuts import render
from django.http import JsonResponse
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.drawing.image import Image

from openpyxl.chart import (
    LineChart,
    Reference,
    Series,
)
from sedc.settings import BASE_DIR


class ReportesAnuario(FormView):
    template_name = 'reportes/anuario_reporte.html'
    form_class = AnuarioForm
    success_url = '/reportes/anuario/'
    lista = {}

    def get_context_data(self, **kwargs):
        context = super(ReportesAnuario, self).get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['base_template'] = "index.html"
        else:
            context['base_template'] = "index_invitado.html"
        return context

    def post(self, request, *args, **kwargs):
        form = AnuarioForm(self.request.POST or None)
        if form.is_valid() and self.request.is_ajax():
            self.lista = filtrar(form)
            return render(request,'reportes/anuario/anuario.html',self.lista)
        return render(request, 'home/form_error.html', {'form': form})


# vista para comparar tres estaciones una sola variable
class ComparacionValores(FormView):
    template_name = 'reportes/comparacion_reporte.html'
    form_class = ComparacionForm
    success_url = '/reportes/comparacion/'
    grafico = []

    def get_context_data(self, **kwargs):
        context = super(ComparacionValores, self).get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['base_template'] = "index.html"
        else:
            context['base_template'] = "index_invitado.html"
        return context

    def post(self, request, *args, **kwargs):
        form = ComparacionForm(self.request.POST or None)
        if form.is_valid() and self.request.is_ajax():
            self.grafico = comparar(form)
            plantilla = 'reportes/consultas/grafico.html'
            diccionario = {'grafico': self.grafico}
            return render(request, plantilla, diccionario)

        return render(request, 'home/form_error.html', {'form': form})


# vista para comparar 2 estaciones y dos Variables
class ComparacionVariables(FormView):
    template_name = 'reportes/comparacion_variable.html'
    form_class = VariableForm
    success_url = '/reportes/compararvariable/'
    grafico = []

    def get_context_data(self, **kwargs):
        context = super(ComparacionVariables, self).get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['base_template'] = "index.html"
        else:
            context['base_template'] = "index_invitado.html"
        return context

    def post(self, request, *args, **kwargs):
        form = VariableForm(self.request.POST or None)
        if form.is_valid() and self.request.is_ajax():
            '''self.grafico = comparar_variable(form)
            plantilla = 'reportes/consultas/grafico.html'
            diccionario = {'grafico': self.grafico}
            return render(request, plantilla, diccionario)'''
            datos = comparar_variables(form)
            return JsonResponse(datos, safe=False)
        return render(request, 'home/form_error.html', {'form': form})


# consultas por periodo y frecuencia horaria, diaria y mensual
class ConsultasPeriodo(FormView):
    template_name = 'reportes/consultas_periodo.html'
    form_class = MedicionSearchForm
    success_url = '/reportes/consultas'
    frecuencia = str("")
    valores = []
    grafico = []

    def post(self, request, *args, **kwargs):
        form = MedicionSearchForm(self.request.POST or None)
        if form.is_valid():
            self.frecuencia = form.cleaned_data["frecuencia"]
            if self.request.is_ajax():
                '''self.grafico = grafico(form)
                return render(request, 'reportes/consultas/grafico.html',
                              {'grafico': self.grafico, 'frecuencia': self.frecuencia})'''
                datos = consultar_datos(form)
                return JsonResponse(datos, safe=False)
            else:
                if self.frecuencia == "0":
                    return self.export_csv(form)
                else:
                    return self.export_excel(self.frecuencia, form)
        return render(request, 'home/form_error.html', {'form': form})

    def get_context_data(self, **kwargs):
        context = super(ConsultasPeriodo, self).get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['base_template'] = "index.html"
        else:
            context['base_template'] = "index_invitado.html"
        context['grafico'] = self.grafico
        return context

    def export_csv(self, form):
        estacion = form.cleaned_data['estacion']
        variable = form.cleaned_data['variable']
        fecha_inicio = form.cleaned_data['inicio']
        fecha_fin = form.cleaned_data['fin']
        if fecha_inicio is None:
            fecha_inicio = estacion.est_fecha_inicio
        if fecha_fin is None:
            fecha_fin = date.today()
        valores, maximos, minimos, tiempo = datos_instantaneos(estacion, variable,fecha_inicio, fecha_fin)
        print(len(valores))
        # Establecemos el nombre del archivo
        nombre_archivo = str('"') + str(estacion.est_codigo) + str("_") + str(variable.var_nombre) + str('.csv"')
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = contenido
        writer = csv.writer(response)
        if variable.var_id == 1:
            writer.writerow(['fecha', 'valor'])
            for valor, fecha in zip(valores, tiempo):
                writer.writerow([fecha, valor])
        else:
            writer.writerow(['fecha', 'valor', 'maximo', 'minimo'])
            for valor, maximo, minimo, fecha in zip(valores, maximos, minimos, tiempo):
                writer.writerow([fecha, valor, maximo, minimo])

        return response

    def export_excel(self,frecuencia,form):
        estacion = form.cleaned_data['estacion']
        variable = form.cleaned_data['variable']
        fecha_inicio = form.cleaned_data['inicio']
        fecha_fin = form.cleaned_data['fin']
        if fecha_inicio is None:
            fecha_inicio = estacion.est_fecha_inicio
        if fecha_fin is None:
            fecha_fin = date.today()
        if frecuencia == "1":
            valores, maximos_abs, maximos_pro, minimos_abs, minimos_pro, tiempo = datos_5minutos(estacion, variable, fecha_inicio, fecha_fin)
        elif frecuencia == "2":
            valores, maximos_abs, maximos_pro, minimos_abs, minimos_pro, tiempo = datos_horarios(estacion, variable, fecha_inicio, fecha_fin)
        elif frecuencia == "3":
            valores, maximos_abs, maximos_pro, minimos_abs, minimos_pro, tiempo = datos_diarios(estacion, variable, fecha_inicio, fecha_fin)
        else:
            valores, maximos_abs, maximos_pro, minimos_abs, minimos_pro, tiempo = datos_mensuales(estacion, variable, fecha_inicio, fecha_fin)
        # ruta de la imagen
        ruta = str(BASE_DIR) + '/media/logo_fonag.jpg'
        img = Image(ruta)
        # estilo de negrita
        font_bold = Font(bold=True)
        # Creamos el libro de trabajo
        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws['B1'] = 'Reporte de Datos Hidrometerológicos'
        ws['B1'].font=font_bold
        ws.add_image(img, 'G1')
        ws.merge_cells('B1:F1')
        ws['A4'] = 'Estación'
        ws['A4'].font = font_bold
        ws['B4'] = estacion.est_codigo
        ws['C4'] = estacion.est_nombre
        ws.merge_cells('C4:E4')
        ws['F4'] = 'Variable'
        ws['F4'].font = font_bold
        ws['G4'] = variable.var_nombre
        ws['B6'] = 'Coordenadas Geográfica UTM (DATUM WGS 84)'
        ws['B6'].font = font_bold
        ws.merge_cells('B6:G6')
        ws['A7'] = 'Latitud'
        ws['A7'].font = font_bold
        ws['B7'] = estacion.est_latitud
        ws['F7'] = 'Longitud'
        ws['F7'].font = font_bold
        ws['G7'] = estacion.est_longitud

        # Creamos los encabezados desde la celda B9 hasta la E9
        ws['A9'] = 'Fecha'
        ws['B9'] = 'Valor'

        cont = 10
        # Recorremos el conjunto de datos
        if len(maximos_abs) != maximos_abs.count(None) and len(maximos_pro) != maximos_pro.count(None):
            ws['C9'] = 'Max absoluto'
            ws['D9'] = 'Min absoluto'
            ws['E9'] = 'Max promedio'
            ws['F9'] = 'Min promedio'
            for valor, maximo_abs, maximos_pro, minimo_abs, minimo_pro, fecha in zip(valores, maximos_abs, maximos_pro, minimos_abs, minimos_pro, tiempo):
                ws.cell(row=cont, column=1).value = fecha
                ws.cell(row=cont, column=2).value = valor
                ws.cell(row=cont, column=3).value = maximo_abs
                ws.cell(row=cont, column=4).value = minimo_abs
                ws.cell(row=cont, column=5).value = maximos_pro
                ws.cell(row=cont, column=6).value = minimo_pro
                cont = cont + 1
        else:
            ws['C9'] = 'Max promedio'
            ws['D9'] = 'Min promedio'
            for valor, maximos_pro, minimo_pro, fecha in zip(valores, maximos_pro, minimos_pro, tiempo):
                ws.cell(row=cont, column=1).value = fecha
                ws.cell(row=cont, column=2).value = valor
                ws.cell(row=cont, column=3).value = maximos_pro
                ws.cell(row=cont, column=4).value = minimo_pro
                cont = cont + 1

        # grafico
        chart = LineChart()
        chart.title = variable.var_nombre
        chart.style = 12
        chart.x_axis.title = 'Tiempo'
        chart.x_axis.number_format = 'dd-mm-yyyy'
        chart.x_axis.majorTimeUnit = 'days'
        chart.y_axis.title = variable.var_nombre
        final = 10+len(tiempo)
        data = Reference(ws, min_col=2, min_row=9, max_col=4, max_row=final-1)
        chart.add_data(data, titles_from_data=True)
        s0 = chart.series[0]
        s0.graphicalProperties.line.solidFill = "32CD32"
        s0.graphicalProperties.line.width = 10
        s0.smooth = True
        s1 = chart.series[1]
        s1.graphicalProperties.line.solidFill = "1645A7"
        s1.graphicalProperties.line.width = 10
        s1.smooth = True
        s2 = chart.series[2]
        s2.graphicalProperties.line.solidFill = "CD0C18"
        s2.graphicalProperties.line.width = 10
        s2.smooth = True
        dates = Reference(ws, min_col=1, min_row=10, max_row=final)
        chart.set_categories(dates)
        if frecuencia != "0":
            ws.add_chart(chart, "H9")
        # Establecemos el nombre del archivo
        nombre_archivo = str('"')+str(estacion.est_codigo) + str("_") + str(variable.var_nombre) + str('.xlsx"')
        # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


# consultas por periodo de todas las variables
class ConsultasEstacionVariable(FormView):
    template_name = 'reportes/consultas_estacion.html'
    form_class = EstacionVariableSearchForm
    success_url = '/reportes/estacionvariable'
    valores = []

    def post(self, request, *args, **kwargs):
        form = EstacionVariableSearchForm(self.request.POST or None)
        if form.is_valid():
            return self.export_csv(form)
        return render(request, 'home/form_error.html', {'form': form})

    def get_context_data(self, **kwargs):
        context = super(ConsultasEstacionVariable, self).get_context_data(**kwargs)
        return context

    def export_csv(self, form):
        estacion = form.cleaned_data['estacion']
        fecha_inicio = form.cleaned_data['inicio']
        fecha_fin = form.cleaned_data['fin']
        valores = datos_estacion(estacion, fecha_inicio, fecha_fin)
        # Establecemos el nombre del archivo
        nombre_archivo = str('"') + str(estacion.est_codigo) + str('.csv"')
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = contenido
        writer = csv.writer(response)
        num_col = len(valores)
        num_fil = len(valores[0])
        for i in range(num_fil):
            fila=[]
            for j in range(num_col):
                if i < len(valores[j]):
                    fila.append(valores[j][i])
                else:
                    fila.append('')
            writer.writerow(fila)
        return response


# web service para consultar datos horarios
def datos_json_horarios(request, est_id, var_id, fec_ini, fec_fin):
    datos = datos_horarios_json(est_id, var_id, fec_ini, fec_fin)
    return JsonResponse(datos, safe=False)
