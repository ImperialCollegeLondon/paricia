# -*- coding: utf-8 -*-

################################################################################################
# Plataforma para la Iniciativa Regional de Monitoreo Hidrológico de Ecosistemas Andinos (iMHEA)
# basada en los desarrollos realizados por:
#     1) FONDO PARA LA PROTECCIÓN DEL AGUA (FONAG), Ecuador.
#         Contacto: info@fonag.org.ec
#     2) EMPRESA PÚBLICA METROPOLITANA DE AGUA POTABLE Y SANEAMIENTO DE QUITO (EPMAPS), Ecuador.
#         Contacto: paramh2o@aguaquito.gob.ec
#
#  IMPORTANTE: Mantener o incluir esta cabecera con la mención de las instituciones creadoras,
#              ya sea en uso total o parcial del código.


import calendar

import plotly.graph_objs as go
import plotly.offline as opy
from django.db.models import Avg
from openpyxl.chart import Reference, ScatterChart, Series

from anuarios.models import Var3Anuarios
from reportes_v2.titulos import Titulos


# clase para anuario de la variable HAI
class TypeIV(Titulos):

    """@staticmethod
    def consulta(estacion, periodo):
        # annotate agrupa los valores en base a un campo y a una operacion
        informacion = list(Var3Anuarios.objects.filter(est_id=estacion).filter(hai_periodo=periodo))
        return informacion"""

    def matriz(self, estacion, variable, periodo):
        datos = self.consulta(estacion, variable, periodo)
        return datos

    def grafico(self, estacion, variable, periodo):
        datos = self.consulta(estacion, variable, periodo)
        historico = self.datos_historicos(estacion, variable, periodo)
        if datos:
            meses = []
            max_simple = []
            min_simple = []
            avg_simple = []
            for item in datos:
                meses.append(str(calendar.month_abbr[item.hai_mes]))
                max_simple.append(item.hai_maximo)
                min_simple.append(item.hai_minimo)
                avg_simple.append(item.hai_promedio)
            trace0 = go.Scatter(
                x=meses,
                y=max_simple,
                name="Max",
                line=dict(color="rgb(22, 96, 167)", width=4),
            )
            trace1 = go.Scatter(
                x=meses,
                y=min_simple,
                name="Min",
                line=dict(
                    color="rgb(205, 12, 24)",
                    width=4,
                ),
            )
            trace2 = go.Scatter(
                x=meses,
                y=avg_simple,
                name="Media",
                line=dict(
                    color="rgb(50, 205, 50)",
                    width=4,
                ),
            )
            # """  trace3 = go.Scatter(
            #     x=meses,
            #     y=historico,
            #     name='Media',
            #     line=dict(
            #         color='rgb(125, 96, 160)',
            #         width=4, )
            # )

            # data = [trace0, trace1, trace2, trace3] """
            data = [trace0, trace1, trace2]
            layout = go.Layout(
                title=str(variable.var_nombre)
                + str(" (")
                + str(variable.uni_id.uni_sigla)
                + str(") ")
            )
            figure = go.Figure(data=data, layout=layout)
            figure.update_layout(legend_orientation="h")
            div = opy.plot(figure, auto_open=False, output_type="div")
            return div
        return False

    def tabla_excel(self, ws, estacion, variable, periodo):
        fila = 5
        col_fin = 11
        col = 1
        ws.merge_cells(
            start_row=fila, start_column=col, end_row=fila, end_column=col_fin
        )
        subtitle = ws.cell(row=fila, column=col)
        subtitle.value = (
            "Humedad Relativa del Aire - Valores medios mensuales, "
            "absolutos maximos y mimimos "
        )
        self.set_style(
            cell=subtitle,
            font="font_bold_10",
            alignment="center",
            border="border_thin",
            fill="light_salmon",
        )

        fila += 1

        ws.merge_cells(
            start_row=fila, start_column=col, end_row=fila + 1, end_column=col
        )
        cell = ws.cell(row=fila, column=col)
        cell.value = "MES"
        self.set_style(
            cell=cell, font="font_10", alignment="center", border="border_thin"
        )

        col += 1

        ws.merge_cells(
            start_row=fila, start_column=col, end_row=fila, end_column=col + 5
        )
        cell = ws.cell(row=fila, column=col)
        cell_final = ws.cell(row=fila, column=col + 5)
        cell.value = "Humedad Relativa del Aire (%)"
        self.set_style(
            cell=cell, font="font_10", alignment="center", border="border_thin"
        )
        self.set_style(
            cell=cell_final, font="font_10", alignment="center", border="border_thin"
        )

        fila += 1
        col = 2

        cell = ws.cell(row=fila, column=col)
        cell.value = "Max"
        self.set_style(
            cell=cell, font="font_10", alignment="center", border="border_thin"
        )

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Día"
        self.set_style(
            cell=cell, font="font_10", alignment="center", border="border_thin"
        )

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Min"
        self.set_style(
            cell=cell, font="font_10", alignment="center", border="border_thin"
        )

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Día"
        self.set_style(
            cell=cell, font="font_10", alignment="center", border="border_thin"
        )

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Mensual"
        self.set_style(
            cell=cell, font="font_10", alignment="center", border="border_thin"
        )

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Histórica"
        self.set_style(
            cell=cell, font="font_10", alignment="center", border="border_thin"
        )

        matriz = self.matriz(estacion, variable, periodo)
        media_historica = self.datos_historicos(estacion, variable, periodo)
        fila += 1
        col = 1

        for item in matriz:
            cell = ws.cell(row=fila, column=col)
            cell.value = self.get_mes_anio(item.hai_mes)
            self.set_style(
                cell=cell, font="font_10", alignment="left", border="border_thin"
            )
            cell = ws.cell(row=fila, column=col + 1)
            cell.value = item.hai_maximo
            self.set_style(
                cell=cell, font="font_10", alignment="center", border="border_thin"
            )

            cell = ws.cell(row=fila, column=col + 2)
            cell.value = item.hai_maximo_dia
            self.set_style(
                cell=cell, font="font_10", alignment="center", border="border_thin"
            )

            cell = ws.cell(row=fila, column=col + 3)
            cell.value = item.hai_minimo
            self.set_style(
                cell=cell, font="font_10", alignment="center", border="border_thin"
            )

            cell = ws.cell(row=fila, column=col + 4)
            cell.value = item.hai_minimo_dia
            self.set_style(
                cell=cell, font="font_10", alignment="center", border="border_thin"
            )

            cell = ws.cell(row=fila, column=col + 5)
            cell.value = item.hai_promedio
            self.set_style(
                cell=cell, font="font_10", alignment="center", border="border_thin"
            )
            if len(media_historica) > 0 and len(media_historica) > item.hai_mes:
                cell = ws.cell(row=fila, column=col + 6)
                cell.value = round(media_historica[item.hai_mes - 1], 2)
                self.set_style(
                    cell=cell, font="font_10", alignment="center", border="border_thin"
                )

            fila += 1

    @staticmethod
    def grafico_excel(ws, variable, periodo):
        c1 = ScatterChart()
        c1.title = (
            str(variable.var_nombre)
            + str(" (")
            + str(variable.uni_id.uni_sigla)
            + str(") ")
            + str(periodo)
        )
        # c1.style = 13
        # c1.y_axis.title = str(variable.uni_id.uni_sigla)
        c1.x_axis.title = "Meses"

        xvalues = Reference(ws, min_col=1, min_row=8, max_row=19)
        cols_data = [2, 4, 6, 7]
        for i in cols_data:
            values = Reference(ws, min_col=i, min_row=7, max_row=19)
            series = Series(values, xvalues, title_from_data=True)
            c1.series.append(series)

        serie_max = c1.series[0]
        serie_max.marker.symbol = "diamond"
        serie_max.marker.graphicalProperties.solidFill = "1660a7"
        serie_max.marker.graphicalProperties.line.solidFill = "1660a7"
        serie_max.graphicalProperties.line.solidFill = "1660a7"

        serie_min = c1.series[1]
        serie_min.marker.symbol = "triangle"
        serie_min.marker.graphicalProperties.solidFill = "cd0c18"
        serie_min.marker.graphicalProperties.line.solidFill = "cd0c18"
        serie_min.graphicalProperties.line.solidFill = "cd0c18"

        serie_pro = c1.series[2]
        serie_pro.marker.symbol = "square"
        serie_pro.marker.graphicalProperties.solidFill = "32cd32"
        serie_pro.marker.graphicalProperties.line.solidFill = "32cd32"
        serie_pro.graphicalProperties.line.solidFill = "32cd32"

        serie_pro = c1.series[3]
        serie_pro.marker.symbol = "x"
        serie_pro.marker.graphicalProperties.solidFill = "7d60a0"
        serie_pro.marker.graphicalProperties.line.solidFill = "7d60a0"
        serie_pro.graphicalProperties.line.solidFill = "7d60a0"

        cats = Reference(ws, min_col=1, min_row=8, max_row=19)
        c1.set_categories(cats)
        c1.legend.position = "b"
        ws.add_chart(c1, "A20")
