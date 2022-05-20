# -*- coding: utf-8 -*-

# Modelos
import csv

# Funciones de django
from datetime import datetime

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.drawing.image import Image

# librerias para manejar los archivos EXCEL
from openpyxl.styles import Font

from djangomain.settings import BASE_DIR
from estacion.models import Estacion

# Funcionalidades Comunes
from reportes_v2.functions import (
    get_data_graph,
    get_elemento_data_json,
    get_error_valor,
    get_layout_grafico,
    get_layout_grafico_agua,
    get_layout_grafico_viento,
    get_radar_chart,
    get_segundo_eje,
)
from reportes_v2.models import (
    ConsultaGenericaFecha,
    ConsultaGenericaFechaHora,
    ConsultaGenericaFechaHoraGrafico,
)

__frecuencia__dict = {
    "subhorario-crudo": "Dato crudo",
    "subhorario-validado": "Dato validado",
    "horario": "Horario",
    "diario": "Diario",
    "mensual": "Mensual",
}


# consultar datos crudos, validados, horarios, diarios y mensuales
def get_datos_graficar(estacion, variable, inicio, fin, frecuencia, profundidad):
    titulo = estacion.est_codigo + " - " + variable.var_nombre
    titulo_grafico = titulo + "<br>(" + __frecuencia__dict[frecuencia] + ")"

    if frecuencia == "subhorario-crudo":
        datos = consulta_crudos_graficar(
            estacion.est_id, variable, inicio, fin, profundidad
        )
    elif frecuencia == "subhorario-validado":
        datos = consulta_validados_graficar(
            estacion.est_id, variable, inicio, fin, profundidad
        )
    elif frecuencia == "horario":
        datos = consulta_horario_graficar(
            estacion.est_id, variable, inicio, fin, profundidad
        )
    elif frecuencia == "diario":
        datos = consulta_diario_graficar(
            estacion.est_id, variable, inicio, fin, profundidad
        )
    else:
        datos = consulta_mensual_graficar(
            estacion.est_id, variable, inicio, fin, profundidad
        )

    titulo_yaxis = variable.var_nombre + " (" + variable.uni_id.uni_sigla + ")"
    hay_datos = False

    if "fecha" in datos:
        if len(datos["fecha"]) > 0:
            hay_datos = True
    else:
        hay_datos = True

    if hay_datos:
        response = True
        if variable.var_id == 1:
            layout = get_layout_grafico(titulo_grafico, titulo_yaxis, inicio, fin)
            data_valor = get_elemento_data_json(
                "bar", datos["fecha"], datos["valor"], "Suma", "#1660A7"
            )
            data = [data_valor]
            if (
                frecuencia == "horario"
                or frecuencia == "diario"
                or frecuencia == "mensual"
            ):
                data_error = get_error_valor(
                    datos["fecha_error"],
                    datos["valor_error"],
                    datos["text_error"],
                    variable,
                )
                data.append(data_error)
        elif variable.var_id == 4 or variable.var_id == 5:
            layout = get_layout_grafico_viento(titulo_grafico)
            data = []
            for item in datos.values():
                data_graph = get_radar_chart(
                    item["datos"], item["color"], item["nombre"], item["hovertext"]
                )
                data.append(data_graph)
        else:
            layout = get_layout_grafico(titulo_grafico, titulo_yaxis, inicio, fin)
            data_valor = get_elemento_data_json(
                "scatter", datos["fecha"], datos["valor"], "Valor", "#1660A7"
            )
            data = [data_valor]
            if (
                frecuencia == "horario"
                or frecuencia == "diario"
                or frecuencia == "mensual"
            ):
                data_error = get_error_valor(
                    datos["fecha_error"],
                    datos["valor_error"],
                    datos["text_error"],
                    variable,
                )
                data.append(data_error)

        grafico = dict(
            data=data,
            layout=layout,
        )
    else:
        response = False
        grafico = {}

    data = {"response": response, "grafico": grafico}

    return data


def get_datos_exportar(estacion, variable, inicio, fin, frecuencia, profundidad):
    nombre_archivo = estacion.est_codigo + "-" + variable.var_nombre
    if profundidad:
        nombre_archivo = nombre_archivo + " a " + str(profundidad / 100.0) + "[m]"
    nombre_archivo = nombre_archivo + " " + __frecuencia__dict[frecuencia]
    nombre_archivo = nombre_archivo.replace(" ", "_")

    if frecuencia == "subhorario-crudo":
        if variable.var_id == 4 or variable.var_id == 5:
            datos = consulta_crudos_viento(estacion.est_id, inicio, fin)
        else:
            datos = consulta_crudos(estacion.est_id, variable, inicio, fin, profundidad)
        formato_fecha = "yyyy/mm/dd hh:mm:ss"
    elif frecuencia == "subhorario-validado":
        if variable.var_id == 4 or variable.var_id == 5:
            datos = consulta_validados_viento(estacion.est_id, inicio, fin)
        else:
            datos = consulta_validados(
                estacion.est_id, variable, inicio, fin, profundidad
            )
        formato_fecha = "yyyy/mm/dd hh:mm:ss"
    elif frecuencia == "horario":
        if variable.var_id == 4 or variable.var_id == 5:
            datos = consulta_horario_viento(estacion.est_id, inicio, fin)
        else:
            datos = consulta_horario(
                estacion.est_id, variable, inicio, fin, profundidad
            )
        formato_fecha = "yyyy/mm/dd hh:mm:ss"
    elif frecuencia == "diario":
        if variable.var_id == 4 or variable.var_id == 5:
            datos = consulta_diario_viento(estacion.est_id, inicio, fin)
        else:
            datos = consulta_diario(estacion.est_id, variable, inicio, fin, profundidad)
        formato_fecha = "yyyy/mm/dd"
    else:
        if variable.var_id == 4 or variable.var_id == 5:
            datos = consulta_mensual_viento(estacion.est_id, inicio, fin)
        else:
            datos = consulta_mensual(
                estacion.est_id, variable, inicio, fin, profundidad
            )
        formato_fecha = "yyyy/mm"

    return datos, nombre_archivo, formato_fecha


def export_csv(estacion, variable, inicio, fin, frecuencia, profundidad):
    datos, nombre_archivo, formato_fecha = get_datos_exportar(
        estacion, variable, inicio, fin, frecuencia, profundidad
    )
    for fila in datos:
        hay_datos = True
        break
    if not hay_datos:
        return None

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        'attachment; filename="' + nombre_archivo + '.csv"'
    )
    writer = csv.writer(response)
    if frecuencia == "horario" or frecuencia == "diario" or frecuencia == "mensual":
        if variable.var_id == 4 or variable.var_id == 5:
            writer.writerow(
                [
                    "Fecha",
                    "Velocidad (" + variable.uni_id.uni_sigla + ")",
                    "Dirección",
                    "Punto Cardinal",
                    "Predomninancia",
                    "Porcentaje",
                ]
            )
        else:
            writer.writerow(
                ["Fecha", "Valor (" + variable.uni_id.uni_sigla + ")", "Porcentaje"]
            )
    else:
        if variable.var_id == 4 or variable.var_id == 5:
            writer.writerow(
                ["Fecha", "Velocidad (" + variable.uni_id.uni_sigla + ")", "Dirección"]
            )
        else:
            writer.writerow(["Fecha", "Valor (" + variable.uni_id.uni_sigla + ")"])
    formato_fecha = (
        formato_fecha.replace("yyyy", "%Y").replace("/mm", "/%m").replace("dd", "%d")
    )
    formato_fecha = (
        formato_fecha.replace("hh", "%H").replace(":mm", ":%M").replace("ss", "%S")
    )

    if frecuencia == "horario" or frecuencia == "diario" or frecuencia == "mensual":
        if variable.var_id == 4 or variable.var_id == 5:
            for fila in datos:
                writer.writerow(
                    [
                        fila.fecha.strftime(formato_fecha),
                        fila.valor,
                        fila.direccion,
                        fila.categoria,
                        fila.predominancia,
                        fila.porcentaje,
                    ]
                )
        else:
            for fila in datos:
                writer.writerow(
                    [fila.fecha.strftime(formato_fecha), fila.valor, fila.porcentaje]
                )
    else:
        if variable.var_id == 4 or variable.var_id == 5:
            for fila in datos:
                writer.writerow(
                    [fila.fecha.strftime(formato_fecha), fila.valor, fila.direccion]
                )
        else:
            for fila in datos:
                writer.writerow([fila.fecha.strftime(formato_fecha), fila.valor])

    return response


def export_excel(estacion, variable, inicio, fin, frecuencia, profundidad):
    datos, nombre_archivo, formato_fecha = get_datos_exportar(
        estacion, variable, inicio, fin, frecuencia, profundidad
    )
    for fila in datos:
        hay_datos = True
        break
    if not hay_datos:
        return None

    # ruta de la imagen
    ruta_fonag = str(BASE_DIR) + "/media/logo_fonag.jpg"
    ruta_imhea = str(BASE_DIR) + "/media/imhea_logo2.jpg"
    img_fonag = Image(ruta_fonag)
    img_imhea = Image(ruta_imhea)
    # estilo de negrita
    font_bold = Font(bold=True)
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    ws.add_image(img_fonag, "A1")
    ws.add_image(img_imhea, "G1")

    ws["B4"] = "Reporte de Datos Hidrometerológicos"
    ws["B4"].font = font_bold

    ws.merge_cells("B4:F4")
    ws["A7"] = "Estación"
    ws["A7"].font = font_bold
    ws["B7"] = estacion.est_codigo
    ws["C7"] = estacion.est_nombre
    ws.merge_cells("C4:E4")
    ws["F7"] = "Variable"
    ws["F7"].font = font_bold
    ws["G7"] = variable.var_nombre
    ws["B9"] = "Coordenadas Geográficas"
    ws["B9"].font = font_bold
    ws.merge_cells("B6:G6")
    ws["A10"] = "Latitud"
    ws["A10"].font = font_bold
    ws["B10"] = estacion.est_latitud
    ws["F10"] = "Longitud"
    ws["F10"].font = font_bold
    ws["G10"] = estacion.est_longitud

    # Creamos los encabezados desde la celda B9 hasta la E9
    ws["A12"] = "Fecha"
    ws["B12"] = "Valor"
    if frecuencia == "horario" or frecuencia == "diario" or frecuencia == "mensual":
        if variable.var_id == 4 or variable.var_id == 5:
            ws["B12"] = "Velocidad"
            ws["C12"] = "Dirección"
            ws["D12"] = "Punto Cardinal"
            ws["E12"] = "Predominancia"
            ws["F12"] = "Porcentaje"
        else:
            ws["C12"] = "Porcentaje"
    else:
        if variable.var_id == 4 or variable.var_id == 5:
            ws["B12"] = "Velocidad"
            ws["C12"] = "Direccion"

    cont = 13

    for fila in datos:
        ws.cell(row=cont, column=1).value = fila.fecha
        ws.cell(row=cont, column=2).value = fila.valor
        if frecuencia == "horario" or frecuencia == "diario" or frecuencia == "mensual":
            if variable.var_id == 4 or variable.var_id == 5:
                ws.cell(row=cont, column=3).value = fila.direccion
                ws.cell(row=cont, column=4).value = fila.categoria
                ws.cell(row=cont, column=5).value = fila.predominancia
                ws.cell(row=cont, column=6).value = fila.porcentaje
            else:
                ws.cell(row=cont, column=3).value = fila.porcentaje
        else:
            if variable.var_id == 4 or variable.var_id == 5:
                ws.cell(row=cont, column=3).value = fila.direccion
        cont = cont + 1
    # última fila de datos
    final = 13 + len(datos)

    # grafico
    if variable.var_id == 1:
        chart_pre = BarChart()
        chart_pre.type = "col"
        chart_pre.style = 11
        chart_pre.title = variable.var_nombre
        chart_pre.y_axis.title = variable.var_nombre
        chart_pre.x_axis.title = "Tiempo"

        data = Reference(ws, min_col=2, min_row=12, max_row=final - 1, max_col=2)
        cats = Reference(ws, min_col=1, min_row=13, max_row=final - 1)
        chart_pre.add_data(data, titles_from_data=True)
        chart_pre.set_categories(cats)

        """labels = Reference(ws, min_col=1, min_row=13, max_row=final - 1)
        data = Reference(ws, min_col=2, min_row=13, max_row=final - 1)
        
        chart_pre.series = (Series(data),)
        chart_pre.title = variable.var_nombre"""
        chart_pre.shape = 4

        # Change bar filling and line color
        s = chart_pre.series[0]
        s.graphicalProperties.line.solidFill = "1660a7"
        s.graphicalProperties.solidFill = "1660a7"

        ws.add_chart(chart_pre, "H9")
    elif variable.var_id == 4 or variable.var_id == 5:
        pass
    else:

        chart = LineChart()
        chart.title = variable.var_nombre
        chart.style = 12
        chart.x_axis.title = "Tiempo(dias)"
        chart.x_axis.number_format = "dd-mm-yyyy"
        chart.x_axis.majorTimeUnit = "days"
        chart.y_axis.title = variable.var_nombre

        data = Reference(ws, min_col=2, min_row=12, max_col=2, max_row=final - 1)
        chart.add_data(data, titles_from_data=True)
        s0 = chart.series[0]
        s0.graphicalProperties.line.solidFill = "32cd32"
        s0.graphicalProperties.line.width = 10

        dates = Reference(ws, min_col=1, min_row=13, max_row=final)
        chart.set_categories(dates)
        ws.add_chart(chart, "H9")

    nombre_archivo = nombre_archivo + str(".xlsx")

    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


# consulta de datos crudos de viento para exportar
def consulta_crudos_viento(estacion_id, inicio, fin):
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    query = "select * FROM reporte_crudos_subhorario_viento_export(%s, %s, %s);"

    consulta = ConsultaGenericaFechaHoraGrafico.objects.raw(
        query, [estacion_id, inicio, fin]
    )

    return consulta


def consulta_crudos(estacion_id, variable, inicio, fin, profundidad):
    modelo = str(variable.var_modelo)
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    if variable.var_id == 1:
        query = "select * FROM reporte_crudos_subhorario_precipitacion(%s, %s, %s);"
    elif variable.var_id == 4 or variable.var_id == 5:
        query = "select * FROM reporte_crudos_subhorario_viento(%s, %s, %s);"
    else:
        query = "select * FROM reporte_crudos_subhorario_" + modelo + "(%s, %s, %s);"

    consulta = ConsultaGenericaFechaHoraGrafico.objects.raw(
        query, [estacion_id, inicio, fin]
    )
    return consulta


# consultar datos crudos para graficar
def consulta_crudos_graficar(estacion_id, variable, inicio, fin, profundidad):
    consulta = consulta_crudos(estacion_id, variable, inicio, fin, profundidad)
    fecha = []
    valor = []

    if variable.var_id == 4 or variable.var_id == 5:
        pass
    else:
        for fila in consulta:
            # if fila.valor is not None:
            #     resultado.append([fila.fecha, fila.valor])
            if fila.salto is True:
                fecha_intermedia = fecha_anterior + (fila.fecha - fecha_anterior) / 2
                fecha.append(fecha_intermedia)
                valor.append(None)
            fecha.append(fila.fecha)
            fecha_anterior = fila.fecha
            valor.append(fila.valor)

    if variable.var_id == 4 or variable.var_id == 5:
        # datos = {'grupo1': grupo1, 'grupo2': grupo2, 'grupo3': grupo3, 'grupo4': grupo4}
        datos = get_matriz_viento(consulta, variable)
    else:
        datos = {"fecha": fecha, "valor": valor}

    return datos


# procesar datos de viento
def get_matriz_viento(consulta, variable):
    grupo1 = []
    grupo2 = []
    grupo3 = []
    grupo4 = []
    lim_sup = []
    lim_min = []
    hovertext = dict(grupo1=[], grupo2=[], grupo3=[], grupo4=[])
    for fila in consulta:
        if fila.grupo == 1:
            grupo1.append(fila.por_acum)
            hovertext["grupo1"].append(fila.porcentaje)
            if fila.categoria == 1:
                lim_sup.append(fila.maximo)
                lim_min.append(fila.minimo)
        elif fila.grupo == 2:
            grupo2.append(fila.por_acum)
            hovertext["grupo2"].append(fila.porcentaje)
            if fila.categoria == 1:
                lim_sup.append(fila.maximo)
                lim_min.append(fila.minimo)
        elif fila.grupo == 3:
            grupo3.append(fila.por_acum)
            hovertext["grupo3"].append(fila.porcentaje)
            if fila.categoria == 1:
                lim_sup.append(fila.maximo)
                lim_min.append(fila.minimo)
        elif fila.grupo == 4:
            grupo4.append(fila.por_acum)
            hovertext["grupo4"].append(fila.porcentaje)
            if fila.categoria == 1:
                lim_sup.append(fila.maximo)
                lim_min.append(fila.minimo)

    datos = dict(
        grupo4=dict(
            datos=grupo4,
            hovertext=hovertext.get("grupo4"),
            color="#0080ff",
            nombre="> " + str(lim_min[3]) + " " + variable.uni_id.uni_sigla,
        ),
        grupo3=dict(
            datos=grupo3,
            hovertext=hovertext.get("grupo3"),
            color="#3399ff",
            nombre=str(lim_min[2])
            + " - "
            + str(lim_sup[2])
            + " "
            + variable.uni_id.uni_sigla,
        ),
        grupo2=dict(
            datos=grupo2,
            hovertext=hovertext.get("grupo2"),
            color="#66b3ff",
            nombre=str(lim_min[1])
            + " - "
            + str(lim_sup[1])
            + " "
            + variable.uni_id.uni_sigla,
        ),
        grupo1=dict(
            datos=grupo1,
            hovertext=hovertext.get("grupo1"),
            color="#99ccff",
            nombre="0 - " + str(lim_sup[0]) + " " + variable.uni_id.uni_sigla,
        ),
    )
    return datos


# consulta de datos validados de viento para exportar
def consulta_validados_viento(estacion_id, inicio, fin):
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    sql = """
        SELECT row_number() OVER (ORDER BY v.fecha ASC) as fila, v.id, v.fecha, v.valor, v.direccion
        FROM validacion_viento v
        WHERE estacion_id = %s
        AND v.fecha >= '%%fecha_inicio%%' AND v.fecha <= '%%fecha_fin%%'
    """

    sql = sql.replace("%%fecha_inicio%%", str(inicio)).replace(
        "%%fecha_fin%%", str(fin)
    )
    consulta = ConsultaGenericaFechaHoraGrafico.objects.raw(sql, [str(estacion_id)])

    return consulta


# datos validados
def consulta_validados(estacion_id, variable, inicio, fin, profundidad):
    modelo = str(variable.var_modelo).lower()
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    if variable.var_id == 1:
        query = "select * FROM reporte_validados_subhorario_precipitacion(%s, %s, %s);"
    elif variable.var_id == 4 or variable.var_id == 5:
        query = "select * FROM reporte_validados_subhorario_viento(%s, %s, %s);"
    else:
        query = "select * FROM reporte_validados_subhorario_" + modelo + "(%s, %s, %s);"

    consulta = ConsultaGenericaFechaHoraGrafico.objects.raw(
        query, [estacion_id, inicio, fin]
    )
    return consulta


# consulta de datos validados para graficar
def consulta_validados_graficar(estacion_id, variable, inicio, fin, profundidad):
    consulta = consulta_validados(estacion_id, variable, inicio, fin, profundidad)
    fecha = []
    valor = []
    if variable.var_id == 4 or variable.var_id == 5:
        pass
    else:
        for fila in consulta:
            # if fila.valor is not None:
            #     resultado.append([fila.fecha, fila.valor])
            if fila.salto is True:
                fecha_intermedia = fecha_anterior + (fila.fecha - fecha_anterior) / 2
                fecha.append(fecha_intermedia)
                valor.append(None)
            fecha.append(fila.fecha)
            fecha_anterior = fila.fecha
            valor.append(fila.valor)

    if variable.var_id == 4 or variable.var_id == 5:
        # datos = {'grupo1': grupo1, 'grupo2': grupo2, 'grupo3': grupo3, 'grupo4': grupo4}
        datos = get_matriz_viento(consulta, variable)
    else:
        datos = {"fecha": fecha, "valor": valor}
    return datos


# consulta de datos horarios de viento
def consulta_horario_viento(estacion_id, inicio, fin):
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    query = "select * FROM reporte_validados_horario_viento(%s, %s, %s);"
    consulta = ConsultaGenericaFechaHora.objects.raw(query, [estacion_id, inicio, fin])
    return consulta


# consulta del modelo de datos horarios
def consulta_horario(estacion_id, variable, inicio, fin, profundidad):
    modelo = str(variable.var_modelo)
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    if variable.var_id == 1:
        query = "select * FROM reporte_validados_horario_precipitacion(%s, %s, %s);"
    elif variable.var_id == 4 or variable.var_id == 5:
        query = "select * FROM reporte_validados_subhorario_viento(%s, %s, %s);"
    else:
        query = "select * FROM reporte_validados_horario_" + modelo + "(%s, %s, %s);"

    consulta = ConsultaGenericaFechaHoraGrafico.objects.raw(
        query, [estacion_id, inicio, fin]
    )
    return consulta


# consulta de datos horarios para graficar
def consulta_horario_graficar(estacion_id, variable, inicio, fin, profundidad):
    consulta = consulta_horario(estacion_id, variable, inicio, fin, profundidad)
    fecha = []
    valor = []
    porcentaje = []
    fecha_error = []
    valor_error = []
    text_error = []
    if variable.var_id == 4 or variable.var_id == 5:
        pass
    else:
        for fila in consulta:
            fecha.append(fila.fecha)
            valor.append(fila.valor)
            porcentaje.append(fila.porcentaje)
            if fila.error:
                fecha_error.append(fila.fecha)
                valor_error.append(fila.valor)
                text_error.append(str(fila.porcentaje) + "%")
    if variable.var_id == 4 or variable.var_id == 5:
        datos = get_matriz_viento(consulta, variable)
    else:
        datos = {
            "fecha": fecha,
            "valor": valor,
            "fecha_error": fecha_error,
            "valor_error": valor_error,
            "text_error": text_error,
        }

    return datos


# consulta de datos horarios de viento
def consulta_diario_viento(estacion_id, inicio, fin):
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    query = "select * FROM reporte_validados_diario_viento(%s, %s, %s);"
    consulta = ConsultaGenericaFechaHora.objects.raw(query, [estacion_id, inicio, fin])
    return consulta


# consulta de datos diarios
def consulta_diario(estacion_id, variable, inicio, fin, profundidad):
    modelo = str(variable.var_modelo)
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    if variable.var_id == 1:
        query = "select * FROM reporte_validados_diario_precipitacion(%s, %s, %s);"
    elif variable.var_id == 4 or variable.var_id == 5:
        query = "select * FROM reporte_validados_subhorario_viento(%s, %s, %s);"
    else:
        query = "select * FROM reporte_validados_diario_" + modelo + "(%s, %s, %s);"

    consulta = ConsultaGenericaFechaHoraGrafico.objects.raw(
        query, [estacion_id, inicio, fin]
    )
    return consulta


# consulta de datos diarios para graficar
def consulta_diario_graficar(estacion_id, variable, inicio, fin, profundidad):
    consulta = consulta_diario(estacion_id, variable, inicio, fin, profundidad)
    fecha = []
    valor = []
    porcentaje = []
    fecha_error = []
    valor_error = []
    text_error = []

    if variable.var_id == 4 or variable.var_id == 5:
        pass
    else:
        for fila in consulta:
            fecha.append(fila.fecha)
            valor.append(fila.valor)
            porcentaje.append(fila.porcentaje)
            if fila.error:
                fecha_error.append(fila.fecha)
                valor_error.append(fila.valor)
                text_error.append(str(fila.porcentaje) + "%")
    if variable.var_id == 4 or variable.var_id == 5:
        datos = get_matriz_viento(consulta, variable)
    else:
        datos = {
            "fecha": fecha,
            "valor": valor,
            "fecha_error": fecha_error,
            "valor_error": valor_error,
            "text_error": text_error,
        }
    return datos


# consulta de datos horarios de viento
def consulta_mensual_viento(estacion_id, inicio, fin):
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    query = "select * FROM reporte_validados_mensual_viento(%s, %s, %s);"
    consulta = ConsultaGenericaFechaHora.objects.raw(query, [estacion_id, inicio, fin])
    return consulta


def consulta_mensual(estacion_id, variable, inicio, fin, profundidad):
    modelo = str(variable.var_modelo)
    if inicio:
        inicio = datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    if variable.var_id == 1:
        query = "select * FROM reporte_validados_mensual_precipitacion(%s, %s, %s);"
    elif variable.var_id == 4 or variable.var_id == 5:
        query = "select * FROM reporte_validados_subhorario_viento(%s, %s, %s);"
    else:
        query = "select * FROM reporte_validados_mensual_" + modelo + "(%s, %s, %s);"

    consulta = ConsultaGenericaFechaHoraGrafico.objects.raw(
        query, [estacion_id, inicio, fin]
    )
    return consulta


def consulta_mensual_graficar(estacion_id, variable, inicio, fin, profundidad):
    consulta = consulta_mensual(estacion_id, variable, inicio, fin, profundidad)
    fecha = []
    valor = []
    porcentaje = []
    fecha_error = []
    valor_error = []
    text_error = []
    if variable.var_id == 4 or variable.var_id == 5:
        pass
    else:
        for fila in consulta:
            fecha.append(fila.fecha)
            valor.append(fila.valor)
            porcentaje.append(fila.porcentaje)
            if fila.error:
                fecha_error.append(fila.fecha)
                valor_error.append(fila.valor)
                text_error.append(str(fila.porcentaje) + "%")

    if variable.var_id == 4 or variable.var_id == 5:
        datos = get_matriz_viento(consulta, variable)
    else:
        datos = {
            "fecha": fecha,
            "valor": valor,
            "fecha_error": fecha_error,
            "valor_error": valor_error,
            "text_error": text_error,
        }
    return datos
