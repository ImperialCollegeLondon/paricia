# -*- coding: utf-8 -*-

################################################################################################
# Plataforma para la Iniciativa Regional de Monitoreo Hidrológico de Ecosistemas Andinos (iMHEA)
# basada en los desarrollos realizados por:
#     1) FONDO PARA LA PROTECCIÓN DEL AGUA (FONAG), Ecuador.
#         Contacto: info@fonag.org.ec
#     2) EMPRESA PÚBLICA METROPOLITANA DE AGUA POTABLE Y SANEAMIENTO DE QUITO (EPMAPS), Ecuador.
#         Contacto: paramh2o@aguaquito.gob.ec
#
#  IMPORTANTE: Mantener o incluir esta cabecera con la mención de las instituciones creadoras,
#              ya sea en uso total o parcial del código.


import csv
import datetime
import decimal
import os
import random
from calendar import monthrange

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import plotly.graph_objs as go
import plotly.offline as opy
import xlwt
from django.conf import settings
from django.db import connection

# usados en: datos_horarios_json
from django.db.models import Avg, Count, Max, Min, Sum
from django.db.models.functions import (
    ExtractDay,
    ExtractHour,
    ExtractMonth,
    ExtractYear,
)
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, NamedStyle, Side, borders
from plotly.subplots import make_subplots

from cruce.models import Cruce
from diario.models import *
from horario.models import *
from mensual.models import *
from variable.models import Unidad, Variable

from .models import *

__frecuencia__ = {
    "Subhorario crudo": "",
    "Subhorario validado": "",
    "Horario": "Hora",
    "Diario": "Día",
    "Mensual": "Mes",
}


def max_min(lista):
    max = None
    min = None
    for n in lista:
        if n is None:
            continue

        if max is None:
            max = n
            min = n
            continue

        if n > max:
            max = n

        if n < min:
            min = n

    return max, min


### funciones necesarias para la vista reporte por periodo con gaps


def dataScala(escalaTemp, codEst, codVar, fi, ff):
    et = int(escalaTemp)
    varLine = [2, 3, 7, 8, 10, 11, 13]
    varBar = [1]
    framed = pd.DataFrame()
    print("Data Scala ", type(codVar), codVar)
    print("fecha inici", str(fi), " - fecha de fin ", ff)
    variable = Variable.objects.get(pk=codVar)
    vacios = variable.vacios
    if et == 2:
        sub = "Horario"
        dataD = (
            globals()["Var" + str(codVar) + "Horario"]
            .objects.filter(
                estacion_id=codEst, fecha__gte=fi, fecha__lte=ff, vacios__lt=vacios
            )
            .order_by("fecha")
        )
        if codVar in varBar:
            dataD = dataD.values_list("fecha", "valor")
            # framed = pd.DataFrame.from_records(dataD)
            # ldataD = list(dataD)
            framed = pd.DataFrame(list(dataD), columns=["fecha", "valor"])
        elif codVar in varLine:
            dataD = dataD.values("fecha", "valor", "max_abs", "min_abs")
            # framed = pd.DataFrame.from_records(dataD)
            framed = pd.DataFrame(list(dataD))
    elif et == 3:
        sub = "Diario"
        dataD = (
            globals()["Var" + str(codVar) + "Diario"]
            .objects.filter(
                estacion_id=codEst, fecha__gte=fi, fecha__lte=ff, vacios__lt=vacios
            )
            .order_by("fecha")
        )
        if codVar in varBar:
            dataD = dataD.values("fecha", "valor")
            # framed = pd.DataFrame.from_records(dataD)
            framed = pd.DataFrame(list(dataD))
        elif codVar in varLine:
            dataD = dataD.values("fecha", "valor", "max_del_prom", "min_del_prom")
            # framed = pd.DataFrame.from_records(dataD)
            framed = pd.DataFrame(list(dataD))
    elif et == 4:
        sub = "Mensual"
        dataD = (
            globals()["Var" + str(codVar) + "Mensual"]
            .objects.filter(
                estacion_id=codEst, fecha__gte=fi, fecha__lte=ff, vacios__lt=vacios
            )
            .order_by("fecha")
        )
        if codVar in varBar:
            dataD = dataD.values("fecha", "valor")
            # framed = pd.DataFrame.from_records(dataD)
            framed = pd.DataFrame(list(dataD))
        elif codVar in varLine:
            dataD = dataD.values("fecha", "valor", "max_del_prom", "min_del_prom")
            # framed = pd.DataFrame.from_records(dataD)
            framed = pd.DataFrame(list(dataD))
    # print(framed)
    return framed, sub


# compara tres estaciones y un avariable
def comparar(form):
    estacion01 = form.cleaned_data["estacion01"]
    estacion02 = form.cleaned_data["estacion02"]
    estacion03 = form.cleaned_data["estacion03"]
    variable = form.cleaned_data["variable"]
    fi = form.cleaned_data["inicio"]
    ff = form.cleaned_data["fin"]
    frecuencia = form.cleaned_data["frecuencia"]

    data1, sub = dataScala(frecuencia, estacion01.est_id, variable.var_id, fi, ff)
    data2, sub = dataScala(frecuencia, estacion02.est_id, variable.var_id, fi, ff)
    data3, sub = dataScala(frecuencia, estacion03.est_id, variable.var_id, fi, ff)
    if data1.empty | data2.empty | data3.empty:
        return '<div class= "alert alert-warning" role="alert">Una de la estaciones no tiene datos para comparar</div>'

    trace0 = trace_graph(variable, estacion01, data1["fecha"], data1["valor"])
    trace1 = trace_graph(variable, estacion02, data2["fecha"], data2["valor"])
    trace2 = trace_graph(variable, estacion03, data3["fecha"], data3["valor"])
    data = [trace0, trace1, trace2]
    layout = go.Layout(
        autosize=False,
        # width=,
        # height=100,
        title="Comparación de Estaciones",
        yaxis=dict(title=variable.var_nombre + " (" + variable.uni_id.uni_sigla + ")"),
        xaxis=dict(
            rangeselector=dict(
                buttons=list(
                    [
                        dict(count=1, label="1d", step="day", stepmode="todate"),
                        dict(count=1, label="1m", step="month", stepmode="backward"),
                        dict(count=6, label="6m", step="month", stepmode="backward"),
                        dict(count=1, label="1y", step="year", stepmode="backward"),
                        dict(step="all"),
                    ]
                )
            ),
            rangeslider=dict(),
            type="date",
        ),
    )
    figure = go.Figure(data=data, layout=layout)
    figure.update_yaxes(automargin=True)
    figure.update_xaxes(automargin=True)
    div = opy.plot(figure, auto_open=False, output_type="div")
    return div


def comparar_variable(form):
    estacion01 = form.cleaned_data["estacion01"]
    estacion02 = form.cleaned_data["estacion02"]
    variable01 = form.cleaned_data["variable01"]
    variable02 = form.cleaned_data["variable02"]
    fi = form.cleaned_data["inicio"]
    ff = form.cleaned_data["fin"]
    frecuencia = form.cleaned_data["frecuencia"]
    data1, sub = dataScala(frecuencia, estacion01.est_id, variable01.var_id, fi, ff)
    data2, sub = dataScala(frecuencia, estacion02.est_id, variable02.var_id, fi, ff)
    if data1.empty | data2.empty:
        return '<div class= "alert alert-warning" role="alert">Una de la estaciones no tiene datos para comparar</div>'
    trace0 = trace_graph(variable01, estacion01, data1["fecha"], data1["valor"])
    trace1 = trace_graph(variable02, estacion02, data2["fecha"], data2["valor"])

    data = [trace0, trace1]
    layout = go.Layout(
        autosize=False,
        # width=,
        # height=100,
        title="Comparación de Variables",
        yaxis=dict(
            title=variable01.var_nombre + " (" + variable01.uni_id.uni_sigla + ")"
        ),
        yaxis2=dict(
            title=variable02.var_nombre + " (" + variable02.uni_id.uni_sigla + ")",
            titlefont=dict(color="rgb(148, 103, 189)"),
            tickfont=dict(color="rgb(148, 103, 189)"),
            overlaying="y",
            side="right",
        ),
        xaxis=dict(
            rangeselector=dict(
                buttons=list(
                    [
                        dict(count=1, label="1d", step="day", stepmode="todate"),
                        dict(count=1, label="1m", step="month", stepmode="backward"),
                        dict(count=6, label="6m", step="month", stepmode="backward"),
                        dict(count=1, label="1y", step="year", stepmode="backward"),
                        dict(step="all"),
                    ]
                )
            ),
            rangeslider=dict(),
            type="date",
        ),
    )
    figure = go.Figure(data=data, layout=layout)
    figure.update_yaxes(automargin=True)
    figure.update_xaxes(automargin=True)
    div = opy.plot(figure, auto_open=False, output_type="div")
    return div


def trace_graph(variable, estacion, tiempo, valor):
    if variable.var_id == 1:
        trace = go.Bar(
            x=tiempo,
            y=valor,
            name=estacion.est_codigo,
        )
    else:
        trace = go.Scatter(
            x=tiempo, y=valor, name=estacion.est_codigo, mode="lines", yaxis="y2"
        )
    return trace


def consulta_crudos(estacion_id, variable_id, inicio, fin, *args, **kwargs):
    profundidad = kwargs.get("profundidad", False)
    if inicio:
        inicio = datetime.datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime.datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    sql = """
    SELECT m.id, m.fecha, m.valor
    FROM medicion_var101medicion m
    WHERE m.estacion_id = %s
    """
    if profundidad:
        sql += "        AND m.profundidad = " + str(profundidad)
    if inicio:
        sql += "        AND m.fecha>='" + str(inicio) + "'"
    if fin:
        sql += "        AND m.fecha<='" + str(fin) + "'"
    sql = (
        sql
        + """
    ORDER BY m.fecha ASC, m.id ASC;
    """
    )
    sql = sql.replace("var101", "var" + str(variable_id))
    consulta = ConsultaGenericaFechaHora.objects.raw(
        sql,
        [
            estacion_id,
        ],
    )
    return consulta


def consulta_crudos__columnas(estacion_id, variable_id, inicio, fin, *args, **kwargs):
    profundidad = kwargs.get("profundidad", False)
    consulta = consulta_crudos(estacion_id, variable_id, inicio, fin, profundidad)
    fecha = []
    valor = []
    for fila in consulta:
        fecha.append(fila.fecha)
        valor.append(fila.valor)
    datos = {"fecha": fecha, "valor": valor}
    return datos


def consulta_crudos_saltos(estacion_id, variable_id, inicio, fin, *args, **kwargs):
    profundidad = kwargs.get("profundidad", False)
    if inicio:
        inicio = datetime.datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime.datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    sql = """
    WITH
    estacion AS (SELECT * FROM estacion_estacion e WHERE e.est_id = %s),
    variable AS (SELECT * FROM variable_variable v WHERE v.var_id = %s),
    medicion AS (
        SELECT m.id, m.fecha, m.valor
        FROM medicion_var101medicion m
        WHERE m.estacion_id = (SELECT e.est_id FROM estacion e)
    """
    if profundidad:
        sql = sql + "        AND m.profundidad = " + str(profundidad)
    if inicio:
        sql = sql + "        AND m.fecha>='" + str(inicio) + "'"
    if fin:
        sql = sql + "        AND m.fecha<='" + str(fin) + "'"
    sql = (
        sql
        + """
        ORDER BY m.fecha ASC
    ),
    fechas AS (
        SELECT 
            m.id, m.fecha, m.valor, 
            EXTRACT(EPOCH FROM m.fecha - lag(m.fecha) OVER (ORDER BY m.fecha ASC))/60  as lapso_tiempo,
            (SELECT f.fre_valor FROM frecuencia_frecuencia f
                    WHERE f.var_id_id = (SELECT v.var_id FROM variable v) 
                    AND f.est_id_id = (SELECT e.est_id FROM estacion e) AND f.fre_fecha_ini < m.fecha
                    ORDER BY f.fre_fecha_ini DESC LIMIT 1) AS periodo_esperado
        FROM medicion m ORDER BY m.fecha ASC, m.id ASC
    ),
    tabla AS (
        SELECT 
            f.id, 
            f.fecha, f.valor, 
            --f.lapso_tiempo, f.periodo_esperado, 
            CASE WHEN f.fecha = (SELECT m.fecha FROM medicion m ORDER BY m.fecha ASC LIMIT 1) THEN FALSE ELSE 
                CASE WHEN f.lapso_tiempo > (f.periodo_esperado * 1.05) THEN TRUE ELSE FALSE END
            END AS salto
        FROM fechas f ORDER BY f.fecha, f.id
    )
    SELECT t.id, t.fecha, t.valor, t.salto FROM tabla t ORDER BY t.fecha ASC, t.id ASC;    
    """
    )
    sql = sql.replace("var101", "var" + str(variable_id))
    consulta = ConsultaGenericaFechaHora_Saltos.objects.raw(
        sql, [estacion_id, variable_id]
    )
    return consulta


def consulta_crudos_saltos__lista(
    estacion_id, variable_id, inicio, fin, *args, **kwargs
):
    profundidad = kwargs.get("profundidad", False)
    if inicio:
        inicio = datetime.datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime.datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    sql = """
    WITH
    estacion AS (SELECT * FROM estacion_estacion e WHERE e.est_id = %s),
    variable AS (SELECT * FROM variable_variable v WHERE v.var_id = %s),
    medicion AS (
        SELECT m.id, m.fecha, m.valor
        FROM medicion_var101medicion m
        WHERE m.estacion_id = (SELECT e.est_id FROM estacion e)
    """
    if profundidad:
        sql = sql + "        AND m.profundidad = " + str(profundidad)
    if inicio:
        sql = sql + "        AND m.fecha>='" + str(inicio) + "'"
    if fin:
        sql = sql + "        AND m.fecha<='" + str(fin) + "'"
    sql = (
        sql
        + """
        ORDER BY m.fecha ASC
    ),
    fechas AS (
        SELECT 
            m.id, m.fecha, m.valor, 
            EXTRACT(EPOCH FROM m.fecha - lag(m.fecha) OVER (ORDER BY m.fecha ASC))/60  as lapso_tiempo,
            (SELECT f.fre_valor FROM frecuencia_frecuencia f
                    WHERE f.var_id_id = (SELECT v.var_id FROM variable v) 
                    AND f.est_id_id = (SELECT e.est_id FROM estacion e) AND f.fre_fecha_ini < m.fecha
                    ORDER BY f.fre_fecha_ini DESC LIMIT 1) AS periodo_esperado
        FROM medicion m ORDER BY m.fecha ASC, m.id ASC
    ),
    tabla AS (
        SELECT 
            f.id, 
            f.fecha, f.valor, 
            --f.lapso_tiempo, f.periodo_esperado, 
            CASE WHEN f.fecha = (SELECT m.fecha FROM medicion m ORDER BY m.fecha ASC LIMIT 1) THEN FALSE ELSE 
                CASE WHEN f.lapso_tiempo > (f.periodo_esperado * 1.05) THEN TRUE ELSE FALSE END
            END AS salto
        FROM fechas f ORDER BY f.fecha, f.id
    )
    --SELECT t.id, t.fecha, t.valor, t.salto FROM tabla t ORDER BY t.fecha ASC, t.id ASC;    
    SELECT t.fecha, t.valor, t.salto FROM tabla t ORDER BY t.fecha ASC, t.id ASC;
    """
    )
    sql = sql.replace("var101", "var" + str(variable_id))
    with connection.cursor() as cursor:
        cursor.execute(sql, [estacion_id, variable_id])
        consulta = cursor.fetchall()
    return consulta


def consulta_crudos_saltos__columnas(
    estacion_id, variable_id, inicio, fin, *args, **kwargs
):
    profundidad = kwargs.get("profundidad", False)
    consulta = consulta_crudos_saltos(
        estacion_id, variable_id, inicio, fin, profundidad=profundidad
    )
    fecha = []
    valor = []
    for fila in consulta:
        if fila.salto is True:
            fecha_intermedia = fecha_anterior + (fila.fecha - fecha_anterior) / 2
            fecha.append(fecha_intermedia)
            valor.append(None)
        fecha.append(fila.fecha)
        valor.append(fila.valor)
        fecha_anterior = fila.fecha
    datos = {"fecha": fecha, "valor": valor}
    return datos


def consulta_validados(estacion_id, variable_id, inicio, fin, *args, **kwargs):
    profundidad = kwargs.get("profundidad", False)
    if inicio:
        inicio = datetime.datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime.datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    sql = """
    SELECT v.id, v.fecha, v.valor
    FROM validacion_var101validado v
    WHERE v.estacion_id = %s
    """
    if profundidad:
        sql += "        AND v.profundidad = " + str(profundidad)
    if inicio:
        sql += "        AND v.fecha>='" + str(inicio) + "'"
    if fin:
        sql += "        AND v.fecha<='" + str(fin) + "'"
    sql = (
        sql
        + """
    ORDER BY v.fecha ASC, v.id ASC;
    """
    )
    sql = sql.replace("var101", "var" + str(variable_id))
    consulta = ConsultaGenericaFechaHora.objects.raw(
        sql,
        [
            estacion_id,
        ],
    )
    return consulta


def consulta_validados__columnas(
    estacion_id, variable_id, inicio, fin, *args, **kwargs
):
    profundidad = kwargs.get("profundidad", False)
    consulta = consulta_validados(estacion_id, variable_id, inicio, fin, profundidad)
    fecha = []
    valor = []
    for fila in consulta:
        fecha.append(fila.fecha)
        valor.append(fila.valor)
    datos = {"fecha": fecha, "valor": valor}
    return datos


def consulta_validados_saltos(estacion_id, variable_id, inicio, fin, *args, **kwargs):
    profundidad = kwargs.get("profundidad", False)
    if inicio:
        inicio = datetime.datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime.datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)

    sql = """
    WITH
    estacion AS (SELECT * FROM estacion_estacion e WHERE e.est_id = %s),
    variable AS (SELECT * FROM variable_variable v WHERE v.var_id = %s),
    validados AS (
        SELECT v.id, v.fecha, v.valor
        FROM validacion_var101validado v
        WHERE v.estacion_id = (SELECT e.est_id FROM estacion e)
    """
    if profundidad:
        sql += "        AND v.profundidad = " + str(profundidad)
    if inicio:
        sql += "        AND v.fecha>='" + str(inicio) + "'"
    if fin:
        sql += "        AND v.fecha<='" + str(fin) + "'"
    sql = (
        sql
        + """
        ORDER BY v.fecha ASC, v.id
    ),
    fechas AS (
        SELECT 
            v.id, v.fecha, v.valor, 
            EXTRACT(EPOCH FROM v.fecha - lag(v.fecha) OVER (ORDER BY v.fecha ASC))/60  as lapso_tiempo,
            (SELECT f.fre_valor FROM frecuencia_frecuencia f
                    WHERE f.var_id_id = (SELECT var.var_id FROM variable var) 
                    AND f.est_id_id = (SELECT e.est_id FROM estacion e) AND f.fre_fecha_ini < v.fecha
                    ORDER BY f.fre_fecha_ini DESC LIMIT 1) AS periodo_esperado
        FROM validados v ORDER BY v.fecha ASC
    ),
    tabla AS (
        SELECT 
            f.id, f.fecha, f.valor, 
            --f.lapso_tiempo, f.periodo_esperado, 
            CASE WHEN f.fecha = (SELECT v.fecha FROM validados v ORDER BY v.fecha ASC LIMIT 1) THEN FALSE ELSE
                CASE WHEN f.lapso_tiempo > (f.periodo_esperado * 1.05) THEN TRUE ELSE FALSE END
            END AS salto
        FROM fechas f ORDER BY f.fecha, f.id ASC
    )
    SELECT t.id, t.fecha, t.valor, t.salto FROM tabla t ORDER BY t.fecha ASC, t.id ASC;
    """
    )
    sql = sql.replace("var101", "var" + str(variable_id))
    consulta = ConsultaGenericaFechaHora_Saltos.objects.raw(
        sql, [estacion_id, variable_id]
    )
    return consulta


def consulta_validados_saltos__columnas(
    estacion_id, variable_id, inicio, fin, *args, **kwargs
):
    profundidad = kwargs.get("profundidad", False)
    consulta = consulta_validados_saltos(
        estacion_id, variable_id, inicio, fin, profundidad=profundidad
    )
    fecha = []
    valor = []
    for fila in consulta:
        # if fila.valor is not None:
        #     resultado.append([fila.fecha, fila.valor])
        if fila.salto is True:
            fecha_intermedia = fecha_anterior + (fila.fecha - fecha_anterior) / 2
            fecha.append(fecha_intermedia)
            valor.append(None)
        fecha.append(fila.fecha)
        valor.append(fila.valor)
        fecha_anterior = fila.fecha
    datos = {"fecha": fecha, "valor": valor}
    return datos


def consulta_horario(estacion_id, variable_id, inicio, fin, *args, **kwargs):
    profundidad = kwargs.get("profundidad", False)
    excluir_vacios = kwargs.get("excluir_vacios", True)

    if inicio:
        inicio = datetime.datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime.datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)
    sql = """
    WITH 
    variable AS (
        SELECT * FROM variable_variable v WHERE v.var_id = %%var_id%%
    ),
    consulta AS (
        SELECT 
            h.fecha 
            , h.valor 
            , h.vacios 
        FROM horario_var%%var_id%%horario h 
        WHERE h.estacion_id = %%est_id%% 
        %%filtro%%
    ),
    secuencia AS (
        SELECT * FROM generate_series( 
            (SELECT c.fecha FROM consulta c ORDER BY c.fecha ASC LIMIT 1 ),
            (SELECT c.fecha FROM consulta c ORDER BY c.fecha DESC LIMIT 1 ),
            '1 hour'::interval) fecha
    )
    SELECT s.fecha, c.valor, c.vacios
        FROM secuencia s LEFT JOIN consulta c USING (fecha) ORDER BY fecha ASC;    
    """
    filtro = ""
    if profundidad:
        filtro += "    AND h.profundidad = " + str(profundidad)
    if inicio:
        filtro += "    AND h.fecha >='" + str(inicio) + "'"
    if fin:
        filtro += "    AND h.fecha <='" + str(fin) + "'"
    if excluir_vacios:
        filtro += "\n        AND h.vacios < (SELECT v.vacios FROM variable v)"

    sql = sql.replace("%%est_id%%", str(estacion_id))
    sql = sql.replace("%%var_id%%", str(variable_id))
    sql = sql.replace("%%filtro%%", filtro)

    consulta = ConsultaReporteFechaHora.objects.raw(sql)
    return consulta


def consulta_horario__columnas(estacion_id, variable_id, inicio, fin, *args, **kwargs):
    profundidad = kwargs.get("profundidad", False)
    excluir_vacios = kwargs.get("excluir_vacios", True)
    consulta = consulta_horario(
        estacion_id,
        variable_id,
        inicio,
        fin,
        profundidad=profundidad,
        excluir_vacios=excluir_vacios,
    )
    fecha = []
    valor = []
    vacios = []
    for fila in consulta:
        fecha.append(fila.fecha)
        valor.append(fila.valor)
        vacios.append(fila.vacios)
    datos = {"fecha": fecha, "valor": valor, "vacios": vacios}
    return datos


def consulta_diario(estacion_id, variable_id, inicio, fin, *args, **kwargs):
    profundidad = kwargs.get("profundidad", False)
    excluir_vacios = kwargs.get("excluir_vacios", True)

    if inicio:
        inicio = datetime.datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime.datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)
    sql = """
    WITH 
    variable AS (
        SELECT * FROM variable_variable v WHERE v.var_id = %%var_id%%
    ),
    consulta AS (
        SELECT d.fecha, d.valor, d.vacios FROM diario_var%%var_id%%diario d 
        WHERE d.estacion_id = %%est_id%% 
        %%filtro%%
    ),
    secuencia AS (
        SELECT * FROM generate_series( 
            (SELECT c.fecha FROM consulta c ORDER BY c.fecha ASC LIMIT 1),
            (SELECT c.fecha FROM consulta c ORDER BY c.fecha DESC LIMIT 1),
            '1 day'::interval) fecha
    )
    SELECT s.fecha, c.valor, c.vacios 
     FROM secuencia s LEFT JOIN consulta c USING (fecha) ORDER BY fecha ASC;
    """
    filtro = ""
    if profundidad:
        filtro += "        AND d.profundidad = " + str(profundidad)
    if inicio:
        filtro += " AND d.fecha >='" + str(inicio) + "'"
    if fin:
        filtro += " AND d.fecha <='" + str(fin) + "'"
    if excluir_vacios:
        filtro += "\n        AND d.vacios < (SELECT v.vacios FROM variable v)"

    sql = sql.replace("%%est_id%%", str(estacion_id))
    sql = sql.replace("%%var_id%%", str(variable_id))
    sql = sql.replace("%%filtro%%", filtro)

    consulta = ConsultaReporteFecha.objects.raw(sql)
    return consulta


def consulta_diario__columnas(estacion_id, variable_id, inicio, fin, *args, **kwargs):
    profundidad = kwargs.get("profundidad", False)
    excluir_vacios = kwargs.get("excluir_vacios", True)
    consulta = consulta_diario(
        estacion_id,
        variable_id,
        inicio,
        fin,
        profundidad=profundidad,
        excluir_vacios=excluir_vacios,
    )
    fecha = []
    valor = []
    vacios = []
    for fila in consulta:
        fecha.append(fila.fecha)
        valor.append(fila.valor)
        vacios.append(fila.vacios)
    datos = {"fecha": fecha, "valor": valor, "vacios": vacios}
    return datos


def consulta_mensual(estacion_id, variable_id, inicio, fin, *args, **kwargs):
    profundidad = kwargs.get("profundidad", False)
    excluir_vacios = kwargs.get("excluir_vacios", True)

    if inicio:
        inicio = datetime.datetime(inicio.year, inicio.month, inicio.day, 0, 0, 0)

    if fin:
        fin = datetime.datetime(fin.year, fin.month, fin.day, 23, 59, 59, 999999)
    sql = """
    WITH 
    variable AS (
        SELECT * FROM variable_variable v WHERE v.var_id = %%var_id%%
    ),
    consulta AS (
        select * from mensual_var%%var_id%%mensual m 
        WHERE m.estacion_id = %%est_id%% 
        %%filtro%%
    ),
    secuencia AS (
        SELECT * FROM generate_series( 
            (SELECT c.fecha FROM consulta c ORDER BY c.fecha ASC LIMIT 1),
            (SELECT c.fecha FROM consulta c ORDER BY c.fecha DESC LIMIT 1),
            '1 month'::interval) fecha
    )
    SELECT s.fecha, c.valor, c.vacios 
     FROM secuencia s LEFT JOIN consulta c USING (fecha) ORDER BY fecha ASC;    
    """
    filtro = ""
    if profundidad:
        filtro += "        AND m.profundidad = " + str(profundidad)
    if inicio:
        filtro += " AND m.fecha >='" + str(inicio) + "'"
    if fin:
        filtro += " AND m.fecha <='" + str(fin) + "'"
    if excluir_vacios:
        filtro += "\n        AND m.vacios < (SELECT v.vacios FROM variable v)"

    sql = sql.replace("%%est_id%%", str(estacion_id))
    sql = sql.replace("%%var_id%%", str(variable_id))
    sql = sql.replace("%%filtro%%", filtro)

    consulta = ConsultaReporteFecha.objects.raw(sql)
    return consulta


def consulta_mensual__columnas(estacion_id, variable_id, inicio, fin, *args, **kwargs):
    profundidad = kwargs.get("profundidad", False)
    excluir_vacios = kwargs.get("excluir_vacios", True)
    consulta = consulta_mensual(
        estacion_id,
        variable_id,
        inicio,
        fin,
        profundidad=profundidad,
        excluir_vacios=excluir_vacios,
    )
    fecha = []
    valor = []
    vacios = []
    for fila in consulta:
        fecha.append(fila.fecha)
        valor.append(fila.valor)
        vacios.append(fila.vacios)
    datos = {"fecha": fecha, "valor": valor, "vacios": vacios}
    return datos


def diaToMesPre(dfd):
    """var = 1 precipitacion"""
    dfm = pd.DataFrame()
    if not dfd.empty:
        head = ["fecha", "vmes", "max", "dmax", "min", "dmin", "ndias"]
        # año dia mes valor
        rows = []
        fechai = dfd.iloc[0, 0]
        for i in range(0, 12):
            row = []
            if i < 11:
                fechai = datetime.date(fechai.year, (i + 1), 1)
                fechaf = fechai.replace(month=(i + 2))
                fechaf = fechaf - datetime.timedelta(days=1)
            else:
                fechai = datetime.date(fechai.year, 12, 1)
                fechaf = datetime.date(fechai.year, 12, 31)
            mes = dfd[(dfd["fecha"] >= fechai) & (dfd["fecha"] <= fechaf)]
            if not mes.empty:
                max = mes[mes["valor"] == mes["valor"].max()]
                # print(max)
                min = mes[mes["valor"] == mes["valor"].min()]
                vmes = mes["valor"].sum()
                ndias = len(mes[mes["valor"] >= 0.1]) + 1
                row.extend(
                    [
                        fechai,
                        vmes,
                        max.iloc[0, 1],
                        max.iloc[0, 0].day,
                        min.iloc[0, 1],
                        min.iloc[0, 0].day,
                        ndias,
                    ]
                )
            else:
                row.extend([fechai, "", "", "", "", "", ""])
            rows.append(row)
        dfm = pd.DataFrame(data=rows, columns=head)
        print(dfm)
    return dfm


def diaToMesTem(dfd):
    """var = 2 temperatura aire"""
    # "año", "mes", "dia", "valor","max_abs","min_abs","max_del_prom","min_del_prom"
    dfm = pd.DataFrame()
    if not dfd.empty:
        head = [
            "fecha",
            "vmes",
            "max_p",
            "dmax_p",
            "min_p",
            "dmin_p",
            "max_a",
            "dmax_a",
            "min_a",
            "dmin_a",
        ]
        # año dia mes valor
        fechai = dfd.iloc[0, 0]
        rows = []
        for i in range(0, 12):
            row = []
            if i < 11:
                fechai = datetime.date(fechai.year, (i + 1), 1)
                fechaf = fechai.replace(month=(i + 2))
                fechaf = fechaf - datetime.timedelta(days=1)
            else:
                fechai = datetime.date(fechai.year, 12, 1)
                fechaf = datetime.date(fechai.year, 12, 31)
            mes = dfd[(dfd["fecha"] >= fechai) & (dfd["fecha"] <= fechaf)]
            if not mes.empty:
                # 'año' 'dia' 'max_abs' 'max_del_prom' 'mes' 'min_abs' 'min_del_prom' 'valor'
                vmes = mes["valor"].mean()
                max_p = mes[mes["max_del_prom"] == mes["max_del_prom"].max()]
                min_p = mes[mes["min_del_prom"] == mes["min_del_prom"].min()]
                max_a = mes[mes["max_abs"] == mes["max_abs"].max()]
                min_a = mes[mes["min_abs"] == mes["min_abs"].min()]
                row.extend(
                    [
                        fechai,
                        vmes,
                        max_p.iloc[0, 2],
                        max_p.iloc[0, 0].day,
                        min_p.iloc[0, 4],
                        min_p.iloc[0, 0].day,
                        max_a.iloc[0, 1],
                        max_a.iloc[0, 0].day,
                        min_a.iloc[0, 3],
                        min_a.iloc[0, 0].day,
                    ]
                )
            else:
                row.extend([fechai, "", "", "", "", "", "", "", "", ""])
            rows.append(row)
        dfm = pd.DataFrame(data=rows, columns=head)
    return dfm


def barchart(title, xlabel, ylabel):
    trace1 = go.Bar(x=xlabel, y=ylabel, name=title)
    data = go.Data([trace1])
    layout = go.Layout(title=title)
    figure = go.Figure(data=data, layout=layout)
    div = opy.plot(figure, auto_open=False, output_type="div")
    ##print(div)
    return div


def lineChart(title, xlabel, ymax, ymin, ymed):
    trace0 = go.Scatter(
        x=xlabel, y=ymax, name="Max", line=dict(color=("rgb(22, 96, 167)"), width=4)
    )
    trace1 = go.Scatter(
        x=xlabel,
        y=ymin,
        name="Min",
        line=dict(
            color=("rgb(205, 12, 24)"),
            width=4,
        ),
    )
    trace2 = go.Scatter(
        x=xlabel,
        y=ymed,
        name="Media",
        line=dict(
            color=("rgb(50, 205, 50)"),
            width=4,
        ),
    )
    data = go.Data([trace0, trace1, trace2])
    layout = go.Layout(title=title)
    figure = go.Figure(data=data, layout=layout)
    div = opy.plot(figure, auto_open=False, output_type="div")
    return div


def getGrafico(estacion, variable, inicio, fin, frecuencia, *args, **kwargs):
    est_id = estacion.est_id
    var_id = variable.var_id
    profundidad = kwargs.get("profundidad", False)
    excluir_vacios = kwargs.get("excluir_vacios", True)

    titulo = estacion.est_codigo + " - " + variable.var_nombre
    if profundidad:
        titulo = titulo + " a " + str(profundidad / 100.0) + " [m]"
    titulo = titulo + "<br>(" + frecuencia.nombre + ")"

    if frecuencia.nombre == "Subhorario crudo":
        datos = consulta_crudos_saltos__columnas(
            est_id, var_id, inicio, fin, profundidad=profundidad
        )
        graf = grafico_simple(datos, variable, estacion, titulo, unir_puntos=True)
    elif frecuencia.nombre == "Subhorario validado":
        datos = consulta_validados_saltos__columnas(
            est_id, var_id, inicio, fin, profundidad=profundidad
        )
        graf = grafico_simple(datos, variable, estacion, titulo, unir_puntos=True)
    elif frecuencia.nombre == "Horario":
        datos = consulta_horario__columnas(
            est_id,
            var_id,
            inicio,
            fin,
            profundidad=profundidad,
            excluir_vacios=excluir_vacios,
        )
        graf = grafico_simple(
            datos,
            variable,
            estacion,
            titulo,
            unir_puntos=True,
            excluir_vacios=excluir_vacios,
        )
    elif frecuencia.nombre == "Diario":
        datos = consulta_diario__columnas(
            est_id,
            var_id,
            inicio,
            fin,
            profundidad=profundidad,
            excluir_vacios=excluir_vacios,
        )
        graf = grafico_simple(
            datos,
            variable,
            estacion,
            titulo,
            unir_puntos=True,
            excluir_vacios=excluir_vacios,
        )
    else:
        datos = consulta_mensual__columnas(
            est_id,
            var_id,
            inicio,
            fin,
            profundidad=profundidad,
            excluir_vacios=excluir_vacios,
        )
        graf = grafico_simple(
            datos,
            variable,
            estacion,
            titulo,
            unir_puntos=True,
            excluir_vacios=excluir_vacios,
        )

    return graf


def getDatos_grafico(estacion, variable, inicio, fin, frecuencia, *args, **kwargs):
    est_id = estacion.est_id
    var_id = variable.var_id
    profundidad = kwargs.get("profundidad", False)
    excluir_vacios = kwargs.get("excluir_vacios", True)

    if frecuencia.nombre == "Subhorario crudo":
        datos = consulta_crudos_saltos__columnas(
            est_id, var_id, inicio, fin, profundidad=profundidad
        )
    elif frecuencia.nombre == "Subhorario validado":
        datos = consulta_validados_saltos__columnas(
            est_id, var_id, inicio, fin, profundidad=profundidad
        )
    elif frecuencia.nombre == "Horario":
        datos = consulta_horario__columnas(
            est_id,
            var_id,
            inicio,
            fin,
            profundidad=profundidad,
            excluir_vacios=excluir_vacios,
        )
    elif frecuencia.nombre == "Diario":
        datos = consulta_diario__columnas(
            est_id,
            var_id,
            inicio,
            fin,
            profundidad=profundidad,
            excluir_vacios=excluir_vacios,
        )
    elif frecuencia.nombre == "Mensual":
        datos = consulta_mensual__columnas(
            est_id,
            var_id,
            inicio,
            fin,
            profundidad=profundidad,
            excluir_vacios=excluir_vacios,
        )

    if len(datos["fecha"]) == 0:
        return None

    periodo = 1440
    for i in range(0, int(len(datos["fecha"]) / 4)):
        periodoi = datos["fecha"][i + 1] - datos["fecha"][i]
        periodoi = round(periodoi.days * 1440 + periodoi.seconds / 60.0)
        periodoN_i = datos["fecha"][-(i + 1)] - datos["fecha"][-(i + 2)]
        periodoN_i = round(periodoN_i.days * 1440 + periodoN_i.seconds / 60.0)
        if periodoi == periodoN_i:
            periodo = periodoi
            break
        periodo = min(periodo, periodoi, periodoN_i)

    intervalo = datos["fecha"][-1] - datos["fecha"][0]
    ndatos_esperado = ((intervalo.days * 1440 + intervalo.seconds / 60.0) / periodo) + 1
    if ndatos_esperado < 1000:
        ndatos_esperado = 1000

    res = {}
    res["estacion"] = {}
    res["estacion"]["codigo"] = estacion.est_codigo
    res["variable"] = {}
    res["variable"]["nombre"] = variable.var_nombre
    res["variable"]["unidad_sigla"] = variable.uni_id.uni_sigla
    res["variable"]["es_acumulada"] = variable.es_acumulada
    res["frecuencia"] = __frecuencia__[frecuencia.nombre]
    res["fecha"] = datos["fecha"]
    res["valor"] = datos["valor"]
    if not excluir_vacios:
        res["vacios"] = datos["vacios"]
    res["periodo"] = periodo
    res["ndatos_esperado"] = ndatos_esperado
    res["profundidad"] = profundidad
    res["excluir_vacios"] = excluir_vacios
    return res


def getDatos_grafico2(estacion, variable, inicio, fin, frecuencia, *args, **kwargs):
    est_id = estacion.est_id
    var_id = variable.var_id
    profundidad = kwargs.get("profundidad", False)
    excluir_vacios = kwargs.get("excluir_vacios", True)

    if frecuencia.nombre == "Subhorario crudo":
        datos = consulta_crudos_saltos__lista(
            est_id, var_id, inicio, fin, profundidad=profundidad
        )
    elif frecuencia.nombre == "Subhorario validado":
        datos = consulta_validados_saltos(
            est_id, var_id, inicio, fin, profundidad=profundidad
        )
    elif frecuencia.nombre == "Horario":
        datos = consulta_horario(
            est_id,
            var_id,
            inicio,
            fin,
            profundidad=profundidad,
            excluir_vacios=excluir_vacios,
        )
    elif frecuencia.nombre == "Diario":
        datos = consulta_diario(
            est_id,
            var_id,
            inicio,
            fin,
            profundidad=profundidad,
            excluir_vacios=excluir_vacios,
        )
    elif frecuencia.nombre == "Mensual":
        datos = consulta_mensual(
            est_id,
            var_id,
            inicio,
            fin,
            profundidad=profundidad,
            excluir_vacios=excluir_vacios,
        )

    if len(datos["fecha"]) == 0:
        return None

    # periodo = 1440
    # for i in range(0, int(len(datos['fecha'])/4)):
    #     periodoi = datos['fecha'][i+1] - datos['fecha'][i]
    #     periodoi = round(periodoi.days*1440 + periodoi.seconds / 60.0)
    #     periodoN_i = datos['fecha'][-(i+1)] - datos['fecha'][-(i+2)]
    #     periodoN_i = round(periodoN_i.days*1440 + periodoN_i.seconds / 60.0)
    #     if periodoi == periodoN_i:
    #         periodo = periodoi
    #         break
    #     periodo = min(periodo, periodoi, periodoN_i)

    # intervalo = datos['fecha'][-1] - datos['fecha'][0]
    # ndatos_esperado = ((intervalo.days * 1440 + intervalo.seconds / 60.0) / periodo) + 1
    # if ndatos_esperado < 1000:
    #     ndatos_esperado = 1000

    res = {}
    res["estacion"] = {}
    res["estacion"]["codigo"] = estacion.est_codigo
    res["variable"] = {}
    res["variable"]["nombre"] = variable.var_nombre
    res["variable"]["unidad_sigla"] = variable.uni_id.uni_sigla
    res["variable"]["es_acumulada"] = variable.es_acumulada
    res["frecuencia"] = __frecuencia__[frecuencia.nombre]
    res["datos"] = datos
    # if (not excluir_vacios):
    #     res['vacios'] = datos['vacios']
    # res['periodo'] = periodo
    # res['ndatos_esperado'] = ndatos_esperado
    res["profundidad"] = profundidad
    res["excluir_vacios"] = excluir_vacios
    return res


def getDatos_exportar(estacion, variable, inicio, fin, frecuencia, *args, **kwargs):
    est_id = estacion.est_id
    var_id = variable.var_id
    profundidad = kwargs.get("profundidad", False)
    excluir_vacios = kwargs.get("excluir_vacios", True)

    nombre_archivo = estacion.est_codigo.replace(" ", "_") + "-" + variable.var_nombre
    if profundidad:
        nombre_archivo = nombre_archivo + " a " + str(profundidad / 100.0) + "[m]"
    nombre_archivo = nombre_archivo + " " + frecuencia.nombre
    nombre_archivo = nombre_archivo.replace(" ", "_")

    if frecuencia.nombre == "Subhorario crudo":
        datos = consulta_crudos(est_id, var_id, inicio, fin, profundidad=profundidad)
        formato_fecha = "yyyy/mm/dd hh:mm:ss"
    elif frecuencia.nombre == "Subhorario validado":
        datos = consulta_validados(est_id, var_id, inicio, fin, profundidad=profundidad)
        formato_fecha = "yyyy/mm/dd hh:mm:ss"
    elif frecuencia.nombre == "Horario":
        datos = consulta_horario(
            est_id,
            var_id,
            inicio,
            fin,
            profundidad=profundidad,
            excluir_vacios=excluir_vacios,
        )
        formato_fecha = "yyyy/mm/dd hh:mm:ss"
    elif frecuencia.nombre == "Diario":
        datos = consulta_diario(
            est_id,
            var_id,
            inicio,
            fin,
            profundidad=profundidad,
            excluir_vacios=excluir_vacios,
        )
        formato_fecha = "yyyy/mm/dd"
    else:
        datos = consulta_mensual(
            est_id,
            var_id,
            inicio,
            fin,
            profundidad=profundidad,
            excluir_vacios=excluir_vacios,
        )
        formato_fecha = "yyyy/mm"

    return datos, nombre_archivo, formato_fecha


def grafico_simple(datos, variable, estacion, titulo, *args, **kwargs):
    unir_puntos = kwargs.get("unir_puntos", False)
    excluir_vacios = kwargs.get("excluir_vacios", True)

    if len(datos["fecha"]) == 0:
        return None

    periodo = 1440
    for i in range(0, int(len(datos["fecha"]) / 4)):
        periodoi = datos["fecha"][i + 1] - datos["fecha"][i]
        periodoi = round(periodoi.days * 1440 + periodoi.seconds / 60.0)
        periodoN_i = datos["fecha"][-(i + 1)] - datos["fecha"][-(i + 2)]
        periodoN_i = round(periodoN_i.days * 1440 + periodoN_i.seconds / 60.0)
        if periodoi == periodoN_i:
            periodo = periodoi
            break
        periodo = min(periodo, periodoi, periodoN_i)

    intervalo = datos["fecha"][-1] - datos["fecha"][0]
    ndatos_esperado = ((intervalo.days * 1440 + intervalo.seconds / 60.0) / periodo) + 1

    fig = None
    if excluir_vacios:
        fig = make_subplots(rows=1, cols=1)
        total_height = 500
    else:
        fig = make_subplots(rows=2, cols=1, row_heights=[0.8, 0.2])
        total_height = 600

    trace = None
    if variable.es_acumulada:
        trace = go.Bar(
            x=datos["fecha"],
            y=datos["valor"],
            name=variable.uni_id.uni_sigla,
            showlegend=False,
            width=60000 * periodo,
            marker=dict(
                line=dict(width=0.3, color="rgb(0,0,0)"),
            ),
        )
    else:
        if unir_puntos:
            trace = go.Scatter(
                x=datos["fecha"],
                y=datos["valor"],
                name=variable.uni_id.uni_sigla,
                # mode='lines+markers',
                mode="lines",
                connectgaps=False,
                line=dict(
                    # shape='spline',
                    color=("rgb(63, 63, 220)"),
                ),
                # marker=dict(
                #     size=2,
                #     color='rgb(32, 32, 110)',
                # ),
                showlegend=False,
            )
        else:
            trace = go.Scatter(
                x=datos["fecha"],
                y=datos["valor"],
                name=variable.uni_id.uni_sigla,
                mode="markers",
                marker=dict(
                    size=2,
                    color="rgb(32, 32, 110)",
                ),
                showlegend=False,
            )
    fig.add_trace(trace, row=1, col=1)

    if not excluir_vacios:
        vacios = go.Bar(
            x=datos["fecha"],
            y=datos["vacios"],
            name="% Vacios",
            showlegend=False,
            width=60000 * periodo,
            marker=dict(
                line=dict(width=0.3, color="rgb(255,0,0)"),
            ),
            marker_color="indianred",
        )
        fig.add_trace(vacios, row=2, col=1)
        fig.update_yaxes(title_text="% Vacíos", row=2, col=1)

    if titulo is None:
        titulo = estacion.est_codigo

    pixels_por_dato = 1.2
    if ndatos_esperado < 1000:
        ndatos_esperado = 1000
    layout = go.Layout(
        autosize=False,
        width=160 + int(ndatos_esperado * pixels_por_dato),
        height=total_height,
        title=titulo,
        yaxis=dict(
            title=variable.var_nombre + " [" + variable.uni_id.uni_sigla + "]",
        ),
        xaxis=dict(
            # rangeslider=dict(visible=True),
            rangeselector=dict(buttons=list([dict(step="all")])),
            type="date",
        ),
    )
    fig.update_layout(layout)
    fig.update_layout(
        title={
            "y": 0.93,
            "x": 0,
            "xanchor": "left",
            "yanchor": "top",
            "pad": {
                "l": 65,
            },
        }
    )
    div = opy.plot(fig, auto_open=False, output_type="div")
    return div


def export_csv(estacion, variable, inicio, fin, frecuencia, *args, **kwargs):
    profundidad = kwargs.get("profundidad", False)
    excluir_vacios = kwargs.get("excluir_vacios", True)
    datos, nombre_archivo, formato_fecha = getDatos_exportar(
        estacion,
        variable,
        inicio,
        fin,
        frecuencia,
        profundidad=profundidad,
        excluir_vacios=excluir_vacios,
    )
    hay_datos = False
    for fila in datos:
        hay_datos = True
        break
    if not hay_datos:
        return None
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        'attachment; filename="' + nombre_archivo + '.csv"'
    )
    writer = csv.writer(response)
    if excluir_vacios:
        writer.writerow(["Fecha", "Valor (" + variable.uni_id.uni_sigla + ")"])
    else:
        writer.writerow(
            ["Fecha", "Valor (" + variable.uni_id.uni_sigla + ")", "% Vacíos"]
        )
    formato_fecha = (
        formato_fecha.replace("yyyy", "%Y").replace("/mm", "/%m").replace("dd", "%d")
    )
    formato_fecha = (
        formato_fecha.replace("hh", "%H").replace(":mm", ":%M").replace("ss", "%S")
    )
    if excluir_vacios:
        for fila in datos:
            writer.writerow([fila.fecha.strftime(formato_fecha), fila.valor])
    else:
        for fila in datos:
            writer.writerow(
                [fila.fecha.strftime(formato_fecha), fila.valor, fila.vacios]
            )
    return response


def export_excel(estacion, variable, inicio, fin, frecuencia, *args, **kwargs):
    profundidad = kwargs.get("profundidad", False)
    excluir_vacios = kwargs.get("excluir_vacios", True)
    datos, nombre_archivo, formato_fecha = getDatos_exportar(
        estacion,
        variable,
        inicio,
        fin,
        frecuencia,
        profundidad=profundidad,
        excluir_vacios=excluir_vacios,
    )
    hay_datos = False
    for fila in datos:
        hay_datos = True
        break
    if not hay_datos:
        return None

    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = (
        'attachment; filename="' + nombre_archivo + '.xls"'
    )
    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Datos")
    ws.col(0).width = len(formato_fecha) * 256
    ws.write(0, 0, "Fecha")
    ws.write(0, 1, "Valor (" + variable.uni_id.uni_sigla + ")")
    if not excluir_vacios:
        ws.write(0, 2, "% Vacíos")

    date_format = xlwt.XFStyle()
    date_format.num_format_str = formato_fecha
    row_num = 1
    if excluir_vacios:
        for fila in datos:
            ws.write(row_num, 0, fila.fecha, date_format)
            ws.write(row_num, 1, fila.valor)
            row_num = row_num + 1
    else:
        for fila in datos:
            ws.write(row_num, 0, fila.fecha, date_format)
            ws.write(row_num, 1, fila.valor)
            ws.write(row_num, 2, fila.vacios)
            row_num = row_num + 1

    wb.save(response)
    return response


def export_diario(estacion, variable, año):
    # REQUERIMIENTOS: pip install pillow        ## pil image
    wb = Workbook()
    ws = wb.active

    fecha_ini = datetime.datetime(año, 1, 1, 0, 0, 0)
    fecha_fin = datetime.datetime(año, 12, 31, 23, 59, 59)
    dataD = consulta_diario(
        estacion.est_id, variable.var_id, fecha_ini, fecha_fin, excluir_vacios=True
    )
    año = []
    mes = []
    dia = []
    valor = []
    for row in dataD:
        año.extend([row.fecha.year])
        mes.extend([row.fecha.month])
        dia.extend([row.fecha.day])
        valor.extend([row.valor])
    dfe = {"año": año, "dia": dia, "mes": mes, "valor": valor}
    df = pd.DataFrame(dfe)
    img_path = os.path.join(settings.BASE_DIR, "static/images/imhea_logo_20210419.png")
    img = Image(img_path)
    ws.add_image(img, "L3")

    etiMes = [
        "ene",
        "feb",
        "mar",
        "abr",
        "may",
        "jun",
        "jul",
        "ago",
        "sep",
        "oct",
        "nov",
        "dic",
    ]
    colnames = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]
    # llenamos Cabecera
    ws["A2"] = "PRECIPITACIÓN DIARIA ESTACIÓN " + estacion.est_codigo
    ws["A2"].font = Font(name="Arial", size=12, color="0066b3")
    ws.merge_cells("B5:N5")
    ws["B5"] = "AÑO " + str(año[0])
    ws["B5"].font = Font(name="Arial", size=12, bold=True)
    ws["B5"].alignment = Alignment(horizontal="center")
    ws["B6"] = "Día"
    ws["B38"] = "Tot."

    ws.column_dimensions["B"].width = float(4.5)
    # enumera 31 dias
    for rw in range(7, 38):
        ws.cell(row=rw, column=2, value=rw - 6)
    # recorrer los meses
    col = 3
    for i in range(0, 12):
        ws.cell(row=6, column=col + i, value=etiMes[i])
        ws.column_dimensions[colnames[i]].width = float(5.9)
        # llenar los datos desde le dataframe
        meses = df[(df["mes"] == (i + 1))]
        if not meses.empty:
            # acumulado mensual
            ws.cell(row=38, column=(col + i), value=meses["valor"].sum())
            for j in range(0, len(meses["valor"])):
                dia = meses.iloc[j, 1]
                ws.cell(
                    row=(dia + 6), column=(col + i), value=meses.iloc[j, 3]
                )  # los datos

    if not df.empty:
        # valores mayores a
        gtc1 = len(df[(df["valor"] > 0)])  # valores mayores a 0
        gtc2 = len(df[(df["valor"] >= 0.5)])  # valores mayores a 0.5
        gtc3 = len(df[(df["valor"] >= 10)])  # valores mayores a 10
        dfmax = df[df["valor"] == df["valor"].max()]
        print(dfmax)
        ws["A40"] = (
            str(gtc1)
            + " Lluvias > 0 mm,   "
            + str(gtc2)
            + " Lluvias >= 0.5 mm,   "
            + str(gtc3)
            + " Lluvias >= 10, Max. el "
            + str(int(dfmax.iloc[0, 1]))
            + "/"
            + str(int(dfmax.iloc[0, 2]))
            + "/"
            + str(int(dfmax.iloc[0, 0]))
            + " = "
            + str(round(dfmax.iloc[0, 3], 1))
            + " mm."
        )
        ws["A40"].font = Font(name="Arial", size=10, bold=True)
        ws["F41"] = "Total Anual = " + str(round(df["valor"].sum(), 1)) + " mm."
        ws["F41"].font = Font(name="Arial", size=10, bold=True)
    ws["A43"] = "Lagunas.-"
    ws["A43"].font = Font(name="Arial", size=10, bold=True)
    ws["B43"] = "Verde: 1 a 2 días"  # 62a73b
    ws["B43"].font = Font(name="Arial", size=10, bold=True, color="62a73b")
    ws["F43"] = "Naranja: 3 a 5 días"  #  f58220
    ws["F43"].font = Font(name="Arial", size=10, bold=True, color="f58220")
    ws["J43"] = "Rojo: 6 a 15 días"  #  ed1c24
    ws["J43"].font = Font(name="Arial", size=10, bold=True, color="ed1c24")

    # Formato de las celdas
    # formato de la tablas
    stTabla = NamedStyle(name="stTabla")
    mi_borde = Border(
        top=Side(border_style=borders.BORDER_THIN),
        bottom=Side(border_style=borders.BORDER_THIN),
        left=Side(border_style=borders.BORDER_THIN),
        right=Side(border_style=borders.BORDER_THIN),
    )
    stTabla.border = mi_borde
    fontTa = Font(name="Arial", size=11)
    stTabla.font = fontTa
    stTabla.alignment = Alignment(horizontal="center")
    for cel in ws["B6:N38"]:
        for val in cel:
            val.style = stTabla
    for cel in ws["B6:N6"]:
        for val in cel:
            val.font = Font(bold=True, size=10)
    for cel in ws["B7:B38"]:
        for val in cel:
            val.font = Font(bold=True, size=10)
    for mes_idx, mes_col in enumerate(colnames):
        vacios = 0
        dias_en_mes = monthrange(año[0], mes_idx + 1)[1]
        for _dia in range(1, dias_en_mes + 1):
            if ws[mes_col + str(_dia + 6)].value is None:
                vacios += 1
        if vacios == 0:
            ws[mes_col + str(38)].font = Font(bold=True, size=10, color="000000")
        elif vacios <= 2:
            ws[mes_col + str(38)].font = Font(bold=True, size=10, color="62a73b")
        elif vacios <= 5:
            ws[mes_col + str(38)].font = Font(bold=True, size=10, color="f58220")
        else:
            ws[mes_col + str(38)].font = Font(bold=True, size=10, color="ed1c24")

    ## nombre del archivo
    nf = estacion.est_codigo.replace(" ", "_") + "__Diarios.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nf)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


def export_mensual_multianual(estacion, variable, año_inicio, año_fin):
    # REQUERIMIENTOS: pip install pillow        ## pil image
    wb = Workbook()
    ws = wb.active
    ws["A1"].alignment = Alignment(horizontal="center")
    # ws.cell(row=6, column=col + i, value=etiMes[i])
    ws.cell(
        row=1, column=1, value="REGISTRO MULTIANUAL MENSUAL DE PLUVIOMETRÍA"
    ).font = Font(name="Arial", size=12, bold=True)
    ws.merge_cells("A1:N1")
    ws.cell(row=2, column=1, value="ESTACIÓN:").font = Font(
        name="Arial", size=10, bold=True
    )
    ws.cell(row=3, column=1, value="CÓDIGO:").font = Font(
        name="Arial", size=10, bold=True
    )
    ws.cell(row=4, column=1, value="COORDENADAS:").font = Font(
        name="Arial", size=10, bold=True
    )
    ws.cell(row=6, column=1, value="ELEVACIÓN:").font = Font(
        name="Arial", size=10, bold=True
    )
    ws.cell(row=2, column=3, value="").font = Font(name="Arial", size=10)
    ws.cell(row=3, column=3, value=estacion.est_codigo).font = Font(
        name="Arial", size=10
    )
    ws.cell(
        row=4, column=3, value="Latitud (°): " + str(round(estacion.est_latitud, 8))
    ).font = Font(name="Arial", size=10)
    ws.cell(
        row=5, column=3, value="Longitud (°): " + str(round(estacion.est_longitud, 8))
    ).font = Font(name="Arial", size=10)
    ws.cell(row=6, column=3, value=str(estacion.est_altura) + " msnm.").font = Font(
        name="Arial", size=10
    )
    ws["A8"].alignment = Alignment(horizontal="center")
    ws.cell(row=8, column=1, value="PRECIPITACIÓN MENSUAL").font = Font(
        name="Arial", size=12, bold=True
    )
    ws.merge_cells("A8:N8")
    stTabla = NamedStyle(name="stTabla")
    mi_borde = Border(
        top=Side(border_style=borders.BORDER_THIN),
        bottom=Side(border_style=borders.BORDER_THIN),
        left=Side(border_style=borders.BORDER_THIN),
        right=Side(border_style=borders.BORDER_THIN),
    )
    stTabla.border = mi_borde
    fontTa = Font(name="Arial", size=10, bold=False)
    fontTa_bold = Font(name="Arial", size=10, bold=True)
    stTabla.font = fontTa_bold
    stTabla.alignment = Alignment(horizontal="center")

    etiMes = [
        "AÑO",
        "ENE",
        "FEB",
        "MAR",
        "ABR",
        "MAY",
        "JUN",
        "JUL",
        "AGO",
        "SEP",
        "OCT",
        "NOV",
        "DIC",
        "TOTAL",
    ]
    colnames = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]

    for e in range(1, 15):
        ws.cell(row=10, column=e, value=etiMes[e - 1]).style = stTabla
        ws.column_dimensions[colnames[e - 1]].width = float(6.5)

    # completo_mediciones = Variable.objects.get(var_id=variable.var_id).umbral_completo
    fecha_ini = datetime.datetime(año_inicio, 1, 1, 0, 0, 0)
    fecha_fin = datetime.datetime(año_fin, 12, 31, 23, 59, 59)
    dataD = consulta_mensual(
        estacion.est_id, variable.var_id, fecha_ini, fecha_fin, excluir_vacios=True
    )
    año = []
    mes = []
    valor = []
    for row in dataD:
        año.extend([row.fecha.year])
        mes.extend([row.fecha.month])
        valor.extend([row.valor])
    try:
        dfe = {"año": año, "mes": mes, "valor": valor}
    except:
        return None
    df = pd.DataFrame(dfe)
    if df.shape[0] < 1:
        # No hay datos
        return None

    stTabla.font = fontTa
    fil = -1

    # prmedios mensuales
    dataM = []
    etime = etiMes[1:13]
    for i in range(1, 13):
        mdf = df[df["mes"] == (i)]
        dataM.extend(
            [np.around(np.mean(mdf["valor"].values.astype(float)), decimals=1)]
        )
    # dataM=dataM.astype(float)
    serM = pd.DataFrame({"valor": dataM}, index=etime)
    serM.plot(kind="bar", legend=None)
    plt.xlabel("meses")
    plt.subplots_adjust(bottom=0.15)
    plt.title(
        "Precipitación mensual multianual " + str(min(año)) + " - " + str(max(año))
    )
    plt.ylabel("Precipitación (mm.)")
    nombre_grafico = (
        "grafico_reporte_mensual_multianual_" + str(random.randint(1, 1000)) + ".png"
    )
    grafico_path = os.path.join(settings.MEDIA_ROOT, nombre_grafico)
    plt.savefig(grafico_path, dpi=90)

    plt.close()
    resta = df.iloc[0, 0] - 11
    print(resta)
    for index, row in df.iterrows():
        fil = row[0] - resta
        col = row[1] + 1
        val = row[2]
        ws.cell(row=fil, column=col, value=val)
        ws.cell(row=fil, column=1, value=row[0])
        ws.cell(
            row=fil,
            column=14,
            value="=IF(COUNTBLANK(B"
            + str(fil)
            + ":M"
            + str(fil)
            + ")<=1, SUM(B"
            + str(fil)
            + ":M"
            + str(fil)
            + '), "")',
        )
    for cel in ws["A11:M" + str(fil)]:
        for val in cel:
            val.style = stTabla

    stTabla.font = fontTa_bold
    for cel in ws["N11:N" + str(fil)]:
        for val in cel:
            val.style = stTabla

    ws.cell(row=(fil + 2), column=1, value="PROM").style = stTabla
    ws.cell(row=(fil + 3), column=1, value="MAX").style = stTabla
    ws.cell(row=(fil + 4), column=1, value="MIN").style = stTabla

    stTabla.font = fontTa
    for i in range(1, 13):
        ws.cell(
            row=(fil + 2),
            column=(i + 1),
            value="=TRUNC(AVERAGE("
            + colnames[i]
            + "11:"
            + colnames[i]
            + str(fil)
            + "), 1)",
        )
        ws.cell(
            row=(fil + 3),
            column=(i + 1),
            value="=MAX(" + colnames[i] + "11:" + colnames[i] + str(fil) + ")",
        )
        ws.cell(
            row=(fil + 4),
            column=(i + 1),
            value="=MIN(" + colnames[i] + "11:" + colnames[i] + str(fil) + ")",
        )
    ws.cell(
        row=(fil + 2),
        column=14,
        value="=SUM(B" + str(fil + 2) + ":M" + str(fil + 2) + ")",
    )
    ws.cell(
        row=(fil + 3),
        column=14,
        value="=MAX(B" + str(fil + 3) + ":M" + str(fil + 3) + ")",
    )
    ws.cell(
        row=(fil + 4),
        column=14,
        value="=MIN(B" + str(fil + 4) + ":M" + str(fil + 4) + ")",
    )
    for cel in ws["B" + str(fil + 2) + ":M" + str(fil + 4)]:
        for val in cel:
            val.style = stTabla

    stTabla.font = fontTa_bold
    for cel in ws["N" + str(fil + 2) + ":N" + str(fil + 4)]:
        for val in cel:
            val.style = stTabla

    img = Image(grafico_path)
    # add to worksheet and anchor next to cells
    ws.add_image(img, "B" + str(fil + 6))

    ## nombre del archivo
    nf = estacion.est_codigo.replace(" ", "_") + "__Mensual-Multianual.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nf)
    response["Content-Disposition"] = contenido
    wb.save(response)
    if os.path.exists(grafico_path):
        try:
            os.remove(grafico_path)
        except:
            pass
    return response


def anuario(estacion, año):
    ac = año
    meses = [
        "ene",
        "feb",
        "mar",
        "abr",
        "may",
        "jun",
        "jul",
        "ago",
        "sep",
        "oct",
        "nov",
        "dic",
    ]
    varChar = [2, 3, 7, 8, 10, 11, 13]
    context = {}
    varEst = Cruce.objects.filter(est_id=estacion)
    for value in varEst:
        varlee = Variable.objects.get(pk=value.var_id_id)
        varcod = varlee.var_codigo
        varname = varlee.var_nombre
        varuni = Unidad.objects.get(pk=varlee.uni_id_id)
        vardim = varuni.uni_sigla
        if value.var_id_id == 1:
            dataD = (
                globals()["Var" + str(value.var_id_id) + "Diario"]
                .objects.filter(
                    estacion_id=estacion.est_id,
                    fecha__gte=str(ac) + "-01-01",
                    fecha__lte=str(ac) + "-12-31",
                )
                .order_by("fecha")
            )
            dataD = dataD.values("fecha", "valor")

            framed = pd.DataFrame.from_records(dataD)
            prem = diaToMesPre(framed)
            if not prem.empty:
                context[varcod] = prem.to_dict(orient="records")
                context[str(value.var_id_id) + "_grafico"] = barchart(
                    varname + " " + vardim, xlabel=meses, ylabel=prem["vmes"]
                )
            else:
                context[varcod + "_msg"] = "No hay datos de " + varname
        elif value.var_id_id in varChar:
            dataD = (
                globals()["Var" + str(value.var_id_id) + "Diario"]
                .objects.filter(
                    estacion_id=estacion.est_id,
                    fecha__gte=str(ac) + "-01-01",
                    fecha__lte=str(ac) + "-12-31",
                )
                .order_by("fecha")
            )
            dataD = dataD.values(
                "fecha", "valor", "max_abs", "min_abs", "max_del_prom", "min_del_prom"
            )
            framed = pd.DataFrame.from_records(dataD)

            temm = diaToMesTem(framed)
            # "año", "mes", "vmes", "max_p", "dmax_p", "min_p", "dmin_p", "max_a", "dmax_a", "min_a", "dmin_a"
            if not temm.empty:
                context[varcod] = temm.to_dict(orient="records")
                context[str(value.var_id_id) + "_grafico"] = lineChart(
                    varname + " " + vardim,
                    xlabel=meses,
                    ymax=temm["max_p"],
                    ymin=temm["min_p"],
                    ymed=temm["vmes"],
                )
            else:
                context[varcod + "_msg"] = "No hay datos de " + varname
    return context


# # utiliza un modelo antiguo "Medicion"del SEDC, muy posiblemente ya no se lo necesite
# def datos_horarios_json(est_id, var_id, fec_ini, fec_fin):
#     consulta = (Medicion.objects.filter(est_id=est_id)
#                 .filter(var_id=var_id).filter(med_fecha__range=[fec_ini, fec_fin]))
#     consulta = consulta.annotate(year=ExtractYear('med_fecha'),
#                                  month=ExtractMonth('med_fecha'),
#                                  day=ExtractDay('med_fecha'),
#                                  hour=ExtractHour('med_fecha')
#                                  ).values('year', 'month', 'day', 'hour')
#     if (var_id == 1):
#         consulta = list(consulta.annotate(valor=Sum('med_valor')).
#                         values('valor', 'year', 'month', 'day', 'hour').
#                         order_by('year', 'month', 'day', 'hour'))
#     else:
#         consulta = list(consulta.annotate(valor=Avg('med_valor'),
#                                           maximo=Max('med_maximo'), minimo=Min('med_minimo')).
#                         values('valor', 'maximo', 'minimo', 'year', 'month', 'day', 'hour').
#                         order_by('year', 'month', 'day', 'hour'))
#     datos = []
#     if len(consulta) > 0:
#         for fila in consulta:
#             fecha_str = (str(fila.get('year')) + ":" +
#                          str(fila.get('month')) + ":" + str(fila.get('day')))
#             fecha = datetime.strptime(fecha_str, '%Y:%m:%d').date()
#             hora = datetime.time(fila.get('hour'))
#             fecha_hora = datetime.combine(fecha, hora)
#             dato = {
#                 'fecha': fecha_hora,
#                 'valor': fila.get('valor'),
#                 'maximo': fila.get('maximo'),
#                 'minimo': fila.get('minimo'),
#             }
#             datos.append(dato)
#     else:
#         datos = {
#             'mensaje': 'no hay datos'
#         }
#     return datos
