################################################################################################
# Plataforma para la Iniciativa Regional de Monitoreo Hidrológico de Ecosistemas Andinos (iMHEA)
# basada en los desarrollos realizados por:
#     1) FONDO PARA LA PROTECCIÓN DEL AGUA (FONAG), Ecuador.
#         Contacto: info@fonag.org.ec
#     2) EMPRESA PÚBLICA METROPOLITANA DE AGUA POTABLE Y SANEAMIENTO DE QUITO (EPMAPS), Ecuador.
#         Contacto: paramh2o@aguaquito.gob.ec
#
#  IMPORTANTE: Mantener o incluir esta cabecera con la mención de las instituciones creadoras,
#              ya sea en uso total o parcial del código.

""" These views allow the user to calcualte various derived hydrological indices using
the functions in this app. 
"""

import decimal

# from anuarios.models import *
import json
from datetime import date, datetime, timedelta

import pandas as pd
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.db.models import Max, Min
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.views import generic
from django.views.generic import FormView
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from djangomain.settings import BASE_DIR
from estacion.models import Estacion, SitioCuenca, Tipo
from horario.models import Var1Horario as HorarioPrecipitacion
from indices.forms import *
from validacion.models import Var1Validado
from validacion.models import Var1Validado as vali_rr
from validacion.models import Var10Validado

from .functions import (
    IndicadoresPrecipitacion,
    acumularDoble,
    calcular_dias_sin_caudal,
    calcular_escorrentia,
    consultaPeriodos,
    getCaudalFrec,
    getCaudalFrecMulti,
    getVarValidado,
    indicaCaudal,
    intensidadDiracion,
)


class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, decimal.Decimal):
            return {"__Decimal__": str(obj)}
        # Let the base class default method raise the TypeError
        return json.JSONEncoder.default(self, obj)


class Doblemasa(generic.FormView):
    template_name = "indices/doblemasa.html"
    form_class = SearchForm
    success_url = "indices/doblemasa"

    def post(self, request, *args, **kwargs):
        form = SearchForm(self.request.POST or None)
        # try:
        estacion1_id = int(request.POST.get("estacion1", None))
        estacion2_id = int(request.POST.get("estacion2", None))
        frecuencia = int(request.POST.get("frecuencia", None))
        variable_id = 1
        tinicio = request.POST.get("inicio", None)
        tfin = request.POST.get("fin", None)
        # print("tinico ",tinicio,"tfin ",tfin, " <=> frecuencia : ",frecuencia)
        inicio = datetime.strptime(tinicio + " 00:00:00", "%d/%m/%Y %H:%M:%S")
        fin = datetime.strptime(tfin + " 23:59:00", "%d/%m/%Y %H:%M:%S")
        intervalos = getVarValidado(variable_id, estacion1_id, inicio, fin, frecuencia)
        intervalos1 = getVarValidado(variable_id, estacion2_id, inicio, fin, frecuencia)
        dicAcum = acumularDoble(intervalos, intervalos1, frecuencia)
        per1 = consultaPeriodos(estacion1_id, frecuencia)
        per2 = consultaPeriodos(estacion2_id, frecuencia)
        dicAdd = {
            "f1max": per1[0][0]["fecha"].strftime("%m/%d/%Y %H:%M:%S"),
            "f1mim": per1[1][0]["fecha"].strftime("%m/%d/%Y %H:%M:%S"),
            "e1cod": per1[2][0]["est_codigo"],
            "f2max": per2[0][0]["fecha"].strftime("%m/%d/%Y %H:%M:%S"),
            "f2mim": per2[1][0]["fecha"].strftime("%m/%d/%Y %H:%M:%S"),
            "e2cod": per2[2][0]["est_codigo"],
        }
        dicAcum.append(dicAdd)
        data = json.dumps(dicAcum, allow_nan=True, cls=DecimalEncoder)
        # print(data)
        # intervalosSerial = serializers.serialize('json',intervalos, fields=('id','fecha','valor') )
        return HttpResponse(data, content_type="application/json")
        # return HttpResponse(intervalosSerial, content_type='application/json')


class PeriodoDatos(generic.View):
    template_name = "indices/rangos.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


class IndPrecip(generic.FormView):
    """Calculates available precipitation rates for the selected station"""

    template_name = "indices/precipitacion.html"
    form_class = IndPrecipForm
    success_url = "indices/precipitacion.html"

    def post(self, request, *args, **kwargs):
        form = SelecEstForm(self.request.POST or None)
        # try:
        estacion_id = int(request.POST.get("estacion", None))
        tinicio = request.POST.get("inicio", "vacio")
        tfin = request.POST.get("fin", None)
        print("tinico ", tinicio, "tfin ", tfin)
        #####3
        ###### Eliminar las siguientes filas
        ######
        # tinicio = '01/01/2005'
        # tfin = '31/12/2010'
        ##############################
        completo = True
        inicio = None
        fin = None

        if tinicio != "":
            inicio = datetime.strptime(tinicio, "%Y-%m-%d")
            completo = False
        if tfin != "":
            fin = datetime.strptime(tfin, "%Y-%m-%d")
            # completo =False
        indrr = IndicadoresPrecipitacion(estacion_id, inicio, fin, completo)
        data = indrr.makeDic()
        # data = indicaPreci(estacion_id,inicio,fin, completo)
        print("data en el view")
        print(data)
        data = json.dumps(data, allow_nan=True, cls=DecimalEncoder)
        return HttpResponse(data, content_type="application/json")


class IndCaudal(generic.FormView):
    """Calculates available flow rates for the selected station"""

    template_name = "indices/caudal.html"
    form_class = IndCaudForm
    success_url = "indices/caudal.html"

    def post(self, request, *args, **kwargs):
        form = SelecEstForm(self.request.POST or None)
        # try:
        estacion_id = int(request.POST.get("estacion", None))
        tinicio = request.POST.get("inicio", "vacio")
        tfin = request.POST.get("fin", None)
        print("tinico ", tinicio, "tfin ", tfin)
        completo = True
        inicio = None
        fin = None
        if tinicio != "":
            inicio = datetime.strptime(tinicio, "%Y-%m-%d")
            completo = False
        if tfin != "":
            fin = datetime.strptime(tfin, "%Y-%m-%d")
            # completo =False
        # anualizar(estacion_id)
        data = indicaCaudal(estacion_id, inicio, fin, completo)
        data = json.dumps(data, allow_nan=True, cls=DecimalEncoder)
        print(data)
        return HttpResponse(data, content_type="application/json")


class IntensidadRR(generic.FormView):
    """Plot the graph of intensity vs. duration for a station"""

    template_name = "indices/intendura.html"
    form_class = SelecEstForm
    success_url = "indices/intendura.html"

    def post(self, request, *args, **kwargs):
        estacion_id = int(request.POST.get("estacion", None))
        # print("Estacion a buscar " , estacion_id)
        tinicio = int(request.POST.get("anio", None))
        # print(type(tinicio))
        inicio = datetime(tinicio, 1, 1)
        fin = datetime(tinicio, 12, 31, 23, 59, 59)
        # print(inicio, fin)
        data = intensidadDiracion(estacion_id, inicio, fin)
        data = json.dumps(data, allow_nan=True, cls=DecimalEncoder)
        return HttpResponse(data, content_type="application/json")


class IntensidadRRMultiestacion(generic.FormView):
    template_name = "indices/intendura_multiestacion.html"
    form_class = IntensidadDuracionMultiestacionForm
    success_url = "indices/intendura_multiestacion.html"

    def post(self, request, *args, **kwargs):
        print("Intencidad multiestacion")
        lisEsta = request.POST.getlist("estacion")
        print("Estacion a buscar ", lisEsta)
        tinicio = int(request.POST.get("anio", None))
        inicio = datetime(tinicio, 1, 1)
        fin = datetime(tinicio, 12, 31, 23, 59, 59)
        print("fechas : ", inicio, fin)
        data = []
        if len(lisEsta) > 0:
            print(lisEsta)
            # recorrer la lista de estaciones para calcular
            for es in lisEsta:
                dict = intensidadDiracion(es, inicio, fin)
                data.append(dict)
        else:
            data = {"error": "Debe seleccionar por lo menos un estación"}
        ##data = intensidadDiracion(estacion_id,inicio,fin)
        ret = json.dumps(data, allow_nan=True, cls=DecimalEncoder)
        return HttpResponse(ret, content_type="application/json")


class DuracionCaudal(generic.FormView):
    # duracioncaudal.html
    """Plot a graph of the duration of flow for a station"""

    template_name = "indices/duracioncaudal.html"
    form_class = SelecCaudalForm
    success_url = "indices/duracioncaudal.html"

    def post(self, request, *args, **kwargs):
        estacion_id = int(request.POST.get("estacion", None))
        tinicio = request.POST.get("inicio", None)
        tfin = request.POST.get("fin", None)
        frecuencia = int(request.POST.get("frecuencia", None))
        inicio = None
        fin = None
        if tinicio != "":
            inicio = datetime.strptime(tinicio, "%Y-%m-%d")
        if tfin != "":
            fin = datetime.strptime(tfin, "%Y-%m-%d")
        data = getCaudalFrec(estacion_id, inicio, fin, frecuencia)
        if data is None:
            data = [{}]
        return HttpResponse(
            json.dumps(data, allow_nan=True, cls=DecimalEncoder),
            content_type="application/json",
        )


class DuracionCaudalExport(generic.FormView):
    # duracioncaudal.html
    """Plot a graph of the duration of flow for a station.
    Action of the "Export" button as opposed to "Calculate" button.
    """

    template_name = "indices/duracioncaudal.html"
    form_class = SelecCaudalForm
    success_url = "indices/duracioncaudal.html"

    def post(self, request, *args, **kwargs):
        estacion_id = int(request.POST.get("estacion", None))
        tinicio = request.POST.get("inicio", None)
        tfin = request.POST.get("fin", None)
        frecuencia = int(request.POST.get("frecuencia", None))
        inicio = None
        fin = None
        if tinicio != "":
            inicio = datetime.strptime(tinicio, "%Y-%m-%d")
        if tfin != "":
            fin = datetime.strptime(tfin, "%Y-%m-%d")
        data = getCaudalFrec(estacion_id, inicio, fin, frecuencia)
        if data is None:
            data = [{}]
            return HttpResponse(data, content_type="application/json")
        else:
            print(data)
            est = Estacion.objects.get(pk=estacion_id)
            # libro = pd.DataFrame.from_dict(data)
            # print(libro)
            wb = Workbook()

            nom_arc = "D_C_" + str(est.est_codigo) + ".xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename = {0}".format(nom_arc)
            ## formato del excel
            # ruta de la imagen
            ruta_fonag = str(BASE_DIR) + "/static/images/imhea_logo_20210419.png"
            # ruta_imhea = str(BASE_DIR) + '/media/imhea_logo2.jpg'
            img_fonag = Image(ruta_fonag)
            # img_imhea = Image(ruta_imhea)
            # estilo de negrita
            font_bold = Font(bold=True)
            # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
            ws = wb.active
            ws.title = str(est.est_codigo)
            ws.add_image(img_fonag, "A1")
            # ws.add_image(img_imhea, 'G1')

            ws["B4"] = "Cálculos realizados para las gráficas de duración de caudal"
            ws["B4"].font = font_bold
            ws["B4"].alignment = Alignment(horizontal="center")
            ws.merge_cells("B4:H4")
            if frecuencia == 1:
                ws["B5"] = "en base a datos horarios"
            else:
                ws["B5"] = "en base a datos diarios"
            ws["B5"].font = font_bold
            ws["B5"].alignment = Alignment(horizontal="center")
            ws.merge_cells("B5:H5")
            if data["aporte"]:
                ws["A8"] = "área de aporte "
                ws["A8"].font = font_bold
                ws.merge_cells("A8:B8")
                ws["C8"] = est.influencia_km

            ws["A7"] = "Estación"
            ws["A7"].font = font_bold
            ws["B7"] = est.est_codigo
            ws["C7"] = ""
            ws.merge_cells("C4:E4")
            ws["F7"] = "Variable"
            ws["F7"].font = font_bold
            ws["G7"] = "Caudal"
            ws["B9"] = "Coordenadas Geográfica "
            ws["B9"].font = font_bold
            ws.merge_cells("B6:G6")
            ws["A10"] = "Latitud"
            ws["A10"].font = font_bold
            ws["B10"] = est.est_latitud
            ws["F10"] = "Longitud"
            ws["F10"].font = font_bold
            ws["G10"] = est.est_longitud

            colu = 1
            for d in data["anios"]:
                fila = 13
                print(d, " ----- --- - -- ", colu)
                ws.cell(row=fila, column=colu).value = "Año " + str(d)
                ws.cell(row=fila, column=colu).alignment = Alignment(
                    horizontal="center"
                )
                if data["aporte"]:
                    ws.merge_cells(
                        start_row=fila,
                        start_column=colu,
                        end_row=fila,
                        end_column=(colu + 2),
                    )
                    ws.cell(row=fila + 1, column=colu).value = "Caudal"
                    ws.cell(row=fila + 1, column=colu + 1).value = "CaudalEsp"
                    ws.cell(row=fila + 1, column=colu + 2).value = "Frecuencia"
                    colu = colu + 3
                else:
                    ws.merge_cells(
                        start_row=fila,
                        start_column=colu,
                        end_row=fila,
                        end_column=(colu + 1),
                    )
                    ws.cell(row=fila + 1, column=colu).value = "Caudal"
                    ws.cell(row=fila + 1, column=colu + 1).value = "Frecuencia"
                    colu = colu + 2
                cu = "cau" + str(d)
                ce = "cauEsp" + str(d)
                fe = "fre" + str(d)
                fila = 15
                print(len(data["anuales"][cu]))
                for v in range(0, len(data["anuales"][cu])):
                    if data["aporte"]:
                        ws.cell(row=fila, column=colu - 3).value = data["anuales"][cu][
                            v
                        ]
                        ws.cell(row=fila, column=colu - 2).value = data["anuales"][ce][
                            v
                        ]
                        ws.cell(row=fila, column=colu - 1).value = data["anuales"][fe][
                            v
                        ]
                    else:
                        ws.cell(row=fila, column=colu - 2).value = data["anuales"][cu][
                            v
                        ]
                        ws.cell(row=fila, column=colu - 1).value = data["anuales"][fe][
                            v
                        ]
                    fila = fila + 1

            wb.save(response)
            response["Content-Disposition"] = contenido
            print("response ", response)
            return response


class DuracionCaudalMultiestacion(generic.FormView):
    """Compare flow duration curves of multiple stations."""

    template_name = "indices/duracioncaudal_multiestacion.html"
    form_class = CuvarCaudalMultiestacionForm
    success_url = "indices/duracioncaudal_multiestacion.html"

    def post(self, request, *args, **kwargs):
        estacion_id = request.POST.getlist("estacion")
        listEst = estacion_id
        tinicio = request.POST.get("inicio", None)
        tfin = request.POST.get("fin", None)
        print("fechas => ", tinicio, tfin)
        if len(estacion_id) <= 0:
            mensaje = {
                "menjaseErr": "  Debe seleccionar por lo menos una estación de la lista"
            }
            return HttpResponse(
                json.dumps(mensaje, allow_nan=True), content_type="application/json"
            )
        elif tinicio is None or tfin is None or tinicio == "" or tfin == "":
            mensaje = {
                "menjaseErr": "Debe Seleccionar la fecha de inicio y la fecha de fin",
                "Cualquir": "otracosa",
            }
            return HttpResponse(
                json.dumps(mensaje, allow_nan=True), content_type="application/json"
            )
        else:
            inicio = datetime.strptime(tinicio, "%Y-%m-%d")
            fin = datetime.strptime(tfin, "%Y-%m-%d")
            print("estaciones")
            print(estacion_id)
            tinicio = request.POST.get("inicio", None)
            tfin = request.POST.get("fin", None)
            print(tinicio, tfin)
            dict = getCaudalFrecMulti(listEst, inicio, fin)
            return HttpResponse(dict, content_type="application/json")


# Listar fechaS
def listar_anio(request, estacion):
    # datos = functions.consultar.variable(estacion)
    validados = Var1Validado.objects.filter(estacion_id__exact=estacion).aggregate(
        Max("fecha"), Min("fecha")
    )
    if validados["fecha__max"] != None:
        validados["fecha__max"] = validados["fecha__max"].year
        validados["fecha__min"] = validados["fecha__min"].year
        fechas = list(range(validados["fecha__min"], validados["fecha__max"] + 1))
    else:
        fechas = ["No existen datos"]
    return JsonResponse(fechas, safe=False)


# Listar fechaS multiestacion
def listar_anio_multi(request):
    lisEsta = request.POST.getlist("estacion")
    fecha_max = 0
    fecha_min = 0
    data = []
    if len(lisEsta) > 0:
        validados = Var1Validado.objects.filter(
            estacion_id__exact=lisEsta[0]
        ).aggregate(Max("fecha"), Min("fecha"))
        if validados["fecha__max"] != None:
            fecha_min = validados["fecha__min"].year
            fecha_max = validados["fecha__max"].year
            for es in lisEsta:
                validados = Var1Validado.objects.filter(
                    estacion_id__exact=es
                ).aggregate(Max("fecha"), Min("fecha"))
                if validados["fecha__max"] != None:
                    if fecha_max < validados["fecha__max"].year:
                        fecha_max = validados["fecha__max"].year

                    if fecha_min > validados["fecha__min"].year:
                        fecha_min = validados["fecha__min"].year
                else:
                    data = ["No existen datos"]

            data = list(range(fecha_min, fecha_max + 1))
        else:
            data = ["No existen datos"]
    else:
        data = ["No existen datos"]
    return JsonResponse(data, safe=False)


# Fecha maxima y minima caudal
def listar_fecha_caudal(request, estacion):
    data = Var10Validado.objects.filter(estacion_id__exact=estacion).aggregate(
        Max("fecha"), Min("fecha")
    )
    if data["fecha__max"] != None:
        # data['fecha__max'] = str(data['fecha__max'].day) +'/'+str(data['fecha__max'].month) +'/'+ str(data['fecha__max'].year)
        data["fecha__max"] = (
            str(data["fecha__max"].year)
            + "-"
            + str(data["fecha__max"].month)
            + "-"
            + str(data["fecha__max"].day)
        )
        # data['fecha__min'] = str(data['fecha__min'].day) +'/'+str(data['fecha__min'].month) +'/'+ str(data['fecha__min'].year)
        data["fecha__min"] = (
            str(data["fecha__min"].year)
            + "-"
            + str(data["fecha__min"].month)
            + "-"
            + str(data["fecha__min"].day)
        )
        # fecha = abs(data['fecha__max']  - data['fecha__min'] ).days
    else:
        data = ["No existen datos"]
    print(data)
    return JsonResponse(data, safe=False)


def listar_fecha_precipitacion(request, estacion):
    data = HorarioPrecipitacion.objects.filter(estacion_id__exact=estacion).aggregate(
        Max("fecha"), Min("fecha")
    )
    if data["fecha__max"] != None:
        # data['fecha__max'] = str(data['fecha__max'].day) +'/'+str(data['fecha__max'].month) +'/'+ str(data['fecha__max'].year)
        data["fecha__max"] = (
            str(data["fecha__max"].year)
            + "-"
            + str(data["fecha__max"].month)
            + "-"
            + str(data["fecha__max"].day)
        )
        # data['fecha__min'] = str(data['fecha__min'].day) +'/'+str(data['fecha__min'].month) +'/'+ str(data['fecha__min'].year)
        data["fecha__min"] = (
            str(data["fecha__min"].year)
            + "-"
            + str(data["fecha__min"].month)
            + "-"
            + str(data["fecha__min"].day)
        )
        # fecha = abs(data['fecha__max']  - data['fecha__min'] ).days
    else:
        data = ["No existen datos"]
    print(data)
    return JsonResponse(data, safe=False)


class EscorrentiaView(PermissionRequiredMixin, FormView):
    """Calculate the runoff coefficient for one or more instrument at a station."""

    permission_required = "indices.view_indices"
    template_name = "indices/escorrentia.html"
    form_class = EscorrentiaForm

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if self.request.is_ajax():
            if form.is_valid():
                sitiocuenca = form.cleaned_data.get("sitiocuenca")
                est_caudal = form.cleaned_data.get("est_caudal")
                est_precipitacion = form.cleaned_data.get("est_precipitacion")
                inicio = form.cleaned_data.get("inicio")
                fin = form.cleaned_data.get("fin")
                res = calcular_escorrentia(
                    sitiocuenca, est_caudal, est_precipitacion, inicio, fin
                )
                return JsonResponse(res)
            else:
                return JsonResponse({"mensaje": "No se admite AJAX"})
        else:
            return self.form_invalid(form)


@permission_required("indices.view_indices")
def estaciones_en_cuenca(request, sitiocuenca_id):
    """View to show the precipitation and flow stations in a basin (basic text output)."""
    sitiocuenca = SitioCuenca.objects.filter(pk=sitiocuenca_id)
    tipo_precipitacion = Tipo.objects.filter(
        nombre__in=["Pluviométrica", "Climatológica"]
    )
    est_precipitacion = Estacion.objects.filter(
        tipo__in=tipo_precipitacion, sitiocuenca__in=sitiocuenca
    )
    opciones_precipitacion = [[x.est_id, x.est_codigo] for x in est_precipitacion]

    tipo_caudal = Tipo.objects.filter(nombre__in=["Hidrométrica"])
    est_caudal = Estacion.objects.filter(
        tipo__in=tipo_caudal, sitiocuenca__in=sitiocuenca
    )
    opciones_caudal = [[x.est_id, x.est_codigo] for x in est_caudal]

    lista = {"precipitacion": opciones_precipitacion, "caudal": opciones_caudal}
    return JsonResponse(lista)
