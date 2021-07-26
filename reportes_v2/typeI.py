# -*- coding: utf-8 -*-

################################################################################################
# Plataforma para la Iniciativa Regional de Monitoreo Hidrológico de Ecosistemas Andinos (iMHEA)
# basada en los desarrollos realizados por:
#     1) FONDO PARA LA PROTECCIÓN DEL AGUA (FONAG), Ecuador.
#         Contacto: info@fonag.org.ec
#     2) EMPRESA PÚBLICA METROPOLITANA DE AGUA POTABLE Y SANEAMIENTO DE QUITO (EPMAPS), Ecuador.
#         Contacto: paramh2o@aguaquito.gob.ec
#
#  IMPORTANTE: Mantener o incluir esta cabecera con la mención de las instituciones creadoras,
#              ya sea en uso total o parcial del código.

'''from anuarios.models import HumedadSuelo
from anuarios.models import PresionAtmosferica
from anuarios.models import TemperaturaAgua
from anuarios.models import Caudal
from anuarios.models import NivelAgua
from django.db.models import Avg'''
import plotly.offline as opy
import plotly.graph_objs as go
from reportes_v2.titulos import Titulos
import calendar

from openpyxl.chart import (
    ScatterChart,
    LineChart,
    Reference,
    Series,
)


# clase para anuario de las las variables HSU, PAT, TAG, CAU, NAG
class TypeI(Titulos):

    def matriz(self, estacion, variable, periodo):
        datos = self.consulta(estacion, variable, periodo)
        return datos

    def grafico(self, estacion, variable, periodo):
        datos = self.consulta(estacion, variable, periodo)
        if len(datos) > 0:
            meses = []
            max_simple = []
            min_simple = []
            avg_simple = []
            for item in datos:
                if variable.var_id == 6:
                    meses.append(str(calendar.month_abbr[item.hsu_mes]))
                    max_simple.append(item.hsu_maximo)
                    min_simple.append(item.hsu_minimo)
                    avg_simple.append(item.hsu_promedio)
                elif variable.var_id == 8:
                    meses.append(str(calendar.month_abbr[item.pat_mes]))
                    max_simple.append(item.pat_maximo)
                    min_simple.append(item.pat_minimo)
                    avg_simple.append(item.pat_promedio)
                elif variable.var_id == 9:
                    meses.append(str(calendar.month_abbr[item.tag_mes]))
                    max_simple.append(item.tag_maximo)
                    min_simple.append(item.tag_minimo)
                    avg_simple.append(item.tag_promedio)
                elif variable.var_id == 10:
                    meses.append(str(calendar.month_abbr[item.cau_mes]))
                    max_simple.append(item.cau_maximo)
                    min_simple.append(item.cau_minimo)
                    avg_simple.append(item.cau_promedio)
                elif variable.var_id == 11:
                    meses.append(str(calendar.month_abbr[item.nag_mes]))
                    max_simple.append(item.nag_maximo)
                    min_simple.append(item.nag_minimo)
                    avg_simple.append(item.nag_promedio)

            trace0 = go.Scatter(
                x=meses,
                y=max_simple,
                name='Max',
                line=dict(
                    color='rgb(22, 96, 167)',
                    width=4)
            )
            trace1 = go.Scatter(
                x=meses,
                y=min_simple,
                name='Min',
                line=dict(
                    color='rgb(205, 12, 24)',
                    width=4, )
            )
            trace2 = go.Scatter(
                x=meses,
                y=avg_simple,
                name='Media',
                line=dict(
                    color='rgb(50, 205, 50)',
                    width=4, )
            )
            #trace3 = go.Scatter(
            #    x=meses,
            #    y=historico,
            #    name='Media',
            #    line=dict(
            #        color='rgb(125, 96, 160)',
            #        width=4, )
            #)
            data = [trace0, trace1, trace2]
            #data = [trace0, trace1, trace2, trace3]
            layout = go.Layout(
                title=str(variable.var_nombre) + str(" (") + str(variable.uni_id.uni_sigla) + str(") ")
            )
            figure = go.Figure(data=data, layout=layout)
            figure.update_layout(legend_orientation="h")
            div = opy.plot(figure, auto_open=False, output_type='div')
            return div
        return False

    def tabla_excel(self, ws, estacion, variable, periodo):
        fila = 5
        col_fin = 11
        col = 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col_fin)
        subtitle = ws.cell(row=fila, column=col)
        subtitle_end = ws.cell(row=fila, column=col_fin)
        subtitle.value = self.get_titulo(variable.var_id)
        self.set_style(cell=subtitle, font='font_bold_10', alignment='center',
                       border='border_thin', fill='light_salmon')
        self.set_style(cell=subtitle_end, font='font_bold_10', alignment='center',
                       border='border_thin', fill='light_salmon')

        fila += 1
        col = 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila + 1, end_column=col)
        cell = ws.cell(row=fila, column=col)
        cell.value = "MES"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')
        col += 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col+3)
        cell = ws.cell(row=fila, column=col)
        cell_final = ws.cell(row=fila, column=col+3)
        cell.value = str(variable.var_nombre) + str(" (") + str(variable.uni_id.uni_sigla) + str(")")
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')
        self.set_style(cell=cell_final, font='font_10', alignment='center',
                       border='border_thin')
        fila += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Máximo"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Mínimo"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Media"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Media His"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        matriz = self.matriz(estacion, variable, periodo)
        fila += 1
        col = 1

        media_historica = self.datos_historicos(estacion, variable, periodo)

        for item in matriz:
            mes = self.get_item_mes(variable.var_id, item)
            cell = ws.cell(row=fila, column=col)
            cell.value = self.get_mes_anio(mes)
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')
            cell = ws.cell(row=fila, column=col+1)
            cell.value = self.get_item_maximo(variable.var_id, item)
            self.set_style(cell=cell, font='font_10', alignment='center',
                           border='border_thin')

            cell = ws.cell(row=fila, column=col+2)
            cell.value = self.get_item_minimo(variable.var_id, item)
            self.set_style(cell=cell, font='font_10', alignment='center',
                           border='border_thin')

            cell = ws.cell(row=fila, column=col+3)
            cell.value = self.get_item_promedio(variable.var_id, item)
            self.set_style(cell=cell, font='font_10', alignment='center',
                           border='border_thin')
            if len(media_historica) > 0  and len(media_historica) >= mes:
                cell = ws.cell(row=fila, column=col+4)
                #print('pruebaaaa')
                #print('prueba',media_historica)
                cell.value = round(media_historica[mes-1], 1)
                self.set_style(cell=cell, font='font_10', alignment='center',
                               border='border_thin')

            fila += 1

    @staticmethod
    def grafico_excel(ws, variable, periodo):
        c1 = ScatterChart()
        c1.title = str(variable.var_nombre) + str(" (") + \
            str(variable.uni_id.uni_sigla) + str(") ") + str(periodo)
        # c1.style = 13
        # c1.y_axis.title = str(variable.uni_id.uni_sigla)
        c1.x_axis.title = 'Meses'

        xvalues = Reference(ws, min_col=1, min_row=8, max_row=19)

        for i in range(2, 6):
            values = Reference(ws, min_col=i, min_row=7, max_row=19)
            series = Series(values, xvalues, title_from_data=True)
            c1.series.append(series)

        serie_max = c1.series[0]
        serie_max.marker.symbol = "diamond"
        serie_max.marker.graphicalProperties.solidFill = "1660a7"
        serie_max.marker.graphicalProperties.line.solidFill = "1660a7"
        serie_max.graphicalProperties.line.solidFill = "1660a7"

        serie_min = c1.series[1]
        serie_min.marker.symbol = "triangle"
        serie_min.marker.graphicalProperties.solidFill = "cd0c18"
        serie_min.marker.graphicalProperties.line.solidFill = "cd0c18"
        serie_min.graphicalProperties.line.solidFill = "cd0c18"

        serie_pro = c1.series[2]
        serie_pro.marker.symbol = "square"
        serie_pro.marker.graphicalProperties.solidFill = "32cd32"
        serie_pro.marker.graphicalProperties.line.solidFill = "32cd32"
        serie_pro.graphicalProperties.line.solidFill = "32cd32"

        serie_pro = c1.series[3]
        serie_pro.marker.symbol = "x"
        serie_pro.marker.graphicalProperties.solidFill = "7d60a0"
        serie_pro.marker.graphicalProperties.line.solidFill = "7d60a0"
        serie_pro.graphicalProperties.line.solidFill = "7d60a0"

        cats = Reference(ws, min_col=1, min_row=8, max_row=19)
        c1.set_categories(cats)
        c1.legend.position = "b"
        ws.add_chart(c1, "F6")

    @staticmethod
    def get_titulo(variable):
        titulo = ''
        if variable == 6:
            titulo = 'Humedad del Suelo - Valores medios diarios, absolutos maximos y mimimos'
        elif variable == 8:
            titulo = 'Presion Atomsferica - Valores medios diarios, absolutos maximos y mimimos'
        elif variable == 9:
            titulo = 'Temperatura de Agua - Valores medios diarios, medios maximos y mimimos'
        elif variable == 10:
            titulo = 'Caudal - Valores medios diarios, medios maximos y mimimos'
        elif variable == 11:
            titulo = 'Nivel del agua - Valores medios diarios, medios maximos y mimimos'
        return titulo
    
    @staticmethod
    def get_item_promedio(variable, item):
        promedio = None
        if variable == 6:
            promedio = item.hsu_promedio
        elif variable == 8:
            promedio = item.pat_promedio
        elif variable == 9:
            promedio = item.tag_promedio
        elif variable == 10:
            promedio = item.cau_promedio
        elif variable == 11:
            promedio = item.nag_promedio
        return promedio

    @staticmethod
    def get_item_maximo(variable, item):
        maximo = None
        if variable == 6:
            maximo = item.hsu_maximo
        elif variable == 8:
            maximo = item.pat_maximo
        elif variable == 9:
            maximo = item.tag_maximo
        elif variable == 10:
            maximo = item.cau_maximo
        elif variable == 11:
            maximo = item.nag_maximo
        return maximo

    @staticmethod
    def get_item_minimo(variable, item):
        minimo = None
        if variable == 6:
            minimo = item.hsu_minimo
        elif variable == 8:
            minimo = item.pat_minimo
        elif variable == 9:
            minimo = item.tag_minimo
        elif variable == 10:
            minimo = item.cau_minimo
        elif variable == 11:
            minimo = item.nag_minimo
        return minimo

    @staticmethod
    def get_item_mes(variable, item):
        mes = None
        if variable == 6:
            mes = item.hsu_mes
        elif variable == 8:
            mes = item.pat_mes
        elif variable == 9:
            mes = item.tag_mes
        elif variable == 10:
            mes = item.cau_mes
        elif variable == 11:
            mes = item.nag_mes
        return mes
