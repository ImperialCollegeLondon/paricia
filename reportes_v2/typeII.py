# -*- coding: utf-8 -*-

################################################################################################
# Plataforma para la Iniciativa Regional de Monitoreo Hidrológico de Ecosistemas Andinos (iMHEA)
# basada en los desarrollos realizados por:
#     1) FONDO PARA LA PROTECCIÓN DEL AGUA (FONAG), Ecuador.
#         Contacto: info@fonag.org.ec
#     2) EMPRESA PÚBLICA METROPOLITANA DE AGUA POTABLE Y SANEAMIENTO DE QUITO (EPMAPS), Ecuador.
#         Contacto: paramh2o@aguaquito.gob.ec
#
#  IMPORTANTE: Mantener o incluir esta cabecera con la mención de las instituciones creadoras,
#              ya sea en uso total o parcial del código.


import plotly.offline as opy
import plotly.graph_objs as go
from reportes_v2.titulos import Titulos
from anuarios.models import Var1Anuarios
from django.db.models import Avg, Max, Min
from django.db.models import FloatField
import calendar

from openpyxl.chart import BarChart, Reference, Series, LineChart
from openpyxl.chart.axis import DateAxis, ChartLines


# clase para anuario de la variable PRE
class TypeII(Titulos):

    @staticmethod
    def datos_historicos(estacion, periodo, parametro):
        consulta = Var1Anuarios.objects.filter(est_id=estacion)
        consulta = consulta.exclude(pre_periodo=periodo).values('pre_mes')
        if parametro == 'promedio':
            informacion = list(
                consulta
                .annotate(valor=Avg('pre_suma')).order_by('pre_mes')
            )
        elif parametro == 'maximo':
            informacion = list(
                consulta
                .annotate(valor=Max('pre_suma', output_field=FloatField())-Avg('pre_suma', output_field=FloatField()))
                .order_by('pre_mes')
            )
        else:
            informacion = list(
                consulta
                .annotate(valor=Avg('pre_suma', output_field=FloatField())-Min('pre_suma', output_field=FloatField()))
                .order_by('pre_mes')
            )
        datos = []
        # suma = 0
        for item in informacion:
            datos.append(item['valor'])
            # suma += item['valor']

        # datos.append(suma)
        return datos

    def matriz(self, estacion, variable, periodo):
        datos = self.consulta(estacion, variable, periodo)
        '''sum_pre = 0
        avg_max = 0
        sum_max = 0
        sum_max_dia = 0
        avg_max_dia = 0
        sum_dias = 0
        obj_pre = Var2Anuarios
        for item in datos:
            sum_pre += item.pre_suma
            sum_max += item.pre_maximo
            sum_max_dia += item.pre_maximo_dia
            sum_dias += item.pre_dias
        obj_pre.pre_mes = 13
        obj_pre.pre_suma = sum_pre
        obj_pre.pre_maximo = round(sum_max/12, 1)
        obj_pre.pre_maximo_dia = round(sum_max_dia/12, 1)
        obj_pre.pre_dias = sum_dias
        datos.append(obj_pre)'''
        return datos

    def grafico(self, estacion, variable, periodo):
        datos = self.consulta(estacion, variable, periodo)
        if datos:
            historicos = self.datos_historicos(estacion, periodo, 'promedio')
            max_historico = self.datos_historicos(estacion, periodo, 'maximo')
            min_historico = self.datos_historicos(estacion, periodo, 'minimo')

            meses = []
            mensual_simple = []
            for item in datos:
                meses.append(str(calendar.month_abbr[item.pre_mes]))
                mensual_simple.append(item.pre_suma)
            trace1 = go.Bar(
                x=meses,
                y=mensual_simple,
                name='Var2Anuarios (mm)',
                marker_color='rgb(31, 119, 180)',
            )
            trace2 = go.Bar(
                x=meses,
                y=historicos,
                name='Pre. Historica (mm)',
                marker_color='rgb(255, 127, 14)',
                error_y=dict(
                    type='data',
                    symmetric=False,
                    array=max_historico,
                    arrayminus=min_historico
                )
            )
            data = [trace1, trace2]
            layout = go.Layout(
                title=str(variable.var_nombre) + str(" (") + str(variable.uni_id.uni_sigla) + str(") ")
            )
            figure = go.Figure(data=data, layout=layout)
            figure.update_layout(legend_orientation="h")

            div = opy.plot(figure, auto_open=False, output_type='div')
            return div
        return False

    def tabla_excel(self, ws, estacion, variable, periodo):
        fila = 5
        col_fin = 11
        col = 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col_fin)
        subtitle = ws.cell(row=fila, column=col)
        subtitle.value = "Var2Anuarios - Valores mensuales y maximos diarios"
        self.set_style(cell=subtitle, font='font_bold_10', alignment='center',
                       border='border_thin', fill='light_salmon')
        fila += 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila + 1, end_column=col)
        cell = ws.cell(row=fila, column=col)
        cell.value = "MES"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        ws.merge_cells(start_row=fila, start_column=col+7, end_row=fila + 1, end_column=col+7)
        cell = ws.cell(row=fila, column=col+7)
        cell.value = "# de días con precipitación"
        self.set_style(cell=cell, font='font_8', alignment='wrap',
                       border='border_thin')
        col += 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col+3)
        cell = ws.cell(row=fila, column=col)
        cell.value = "Precipitación (mm)"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')
        fila += 1
        col = 2
        # ws.merge_cells(start_row=fila, start_column=col, end_row=fila+1, end_column=col)
        cell = ws.cell(row=fila, column=col)
        cell.value = "Mensual"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')
        col += 1
        # ws.merge_cells(start_row=fila, start_column=col, end_row=fila + 1, end_column=col)
        cell = ws.cell(row=fila, column=col)
        cell.value = "Media His"
        self.set_style(cell=cell, font='font_10', alignment='wrap',
                       border='border_thin')

        col += 1
        # ws.merge_cells(start_row=fila, start_column=col, end_row=fila + 1, end_column=col)
        cell = ws.cell(row=fila, column=col)
        cell.value = "Delta Máx His (Max His - Media His)"
        self.set_style(cell=cell, font='font_10', alignment='wrap',
                       border='border_thin')

        col += 1
        # ws.merge_cells(start_row=fila, start_column=col, end_row=fila + 1, end_column=col)
        cell = ws.cell(row=fila, column=col)
        cell.value = "Delta Mín His (Media His - Min His)"
        self.set_style(cell=cell, font='font_10', alignment='wrap',
                       border='border_thin')

        col += 1
        fila = 6
        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col+1)
        cell = ws.cell(row=fila, column=col)
        cell.value = "Máxima en"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        fila += 1
        col = 6

        cell = ws.cell(row=fila, column=col)
        cell.value = "24H (mm)"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        col += 1
        cell = ws.cell(row=fila, column=col)
        cell.value = "Día"
        self.set_style(cell=cell, font='font_10', alignment='center',
                       border='border_thin')

        matriz = self.matriz(estacion, variable, periodo)
        fila += 1
        col = 1
        historicos = self.datos_historicos(estacion, periodo, 'promedio')
        max_historico = self.datos_historicos(estacion, periodo, 'maximo')
        min_historico = self.datos_historicos(estacion, periodo, 'minimo')

        for item in matriz:
            cell = ws.cell(row=fila, column=col)
            cell.value = self.get_mes_anio(item.pre_mes)
            self.set_style(cell=cell, font='font_10', alignment='left',
                           border='border_thin')
            cell = ws.cell(row=fila, column=col + 1)
            cell.value = item.pre_suma
            self.set_style(cell=cell, font='font_10', alignment='center',
                           border='border_thin')
            if len(historicos) > 0 and len(historicos) > item.pre_mes:
                cell = ws.cell(row=fila, column=col + 2)
                cell.value = round(historicos[item.pre_mes-1], 1)
                self.set_style(cell=cell, font='font_10', alignment='wrap',
                               border='border_thin')
            if len(max_historico) > 0:
                cell = ws.cell(row=fila, column=col + 3)
                cell.value = round(max_historico[item.pre_mes-1], 1)
                self.set_style(cell=cell, font='font_10', alignment='wrap',
                               border='border_thin')
            if len(min_historico) > 0:
                cell = ws.cell(row=fila, column=col + 4)
                cell.value = round(min_historico[item.pre_mes-1], 1)
                self.set_style(cell=cell, font='font_10', alignment='wrap',
                               border='border_thin')

            cell = ws.cell(row=fila, column=col + 5)
            cell.value = item.pre_maximo
            self.set_style(cell=cell, font='font_10', alignment='center',
                           border='border_thin')

            cell = ws.cell(row=fila, column=col + 6)
            cell.value = item.pre_maximo_dia
            self.set_style(cell=cell, font='font_10', alignment='center',
                           border='border_thin')

            cell = ws.cell(row=fila, column=col + 7)
            cell.value = item.pre_dias
            self.set_style(cell=cell, font='font_10', alignment='center',
                           border='border_thin')

            fila += 1

    @staticmethod
    def grafico_excel(ws, variable, periodo):
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = str(variable.var_nombre) + str(" (") + \
            str(variable.uni_id.uni_sigla) + str(") ") + str(periodo)
        chart1.x_axis.title = 'Meses'

        data = Reference(ws, min_col=2, min_row=7, max_row=19, max_col=3)
        cats = Reference(ws, min_col=1, min_row=8, max_row=19)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 5

        chart1.legend.position = "b"
        ws.add_chart(chart1, "A21")
