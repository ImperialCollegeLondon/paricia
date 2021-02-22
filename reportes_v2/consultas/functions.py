# -*- coding: utf-8 -*-

################################################################################################
# Plataforma para la Iniciativa Regional de Monitoreo Hidrológico de Ecosistemas Andinos (iMHEA)
# basada en los desarrollos realizados por:
#     1) FONDO PARA LA PROTECCIÓN DEL AGUA (FONAG), Ecuador.
#         Contacto: info@fonag.org.ec
#     2) EMPRESA PÚBLICA METROPOLITANA DE AGUA POTABLE Y SANEAMIENTO DE QUITO (EPMAPS), Ecuador.
#         Contacto: paramh2o@aguaquito.gob.ec
#
#  IMPORTANTE: Mantener o incluir esta cabecera con la mención de las instituciones creadoras,
#              ya sea en uso total o parcial del código.


from estacion.models import Estacion
from variable.models import Variable
from cruce.models import Cruce
from importacion.functions import get_modelo

from datetime import timedelta, datetime
from math import ceil
from datetime import date

from django.db import connection
from djangomain.settings import BASE_DIR
# librerias para manejar los archivos EXCEL
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.drawing.image import Image

from openpyxl.chart import (
    LineChart,
    Reference,
)
from django.http import HttpResponse
import csv

suma = 'sum(valor) as valor '
promedio = 'avg(valor) as valor'
promedio_nivel = 'avg(nivel) as valor'
promedio_caudal = 'avg(caudal) as valor'
max_abs = 'max(maximo) as max_abs'
max_pro = 'max(valor) as max_pro'
max_pro_nivel = 'max(nivel) as max_pro'
max_pro_caudal = 'max(caudal) as max_pro'
min_abs = 'min(minimo) as min_abs'
min_pro = 'min(valor) as min_pro'
min_pro_nivel = 'min(nivel) as min_pro'
min_pro_caudal = 'min(caudal) as min_pro'
coma = ', '
group_order = 'GROUP BY 1 ORDER BY 1'


def reporte_excel(form):
    estacion = form.cleaned_data['estacion']
    variable = form.cleaned_data['variable']
    fecha_inicio = form.cleaned_data['inicio']
    fecha_fin = form.cleaned_data['fin']
    frecuencia = form.cleaned_data['frecuencia']
    print(estacion, variable, fecha_inicio, fecha_fin)
    if fecha_inicio is None:
        fecha_inicio = estacion.est_fecha_inicio
    if fecha_fin is None:
        fecha_fin = date.today()
    if frecuencia == "1":
        informacion = datos_horarios(estacion, variable, fecha_inicio, fecha_fin)
    elif frecuencia == "2":
        informacion = datos_diarios(estacion, variable, fecha_inicio, fecha_fin)
    else:
        informacion = datos_mensuales(estacion, variable, fecha_inicio, fecha_fin)
    print("longitud de datos",len(informacion["valor"]))
    tiempo = informacion["tiempo"]
    valores = informacion["valor"]
    maximos_abs = informacion["max_abs"]
    minimos_abs = informacion["min_abs"]
    maximos_pro = informacion["max_pro"]
    minimos_pro = informacion["min_pro"]

    # ruta de la imagen
    ruta = str(BASE_DIR) + '/media/logo_fonag.jpg'
    img = Image(ruta)
    # estilo de negrita
    font_bold = Font(bold=True)
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    ws['B1'] = 'Reporte de Datos Hidrometerológicos'
    ws['B1'].font = font_bold
    ws.add_image(img, 'G1')
    ws.merge_cells('B1:F1')
    ws['A4'] = 'Estación'
    ws['A4'].font = font_bold
    ws['B4'] = estacion.est_codigo
    ws['C4'] = estacion.est_nombre
    ws.merge_cells('C4:E4')
    ws['F4'] = 'Variable'
    ws['F4'].font = font_bold
    ws['G4'] = variable.var_nombre
    ws['B6'] = 'Coordenadas Geográfica TMQ WGS84'
    ws['B6'].font = font_bold
    ws.merge_cells('B6:G6')
    ws['A7'] = 'Latitud'
    ws['A7'].font = font_bold
    ws['B7'] = estacion.est_latitud
    ws['F7'] = 'Longitud'
    ws['F7'].font = font_bold
    ws['G7'] = estacion.est_longitud

    # Creamos los encabezados desde la celda B9 hasta la E9
    ws['A9'] = 'Fecha'
    ws['B9'] = 'Valor'

    cont = 10
    # Recorremos el conjunto de datos
    if len(maximos_abs) != maximos_abs.count(None) and len(maximos_pro) != maximos_pro.count(None):
        ws['C9'] = 'Max absoluto'
        ws['D9'] = 'Min absoluto'
        ws['E9'] = 'Max promedio'
        ws['F9'] = 'Min promedio'
        for valor, maximo_abs, maximos_pro, minimo_abs, minimo_pro, fecha in zip(valores, maximos_abs, maximos_pro,
                                                                                 minimos_abs, minimos_pro, tiempo):
            ws.cell(row=cont, column=1).value = fecha
            ws.cell(row=cont, column=2).value = valor
            ws.cell(row=cont, column=3).value = maximo_abs
            ws.cell(row=cont, column=4).value = minimo_abs
            ws.cell(row=cont, column=5).value = maximos_pro
            ws.cell(row=cont, column=6).value = minimo_pro
            cont = cont + 1
    else:
        ws['C9'] = 'Max promedio'
        ws['D9'] = 'Min promedio'
        for valor, maximos_pro, minimo_pro, fecha in zip(valores, maximos_pro, minimos_pro, tiempo):
            ws.cell(row=cont, column=1).value = fecha
            ws.cell(row=cont, column=2).value = valor
            ws.cell(row=cont, column=3).value = maximos_pro
            ws.cell(row=cont, column=4).value = minimo_pro
            cont = cont + 1

    # grafico
    chart = LineChart()
    chart.title = variable.var_nombre
    chart.style = 12
    chart.x_axis.title = 'Tiempo'
    chart.x_axis.number_format = 'dd-mm-yyyy'
    chart.x_axis.majorTimeUnit = 'days'
    chart.y_axis.title = variable.var_nombre
    final = 10 + len(tiempo)
    data = Reference(ws, min_col=2, min_row=9, max_col=4, max_row=final - 1)
    chart.add_data(data, titles_from_data=True)
    s0 = chart.series[0]
    s0.graphicalProperties.line.solidFill = "32CD32"
    s0.graphicalProperties.line.width = 10
    s0.smooth = True
    s1 = chart.series[1]
    s1.graphicalProperties.line.solidFill = "1645A7"
    s1.graphicalProperties.line.width = 10
    s1.smooth = True
    s2 = chart.series[2]
    s2.graphicalProperties.line.solidFill = "CD0C18"
    s2.graphicalProperties.line.width = 10
    s2.smooth = True
    dates = Reference(ws, min_col=1, min_row=10, max_row=final)
    chart.set_categories(dates)
    if frecuencia != "0":
        ws.add_chart(chart, "H9")
    # Establecemos el nombre del archivo
    nombre_archivo = str('"') + str(estacion.est_codigo) + str("_") + str(variable.var_nombre) + str('.xlsx"')
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


# generar reporte CSV
def reporte_csv(form):
    estacion = form.cleaned_data['estacion']
    variable = form.cleaned_data['variable']
    fecha_inicio = form.cleaned_data['inicio']
    fecha_fin = form.cleaned_data['fin']
    if fecha_inicio is None:
        fecha_inicio = estacion.est_fecha_inicio
    if fecha_fin is None:
        fecha_fin = date.today()
    informacion = datos_instantaneos(estacion, variable, fecha_inicio, fecha_fin)
    valores = informacion['valor']
    maximos = informacion['max_abs']
    minimos = informacion['min_abs']
    tiempo = informacion['tiempo']
    # Establecemos el nombre del archivo
    nombre_archivo = str('"') + str(estacion.est_codigo) + str("_") + str(variable.var_nombre) + str('.csv"')
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = contenido
    writer = csv.writer(response)
    if variable.var_id == 1:
        writer.writerow(['fecha', 'valor'])
        for valor, fecha in zip(valores, tiempo):
            writer.writerow([fecha, valor])
    else:
        writer.writerow(['fecha', 'valor', 'maximo', 'minimo'])
        for valor, maximo, minimo, fecha in zip(valores, maximos, minimos, tiempo):
            writer.writerow([fecha, valor, maximo, minimo])

    return response


def datos_instantaneos(estacion, variable, fecha_inicio, fecha_fin):

    modelo = get_modelo(variable.var_id)
    fecha_inicio = datetime(fecha_inicio.year, fecha_inicio.month, fecha_inicio.day, 0, 0, 0)
    fecha_fin = datetime(fecha_fin.year, fecha_fin.month, fecha_fin.day, 23, 59, 59, 999999)
    tabla = 'medicion_' + str(variable.var_modelo)
    sql = 'SELECT * FROM ' + tabla + ' WHERE estacion_id =' + str(estacion.est_id)
    # if fecha_inicio:
    sql = sql + ' AND fecha>=\'' + str(fecha_inicio) + '\''
    # if fecha_fin:
    sql = sql + ' AND fecha<=\'' + str(fecha_fin) + '\''
    sql = sql + ' order by fecha ASC;'
    print(sql)
    consulta = list(modelo.objects.raw(sql))
    valor = []
    maximo = []
    minimo = []
    frecuencia = []
    for fila in consulta:
        if fila.valor is not None:
            valor.append(fila.valor)
        else:
            valor.append(None)
        if variable.var_id != 1:
            if fila.maximo is not None:
                maximo.append(fila.maximo)
            else:
                maximo.append(None)
            if fila.minimo is not None:
                minimo.append(fila.minimo)
            else:
                minimo.append(None)
        # frecuencia.append(fila.fecha)
        frecuencia.append(fila.fecha.strftime("%Y-%m-%d %H:%M:%S"))
    print(len(valor))
    informacion = dict(
        valor=valor,
        max_abs=maximo,
        min_abs=minimo,
        max_pro=[],
        min_pro=[],
        tiempo=frecuencia
    )
    return informacion


def datos_estacion(estacion, fecha_inicio, fecha_fin):
    cruce = list(Cruce.objects.filter(est_id=estacion))
    i = 0
    i_fecha = 0
    num_variables=len(cruce)
    valores = [[0 for x in range(0)] for y in range(num_variables+1)]
    for fila in cruce:
        valor, maximo, minimo, frecuencia = datos_instantaneos(estacion, fila.var_id, fecha_inicio, fecha_fin)
        # llenar la frecuencia de tiempo
        if i == 0:
            valores[i] = set_valores(frecuencia)
            valores[i].insert(0, "Fecha hora")
            i += 1
        # comprobar si hay otras frecuecnas de tiempo ej: el viento se registra cada 2 min
        elif len(frecuencia) < (len(valores[i_fecha])-1):
            valores[i] = set_valores(frecuencia)
            valores[i].insert(0, "Fecha hora")
            i_fecha = i
            valores.append([0 for x in range(0)])
            i += 1
        # agregar el valor de cada variable
        valores[i] = set_valores(valor)
        valores[i].insert(0, fila.var_id.var_nombre)
        # verificar si hay maximos
        if contar_vacios(maximo) > 0:
            i += 1
            valores.append([0 for x in range(0)])
            valores[i] = set_valores(maximo)
            valores[i].insert(0, fila.var_id.var_nombre + " Máxima")
        # verificar si hay mínimos
        if contar_vacios(minimo)>0:
            i += 1
            valores.append([0 for x in range(0)])
            valores[i] = set_valores(minimo)
            valores[i].insert(0, fila.var_id.var_nombre + " Mínima")
        i += 1
    num_fil = len(valores[0])
    matriz = [[0 for x in range(0)] for y in range(0)]
    for i in range(len(valores)):
        if len(valores[i]) == num_fil:
            matriz.append(valores[i])
    for i in range(len(valores)):
        if len(valores[i]) < num_fil:
            matriz.append(valores[i])
    return matriz


def datos_5minutos(estacion, variable, fecha_inicio, fecha_fin):
    datos = armar_consulta(estacion, variable, 1, fecha_inicio, fecha_fin)
    valor = []
    maximo_abs = []
    maximo_pro = []
    minimo_abs = []
    minimo_pro = []
    frecuencia = []
    intervalo = timedelta(minutes=5)
    fecha = datetime.combine(fecha_inicio, datetime.min.time())
    int_5minutos = ((fecha_fin - fecha_inicio).days + 1) * 288
    num_datos = len(datos)
    item = 0
    for dia in range(int_5minutos):
        if item < num_datos:
            fecha_datos = datos[item].get('fecha')
            if fecha_datos == fecha:
                valor.append(datos[item].get('valor'))
                maximo_abs.append(datos[item].get('max_abs'))
                maximo_pro.append(datos[item].get('max_pro'))
                minimo_abs.append(datos[item].get('min_abs'))
                minimo_pro.append(datos[item].get('min_pro'))
                item += 1
            else:
                valor.append(None)
                maximo_abs.append(None)
                maximo_pro.append(None)
                minimo_abs.append(None)
                minimo_pro.append(None)
            #frecuencia.append(fecha)
            frecuencia.append(fecha.strftime("%Y-%m-%d %H:%M:%S"))
            fecha += intervalo

    informacion = dict(
        valor=valor,
        max_abs=maximo_abs,
        min_abs=minimo_abs,
        max_pro=maximo_pro,
        min_pro=minimo_pro,
        tiempo=frecuencia
    )

    return informacion


def datos_horarios(estacion, variable, fecha_inicio, fecha_fin):
    datos = armar_consulta(estacion, variable, 2, fecha_inicio, fecha_fin)
    valor = []
    maximo_abs = []
    maximo_pro = []
    minimo_abs = []
    minimo_pro = []
    frecuencia = []
    intervalo = timedelta(hours=1)
    fecha = datetime.combine(fecha_inicio, datetime.min.time())
    int_horas = ((fecha_fin - fecha_inicio).days + 1) * 24
    num_datos = len(datos)
    item = 0
    for dia in range(int_horas):
        if item < num_datos:
            fecha_datos = datos[item].get('fecha')
            if fecha_datos == fecha:
                valor.append(datos[item].get('valor'))
                maximo_abs.append(datos[item].get('max_abs'))
                maximo_pro.append(datos[item].get('max_pro'))
                minimo_abs.append(datos[item].get('min_abs'))
                minimo_pro.append(datos[item].get('min_pro'))
                item += 1
            else:
                valor.append(None)
                maximo_abs.append(None)
                maximo_pro.append(None)
                minimo_abs.append(None)
                minimo_pro.append(None)
            # frecuencia.append(fecha)
            frecuencia.append(fecha.strftime("%Y-%m-%d %H:%M:%S"))
            fecha += intervalo

    informacion = dict(
        valor=valor,
        max_abs=maximo_abs,
        min_abs=minimo_abs,
        max_pro=maximo_pro,
        min_pro=minimo_pro,
        tiempo=frecuencia
    )

    return informacion


def datos_diarios(estacion, variable, fecha_inicio, fecha_fin):
    datos = armar_consulta(estacion, variable, 3, fecha_inicio, fecha_fin)
    valor = []
    # maximos absolutos
    maximo_abs = []
    maximo_pro = []
    minimo_abs = []
    minimo_pro = []
    frecuencia = []
    intervalo = timedelta(days=1)

    fecha = fecha_inicio
    int_dias = (fecha_fin - fecha_inicio).days + 1
    num_datos = len(datos)
    item = 0
    for dia in range(int_dias):
        if item < num_datos:
            fecha_datos = datos[item].get('fecha').date()
            if fecha_datos == fecha:

                valor.append(datos[item].get('valor'))
                maximo_abs.append(datos[item].get('max_abs'))
                maximo_pro.append(datos[item].get('max_pro'))
                minimo_abs.append(datos[item].get('min_abs'))
                minimo_pro.append(datos[item].get('min_pro'))
                item += 1
            else:
                valor.append(None)
                maximo_abs.append(None)
                maximo_pro.append(None)
                minimo_abs.append(None)
                minimo_pro.append(None)
            frecuencia.append(fecha)
            fecha += intervalo

    informacion = dict(
        valor=valor,
        max_abs=maximo_abs,
        min_abs=minimo_abs,
        max_pro=maximo_pro,
        min_pro=minimo_pro,
        tiempo=frecuencia
    )
    return informacion


def datos_mensuales(estacion, variable, fecha_inicio, fecha_fin):
    datos = armar_consulta(estacion, variable, 4, fecha_inicio, fecha_fin)
    valor = []
    maximo_abs = []
    maximo_pro = []
    minimo_abs = []
    minimo_pro = []
    frecuencia = []
    fecha = fecha_inicio.replace(day=1)
    # intervalo = timedelta(days=30)
    dias = float((fecha_fin - fecha_inicio).days)
    meses = int(ceil(dias / 30))
    num_datos = len(datos)
    item = 0
    for mes in range(meses):
        if item < num_datos:
            fecha_datos = datos[item].get('fecha').date()
            if fecha == fecha_datos:
                valor.append(datos[item].get('valor'))
                maximo_abs.append(datos[item].get('max_abs'))
                maximo_pro.append(datos[item].get('max_pro'))
                minimo_abs.append(datos[item].get('min_abs'))
                minimo_pro.append(datos[item].get('min_pro'))
                item += 1
            else:
                valor.append(None)
                maximo_abs.append(None)
                maximo_pro.append(None)
                minimo_abs.append(None)
                minimo_pro.append(None)
            frecuencia.append(fecha)
            intervalo = dias_mes(fecha.month, fecha.year)
            fecha += timedelta(days=intervalo)
    informacion = dict(
        valor=valor,
        max_abs=maximo_abs,
        min_abs=minimo_abs,
        max_pro=maximo_pro,
        min_pro=minimo_pro,
        tiempo=frecuencia
    )
    return informacion


def armar_consulta(estacion, variable, frecuencia, fecha_inicio, fecha_fin):

    cursor = connection.cursor()
    # datos = []
    fecha_inicio = datetime(fecha_inicio.year, fecha_inicio.month, fecha_inicio.day, 0, 0, 0)
    fecha_fin = datetime(fecha_fin.year, fecha_fin.month, fecha_fin.day, 23, 59, 59, 999999)
    fecha = ""
    if frecuencia == 1:
        fecha = 'to_timestamp(floor((extract(\'epoch\' '
        fecha += 'from fecha) / 300 )) * 300)'
        fecha += 'AT TIME ZONE \'UTC5\' as fecha, '
    elif frecuencia == 2:
        fecha = 'date_trunc(\'hour\',fecha) as fecha, '
    elif frecuencia == 3:
        fecha = 'date_trunc(\'day\',fecha) as fecha, '
    elif frecuencia == 4:
        fecha = 'date_trunc(\'month\',fecha) as fecha, '

    if (variable.var_id == 4 or variable.var_id == 5) and estacion.est_externa is False:
        tabla = 'validacion_viento'
    elif (variable.var_id == 10 or variable.var_id == 11) and estacion.est_externa is False:
        tabla = 'validacion_agua'
    else:
        tabla = 'validacion_' + str(variable.var_modelo.lower())

    filtro = 'estacion_id=' + str(estacion.est_id)
    if fecha_inicio:
        filtro += ' and fecha>=\'' + str(fecha_inicio) + '\''
    if fecha_fin:
        filtro += ' and fecha<=\'' + str(fecha_fin) + '\' '

    if variable.var_id == 1:
        filtro += ' and valor !=\'NaN\'::numeric '
        sql = 'SELECT ' + fecha + suma
        sql += ' FROM ' + tabla + ' WHERE '
        sql += filtro + group_order
    elif variable.var_id == 5 and estacion.est_externa is False:
        filtro += ' and direccion !=\'NaN\'::numeric '
        sql = 'SELECT ' + fecha
        sql += 'AVG(direccion) as valor' + coma + 'MAX(direccion) as max_pro' + coma + 'MIN(direccion) as min_pro'
        sql += ' FROM ' + tabla + ' WHERE '
        sql += filtro + group_order

    elif variable.var_id == 10 and estacion.est_externa is False:
        filtro += ' and caudal !=\'NaN\'::numeric '
        sql = 'SELECT ' + fecha
        sql += promedio_caudal + coma + max_pro_caudal + coma + min_pro_caudal
        sql += ' FROM ' + tabla + ' WHERE '
        sql += filtro + group_order
    elif variable.var_id == 11 and estacion.est_externa is False:
        filtro += ' and nivel !=\'NaN\'::numeric '
        sql = 'SELECT ' + fecha
        sql += promedio_nivel + coma + max_pro_nivel + coma + min_pro_nivel
        sql += ' FROM ' + tabla + ' WHERE '
        sql += filtro + group_order
    else:
        filtro += ' and valor !=\'NaN\'::numeric '
        sql = 'SELECT ' + fecha
        sql += promedio + coma + max_abs + coma + max_pro + coma + min_abs + coma + min_pro
        sql += ' FROM ' + tabla + ' WHERE '
        sql += filtro + group_order

    cursor.execute(sql)

    datos = dictfetchall(cursor)
    cursor.close()
    return datos




def dias_mes(month, year):
    dias = 30
    if month <= 7:
        if month % 2 == 0 and month != 2:
            dias = 30
        elif month % 2 == 0 and month == 2:
            if year % 4 == 0:
                dias = 29
            else:
                dias = 28
        else:
            dias = 31
    else:
        if month % 2 == 0:
            dias = 31
        else:
            dias = 30
    return dias


def titulo_frecuencia(frecuencia):
    nombre = []
    if frecuencia == '0':
        nombre = 'Instantanea'
    if frecuencia == '1':
        nombre = '5 Minutos'
    elif frecuencia == '2':
        nombre = 'Horaria'
    elif frecuencia == '3':
        nombre = 'Diaria'
    elif frecuencia == '4':
        nombre = 'Mensual'
    return nombre


def dictfetchall(cursor):
    # Return all rows from a cursor as a dict
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]


def datos_horarios_json(est_id, var_id, fec_ini, fec_fin):
    fecha_ini = datetime.strptime(fec_ini, '%Y-%m-%d %H:%M:%S')
    fecha_fin = datetime.strptime(fec_fin, '%Y-%m-%d %H:%M:%S')
    datos = []
    estacion = Estacion.objects.get(est_id=est_id)
    variable = Variable.objects.get(var_id=var_id)

    informacion = datos_horarios(estacion, variable, fecha_ini, fecha_fin)
    val = informacion['valor']
    max_abs = informacion['max_abs']
    min_abs = informacion['min_abs']
    max_pro = informacion['max_pro']
    min_pro = informacion['min_pro']
    tiempo = informacion['tiempo']
    if len(val) > 0:
        for item_val, item_max_abs, item_max_pro, item_min_abs, item_min_pro, item_time \
                in zip(val, max_abs, max_pro, min_abs, min_pro, tiempo):
            if fecha_ini <= datetime.strptime(item_time, '%Y-%m-%d %H:%M:%S') <= fecha_fin:
                dato = {
                    'fecha': item_time,
                    'valor': item_val,
                    'maximo_absoluto': item_max_abs,
                    'minimo_absolulo': item_min_abs,
                    'maximo_promedio': item_max_pro,
                    'minimo_promedio': item_min_pro,
                }
                datos.append(dato)
    else:
        datos = {
            'mensaje': 'no hay datos'
        }
    return datos


def contar_vacios(valores):
    num_vacios = 0
    for item in valores:
        if item is not None:
            num_vacios += 1
    return num_vacios


def set_valores(valores):
    datos = []
    for item in valores:
        if item is not None:
            datos.append(item)
        else:
            datos.append('')
    return datos
