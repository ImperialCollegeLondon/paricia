# -*- coding: utf-8 -*-

################################################################################################
# Plataforma para la Iniciativa Regional de Monitoreo Hidrológico de Ecosistemas Andinos (iMHEA)
# basada en los desarrollos realizados por:
#     1) FONDO PARA LA PROTECCIÓN DEL AGUA (FONAG), Ecuador.
#         Contacto: info@fonag.org.ec
#     2) EMPRESA PÚBLICA METROPOLITANA DE AGUA POTABLE Y SANEAMIENTO DE QUITO (EPMAPS), Ecuador.
#         Contacto: paramh2o@aguaquito.gob.ec
#
#  IMPORTANTE: Mantener o incluir esta cabecera con la mención de las instituciones creadoras,
#              ya sea en uso total o parcial del código.

from variable.models import Variable
# librerias para manejar los archivos EXCEL
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from validacion_v2.functions import normalize
from anuarios.models import Var1Anuarios
from anuarios.models import Var2Anuarios
from anuarios.models import Var3Anuarios
from anuarios.models import Var6Anuarios
from anuarios.models import Var8Anuarios
from anuarios.models import Var9Anuarios
from anuarios.models import Var10Anuarios
from anuarios.models import Var11Anuarios

from django.db.models import Avg


class Titulos:
    # Estilo de Fuentes
    fonts = {
        'font_bold_11': Font(bold=True),
        'font_bold_10': Font(bold=True, size=10),
        'font_10': Font(size=10),
        'font_11': Font(size=11),
        'font_8': Font(size=8)
    }
    # Alineacion de las celdas
    alignments = {
        'center': Alignment(horizontal="center", vertical="center"),
        'right': Alignment(horizontal="right", vertical="center"),
        'left': Alignment(horizontal="left", vertical="center"),
        'wrap': Alignment(horizontal="center", vertical="center", wrap_text=True)
    }
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    # bordes
    borders = {
        'border_thin': Border(top=thin, left=thin, bottom=thin, right=thin),
        'border_medium': Border(top=medium, left=medium, right=medium, bottom=medium)
    }
    # relleno
    colors = {
        'orange': PatternFill("solid", fgColor="FDE9D9"),
        'light_salmon': PatternFill("solid", fgColor="FFA07A")
    }

    @staticmethod
    def titulo_grafico(variable):
        # returns var_nombre given var_id
        consulta = list(Variable.objects.filter(var_id=variable))

        return consulta[0]

    @staticmethod
    def titulo_unidad(variable):
        var = Variable.objects.get(var_id=variable)
        return var.uni_id.uni_sigla

    @staticmethod
    def consulta(estacion, variable, periodo):
        if variable.var_id == 1:
            informacion = list(Var1Anuarios.objects.filter(est_id=estacion).filter(pre_periodo=periodo))
        elif variable.var_id == 2:
            informacion = list(Var2Anuarios.objects.filter(est_id=estacion).filter(tai_periodo=periodo))
        elif variable.var_id == 3:
            informacion = list(Var3Anuarios.objects.filter(est_id=estacion).filter(hai_periodo=periodo))
        elif variable.var_id == 6:
            informacion = list(Var6Anuarios.objects.filter(est_id=estacion).filter(hsu_periodo=periodo))
        elif variable.var_id == 8:
            informacion = list(Var8Anuarios.objects.filter(est_id=estacion).filter(pat_periodo=periodo))
        elif variable.var_id == 9:
            informacion = list(Var9Anuarios.objects.filter(est_id=estacion).filter(tag_periodo=periodo))
        elif variable.var_id == 10:
            informacion = list(Var10Anuarios.objects.filter(est_id=estacion).filter(cau_periodo=periodo))
        elif variable.var_id == 11:
            informacion = list(Var11Anuarios.objects.filter(est_id=estacion).filter(nag_periodo=periodo))
        return informacion

    @staticmethod
    def datos_historicos(estacion, variable, periodo):
        modelo = 'Var' + str(variable.var_id) + 'Anuarios'
        modelo = globals()[modelo]
        consulta = modelo.objects.filter(est_id=estacion)
        mes = str(variable.var_codigo).lower() + "_mes"
        promedio = str(variable.var_codigo).lower() + "_promedio"
        if variable.var_id == 2:
            consulta = consulta.exclude(tai_periodo=periodo).values(mes)
        elif variable.var_id == 3:
            consulta = consulta.exclude(hai_periodo=periodo).values(mes)
        elif variable.var_id == 6:
            consulta = consulta.exclude(hsu_periodo=periodo).values(mes)
        elif variable.var_id == 8:
            consulta = consulta.exclude(pat_periodo=periodo).values(mes)
        elif variable.var_id == 9:
            consulta = consulta.exclude(tag_periodo=periodo).values(mes)
        elif variable.var_id == 10:
            consulta = consulta.exclude(cau_periodo=periodo).values(mes)
        elif variable.var_id == 11:
            consulta = consulta.exclude(nag_periodo=periodo).values(mes)

        informacion = list(
            consulta.annotate(valor=Avg(promedio)).order_by(mes)
        )

        datos = []
        for item in informacion:
            datos.append(item['valor'])
        return datos

    def set_encabezado_excel(self, ws, estacion, periodo):
        fila = 1
        col = 1
        col_fin = 11

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col+2)
        title = ws.cell(row=fila, column=col)
        title.value = estacion.est_codigo
        self.set_style(cell=title, font='font_bold_11', alignment='center',
                       border='border_thin', fill='light_salmon')

        ws.merge_cells(start_row=fila, start_column=col+3, end_row=fila, end_column=col+7)
        title = ws.cell(row=fila, column=col+3)
        title.value = estacion.est_nombre
        self.set_style(cell=title, font='font_bold_11', alignment='center',
                       border='border_thin', fill='light_salmon')

        ws.merge_cells(start_row=fila, start_column=col+8, end_row=fila, end_column=col_fin)
        title = ws.cell(row=fila, column=col+8)
        title.value = periodo
        self.set_style(cell=title, font='font_bold_11', alignment='center',
                       border='border_thin', fill='light_salmon')

        fila += 1

        ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col_fin)
        title = ws.cell(row=fila, column=col)
        title.value = 'Coordenadas Geográficas'
        self.set_style(cell=title, font='font_bold_11', alignment='center',
                       border='border_thin')

        fila += 1

        cell = ws.cell(row=fila, column=col)
        cell.value = "Latitud"
        self.set_style(cell=cell, font='font_bold_10', alignment='center',
                       border='border_thin')

        ws.merge_cells(start_row=fila, start_column=col + 1, end_row=fila, end_column=col + 3)
        title = ws.cell(row=fila, column=col + 1)
        title.value = estacion.est_latitud
        self.set_style(cell=title, font='font_10', alignment='center',
                       border='border_thin')

        cell = ws.cell(row=fila, column=col+4)
        cell.value = "Longitud"
        self.set_style(cell=cell, font='font_bold_10', alignment='center',
                       border='border_thin')

        ws.merge_cells(start_row=fila, start_column=col + 5, end_row=fila, end_column=col + 7)
        title = ws.cell(row=fila, column=col + 5)
        title.value = estacion.est_longitud
        self.set_style(cell=title, font='font_10', alignment='center',
                       border='border_thin')

        cell = ws.cell(row=fila, column=col + 8)
        cell.value = "Altura"
        self.set_style(cell=cell, font='font_bold_10', alignment='center',
                       border='border_thin')

        ws.merge_cells(start_row=fila, start_column=col + 9, end_row=fila, end_column=col_fin)
        title = ws.cell(row=fila, column=col + 9)
        title.value = estacion.est_altura
        self.set_style(cell=title, font='font_10', alignment='center',
                       border='border_thin')

    def bordes_celdas(self, ws, start_row, start_column, end_row, end_column):
        for fil in range(start_row, end_row+1):
            for col in range(start_column, end_column + 1):
                cell = ws.cell(row=fil, column=col)
                self.set_style(cell=cell, border='border_thin')





    @staticmethod
    def get_mes_anio(int_mes):
        # meses = ['Enero', 'Febrero','Marzo', 'Abril', 'Mayo', 'Junio', 'Julio','Agosto',
        # 'Septiembre','Octubre','Noviembre','Diciembre']
        meses = ['ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEP', 'OCT',
                 'NOV', 'DIC', 'ANUAL']

        return meses[int_mes-1]

    def set_style(self, cell, font=None, alignment=None, border=None, fill = None):
        if font is not None:
            cell.font = self.fonts[font]
        if alignment is not None:
            cell.alignment = self.alignments[alignment]
        if border is not None:
            cell.border = self.borders[border]
        if fill is not None:
            cell.fill = self.colors[fill]
